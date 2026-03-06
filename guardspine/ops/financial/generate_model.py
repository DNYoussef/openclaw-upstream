"""
GuardSpine Market Analysis Generator
Produces: Excel workbook (.xlsx), PDF report (.pdf), chart figures (.png)

All constants are at the top for easy tuning.
Excel formulas are preserved as live formulas, not computed values.

Thesis: GuardSpine is a multi-artifact governance spine for the AI office.
Four lanes (CodeGuard, PDFGuard, SheetGuard, ImageGuard) + PII-Shield +
Proprioceptive AI cognitive attestation. L0-L4 are risk tiers driving
governance behavior, not just cost tiers.
"""

import math
import os
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ============================================================
# CONSTANTS (tune these to update the entire model)
# ============================================================

# --- TAM Data ---
# Code-only markets
DEVSECOPS_TAM = 11.72e9
AI_GOVERNANCE_TAM = 0.44e9
CODE_REVIEW_TAM = 3.0e9
CODE_ONLY_TAM = DEVSECOPS_TAM + AI_GOVERNANCE_TAM + CODE_REVIEW_TAM

# Multi-artifact expansion
GRC_TAM = 23.32e9
DOC_MGMT_TAM = 7.0e9
DIGITAL_ASSET_TAM = 5.8e9
MULTI_ARTIFACT_TAM = CODE_ONLY_TAM + GRC_TAM + DOC_MGMT_TAM + DIGITAL_ASSET_TAM
GOVERNANCE_LAYER_PCT_LOW = 0.05
GOVERNANCE_LAYER_PCT_HIGH = 0.15

# Bottom-up
TOTAL_DEVS = 47.2e6
PRS_PER_DEV_MONTH = 20
GOVERNANCE_PRICE_PER_PR = 0.25
KNOWLEDGE_WORKERS = 350e6
CHANGES_PER_WORKER_MONTH = 200
BLENDED_PRICE_PER_CHANGE = 0.12

# --- SAM ---
SAM_CODE_ORGS = 15000
SAM_CODE_ACV = 25000
SAM_DOC_GOV_ORGS = 5000
SAM_DOC_GOV_ACV = 50000
SAM_CROSS_ORGS = 3000
SAM_CROSS_ACV = 75000
SAM_TOTAL = (SAM_CODE_ORGS * SAM_CODE_ACV +
             SAM_DOC_GOV_ORGS * SAM_DOC_GOV_ACV +
             SAM_CROSS_ORGS * SAM_CROSS_ACV)

# --- SOM ---
SOM_Y1_PCT = 0.01
SOM_Y2_PCT = 0.04
SOM_Y3_PCT = 0.10

# --- L0-L4 Risk Tiers (governance behavior, not just cost) ---
# L0: Auto-pass (metadata logged, no review)
# L1: Light review (AI summary, single reviewer)
# L2: Standard review (multi-model consensus, rubric eval, evidence bundle)
# L3: Elevated review (role-based approvers, stop-the-line)
# L4: Full audit (cross-functional, adversarial, cognitive attestation)
TIER_DIST = {"L0": 0.55, "L1": 0.25, "L2": 0.12, "L3": 0.06, "L4": 0.02}
TIER_COGS = {"L0": 0.00, "L1": 0.03, "L2": 0.15, "L3": 0.50, "L4": 1.50}
TIER_REV  = {"L0": 0.00, "L1": 0.15, "L2": 0.60, "L3": 1.50, "L4": 3.00}
TIER_NAMES = {
    "L0": "Auto-pass",
    "L1": "Light review",
    "L2": "Standard review",
    "L3": "Elevated review",
    "L4": "Full audit + cognitive attestation",
}

# --- Open-Core Model (Linux vs Red Hat) ---
# FREE: Spec, verifier, CodeGuard Action (MIT), PII-Shield, YAML rubric format
# PAID: UI, management coordination, enterprise features, cognitive attestation
#
# Analogues:
#   Linux/Red Hat:     kernel free, RHEL $3.4B/yr (IBM acquired for $34B)
#   GitLab CE/EE:      CE free, EE $580M/yr ARR
#   Docker CE/Desktop: engine free, Desktop $200M/yr
#   WordPress/VIP:     WP free, Automattic $700M+/yr

# --- Open-Core Funnel ---
# Free users self-host, $0 revenue, $0 COGS (community support only)
# Conversion rates (industry benchmarks):
FREE_TO_PRO_CONVERSION = 0.03       # 3% (typical open-core: 1-5%)
PRO_TO_BUSINESS_EXPANSION = 0.20    # 20% within 12 months
BUSINESS_TO_ENTERPRISE = 0.10       # 10% within 18 months

# --- Per-Customer Unit Economics (4 tiers) ---
TIERS = {
    # BYOK: Users bring their own API keys. GuardSpine pays NO LLM API costs.
    # COGS = hosting/infra + support + cognitive attestation licensing only.
    "Community (Free)": {
        "monthly": 0, "changes_mo": 500,
        "hosting": 0, "support": 0, "cognitive_license": 0,
        "nrr": 1.00, "cac": 0,
    },
    "Pro (Code UI)": {
        "monthly": 2000, "changes_mo": 800,
        "hosting": 150, "support": 100, "cognitive_license": 0,
        "nrr": 1.15, "cac": 5000,
    },
    "Business (Multi-Lane)": {
        "monthly": 5000, "changes_mo": 2500,
        "hosting": 300, "support": 200, "cognitive_license": 0,
        "nrr": 1.20, "cac": 15000,
    },
    "Enterprise (Full)": {
        "monthly": 12000, "changes_mo": 8000,
        "hosting": 500, "support": 400, "cognitive_license": 150,
        "nrr": 1.30, "cac": 50000,
    },
}
CUSTOMER_LIFETIME_MONTHS = 36

# --- Revenue Scenarios (paid customers only; free users are the funnel) ---
SCENARIOS = {
    "Bear":  {"c0": 5,  "growth": 0.08, "avg_acv": 18000, "free_base": 500},
    "Base":  {"c0": 8,  "growth": 0.15, "avg_acv": 30000, "free_base": 2000},
    "Bull":  {"c0": 15, "growth": 0.20, "avg_acv": 50000, "free_base": 5000},
}

# --- Unicorn Path (with growth decay, NRR, enterprise catalysts) ---
# Model: ARR(Y) = ARR(Y-1) * NRR + new_logos(Y) * ACV + enterprise_deals(Y)
# Growth rate decays annually. Valuation = ARR * multiple(ARR, YoY growth).
# Enterprise catalysts: Netflix (pain-driven pilot March 2026) + IBM (Triangle Strategy Q3 2026)
UNICORN_SCENARIOS = {
    "Bear": {
        "c0": 5, "g0": 0.08, "acv": 18000, "nrr": 1.10, "decay": 0.15,
        "enterprise": {1: 50000, 2: 80000, 3: 100000, 4: 120000, 5: 150000,
                       6: 200000, 7: 250000},
    },
    "Base": {
        "c0": 8, "g0": 0.15, "acv": 30000, "nrr": 1.20, "decay": 0.20,
        # Netflix $200K Y1 + IBM $150K Y1, expanding each year + new enterprise deals
        "enterprise": {1: 350000, 2: 1100000, 3: 1850000, 4: 3300000, 5: 5500000,
                       6: 8000000, 7: 10000000},
    },
    "Bull": {
        "c0": 15, "g0": 0.20, "acv": 50000, "nrr": 1.30, "decay": 0.15,
        # Netflix $500K Y1 + IBM $500K Y1, rapid enterprise expansion
        "enterprise": {1: 1000000, 2: 3300000, 3: 8500000, 4: 18200000, 5: 34000000},
    },
}
# Valuation multiples: base_mult * growth_premium * margin_premium * category_premium
# Base by ARR bracket: <$5M=20x, $5-20M=15x, $20-50M=12x, $50-100M=10x, $100M+=8x
# Growth: >100% YoY = 1.5x, >50% = 1.25x, else 1.0x
# Margin: 1.15x (87-91% vs 75% SaaS baseline)
# Category: 1.10x (category creator, no direct competitor)

# --- Probability Risk Model (pre-mortem analysis, calibrated from team composition) ---
# Each risk factor has a failure probability. Cumulative unicorn probability is stated
# directly (not a pure product of survival rates -- accounts for correlation and
# non-modeled risks). Source: pre-mortem analysis with Igor + Chris + validation gates.
#
# "triangle" column: Post Pre-Mortem + Triangle Strategy (Logan/Ishwar/Jacob -> IBM).
# Triangle reduces 6 of 8 factors:
#   GTM: Ishwar at IBM = warm internal champion, not cold outbound
#   PMF: Z-Inspection = independent validation; IBM pilot = enterprise PMF signal
#   Moat: Logan exclusive patents (MOU s5) + G7 standards reference = regulatory lock-in
#   Capital: IBM pilot + Z-Inspection report + G7 ref = strongest seed narrative
#   Scaling: IBM logo unlocks procurement at peer enterprises
#   Legal: Z-Inspection report = due diligence cover
RISK_FACTORS = {
    #                       solo   igor  igor_chris  pre_mortem  triangle  current (Feb 2026)
    # "current" = Triangle + Kristen GTM + $1M angel + code 85%+ done
    "Technical Execution":  {"solo": 0.35, "igor": 0.15, "igor_chris": 0.15, "pre_mortem": 0.09, "triangle": 0.08, "current": 0.05},
    "Product-Market Fit":   {"solo": 0.40, "igor": 0.35, "igor_chris": 0.23, "pre_mortem": 0.14, "triangle": 0.09, "current": 0.07},
    "GTM / Sales":          {"solo": 0.50, "igor": 0.45, "igor_chris": 0.23, "pre_mortem": 0.20, "triangle": 0.12, "current": 0.08},
    "Scaling":              {"solo": 0.40, "igor": 0.35, "igor_chris": 0.23, "pre_mortem": 0.18, "triangle": 0.15, "current": 0.12},
    "Competitive Moat":     {"solo": 0.30, "igor": 0.25, "igor_chris": 0.15, "pre_mortem": 0.11, "triangle": 0.06, "current": 0.05},
    "Capital Access":       {"solo": 0.35, "igor": 0.30, "igor_chris": 0.18, "pre_mortem": 0.09, "triangle": 0.05, "current": 0.03},
    "Founder Dynamics":     {"solo": 0.00, "igor": 0.00, "igor_chris": 0.00, "pre_mortem": 0.18, "triangle": 0.15, "current": 0.12},
    "Legal / Liability":    {"solo": 0.00, "igor": 0.00, "igor_chris": 0.00, "pre_mortem": 0.07, "triangle": 0.05, "current": 0.04},
}
# Stated cumulative unicorn probabilities (accounts for correlation + unmodeled risks):
# "current" = v3 code audit shows TEAM tier 1-2 weeks, ORG 2-3 weeks,
# Kristen adds GTM/product-to-market experience, $1M angel via her network,
# Ishwar warm (IBM/Z-Inspection), Jacob warm (G7/NIST/Platform One)
UNICORN_PROBABILITIES = {
    "Solo (David)":           0.051,
    "+ Igor (CTO)":           0.08,
    "+ Igor + Chris (CCO)":   0.175,
    "Post Pre-Mortem":        0.21,
    "+ Triangle Strategy":    0.27,
    "Current (Feb 2026)":     0.34,
}

# --- Triangle Strategy Gates (all must pass before joint call) ---
TRIANGLE = {
    "ibm_acv_range": (300000, 1000000),  # $300K-$1M ACV
    "ibm_pilot_month": 7,                # Q3 2026 = ~month 7 from Feb 2026
    "gates": [
        "Logan signs NDA (unblocks technical discussion)",              # DONE
        "Working integration prototype exists",                         # DONE (codeguard-action v1.0.1)
        "Ishwar agrees to Z-Inspection assessment",                     # WARM (Feb 13 meeting positive)
        "Jacob shows interest in cognitive attestation angle",          # WARM (SBOM angle, Platform One)
    ],
    "gate_status": {
        "Logan NDA": "DONE",
        "Prototype": "DONE",
        "Ishwar Z-Inspection": "WARM - comparing vs Z-Inspection methodology",
        "Jacob G7/NIST": "WARM - SBOM complementarity, Canada contact pending",
    },
    "timeline": {
        "mou_signed": "Feb 2026",
        "prototype": "Feb 2026 (shipped)",
        "z_inspection": "March-April 2026",
        "g7_reference": "April-May 2026",
        "ibm_pilot": "Q3 2026",
    },
    "flywheel_months": 3,  # Compressed: 2/4 gates already done
}

# --- Founding Team (updated Feb 2026) ---
# David: vision, architecture, IP, 1-on-1 enterprise sales
# Igor: technical depth, 13yr commercial eng, Rust, crypto, physics MSc
# Kristen: GTM, product-to-market experience, angel investor network
# Chris Hood: advisor (ex-Google 7yr, Noematic AI inventor, USPTO, Google connects)
# Ilya (contributor): PII-Shield WASM integration, open-source contributor
TEAM = {
    "David":   {"equity_pct": 40, "role": "CEO / Architect / IP / Enterprise Sales"},
    "Igor":    {"equity_pct": 40, "role": "CTO (13yr commercial eng, Rust, crypto, physics MSc)"},
    "Kristen": {"equity_pct": 2,  "role": "GTM Advisor (2% equity, no cliff, no cash -- agreed Mar 2)"},
    "Angel":   {"equity_pct": 10, "role": "Pre-seed investor ($1M)"},
    "Pool":    {"equity_pct": 8,  "role": "Future hires (staircase model: ~1.5% per hire)"},
}
DAVID_EQUITY = 0.40  # post-raise; 44.4% pre-raise

# --- Capital Strategy (updated Feb 2026) ---
# PRIMARY: Bootstrap + angel round. BYOK model = 97-99% gross margins.
# Core team: David + Igor + AI = equivalent to 6-8 person team.
# Monthly burn $26.5K (Kristen = equity-only advisor, no AI budget). Cash-flow positive much sooner.
# $1M angel round via Kristen's investor network for 3yr runway.
#
# WHY THIS PATH:
#   - BYOK: zero API COGS. Users bring their own keys.
#   - David + Igor (technical) + Kristen (GTM) + AI = full founding team.
#   - Open-core GitHub Action markets itself. No paid acquisition needed early.
#   - Warm enterprise intros (IBM via Ishwar, gov via Jacob). No sales team yet.
#   - 87-91% margins mean revenue self-funds growth after first 2-3 customers.
#   - $1M angel = 2yr runway, 3 hires (engineer, lawyer, finance person).
#   - Every avoided VC round saves $35-74M in dilution at exit.

# Bootstrap path (PRIMARY): $1M angel for 2yr runway
BOOTSTRAP_PATH = [
    {"round": "Angel", "month": 2, "raise": 1.0e6, "valuation": 15.0e6, "dilution": 0.0625},
]
BOOTSTRAP_BURN_MONTHLY = 40000  # $40K/month (3 founders + infra + legal)
BOOTSTRAP_RUNWAY_MONTHS = 1.0e6 / 40000  # 25 months on angel alone

# VC path (COMPARISON ONLY -- not the plan):
VC_PATH = [
    {"round": "Pre-seed", "month": 7,  "raise": 1.0e6,  "valuation": 8.0e6,  "dilution": 0.12},
    {"round": "Seed",     "month": 14, "raise": 5.0e6,  "valuation": 15.0e6, "dilution": 0.25},
    {"round": "Series A", "month": 23, "raise": 20.0e6, "valuation": 75.0e6, "dilution": 0.20},
    {"round": "Series B", "month": 32, "raise": 50.0e6, "valuation": 250.0e6, "dilution": 0.20},
]

# Default to bootstrap
FUNDRAISING = BOOTSTRAP_PATH

# $316M exit = $100M post-tax at 42.19% ownership (angel only).
# $450M exit = $100M post-tax at 29.7% ownership (VC 2-round).
# $702M exit = $100M post-tax at 19.0% ownership (VC 4-round).
EXIT_VALUATION = 400e6  # Conservative target
EXIT_MONTH_RANGE = (20, 30)  # Faster: product 85%+ built, team complete
TAX_RATE = 0.25  # federal 20% + state ~5%

# $100M personal target math:
# $100M = Exit * David% * (1 - tax)
# At 42.19% ($1M angel at $15M pre): Exit = $100M / (0.4219 * 0.75) = $316M
# At 45.00% (zero raise): Exit = $100M / (0.45 * 0.75) = $296M
PERSONAL_TARGET = 100e6
BOOTSTRAP_OWNERSHIP = DAVID_EQUITY * (1 - BOOTSTRAP_PATH[0]["dilution"])  # ~42.19%
BOOTSTRAP_EXIT_NEEDED = PERSONAL_TARGET / (BOOTSTRAP_OWNERSHIP * (1 - TAX_RATE))

# --- Revenue Milestones (v3 audit, Feb 2026: TEAM tier 1-2 weeks, ORG 2-3 weeks) ---
# Faster than original: product is 85%+ built, Kristen accelerates GTM
REVENUE_MILESTONES = [
    ("First customer",  3.0,  0.88),
    ("$500K ARR",       7.0,  0.72),
    ("$1M ARR",         10.0, 0.62),
    ("$3M ARR",         16.0, 0.48),
    ("$10M ARR",        22.0, 0.33),
    ("$30M ARR",        30.0, 0.24),
    ("$100M ARR",       42.0, 0.14),
]

# --- Netflix Kill Criteria ---
NETFLIX = {
    "engineers": 116,
    "review_velocity_gap": 4.0,       # 4x faster than review capacity
    "false_positive_max": 0.05,       # <5%
    "false_negative_max": 0.02,       # <2%
    "decision_reduction": (1000, 10), # from 1000 to 10
    "pilot_weeks": 4,
    "acv_range": (200000, 2000000),   # $200K-$2M (saves one senior hire: $180-250K)
}

# --- Personal Liquidity Thresholds (Bootstrap: David at ~42.19% post-angel) ---
# After angel (6.25% dilution): 0.45 * 0.9375 = 0.4219
# Need ~$316M company valuation for $100M after tax at ~42.19% ownership
# Compare VC 4-round: 19% ownership, needs $702M for same $100M.
LIQUIDITY_THRESHOLDS = [
    ("$100K",  (6, 10),   0.80),
    ("$1M",    (10, 16),  0.60),
    ("$30M",   (20, 28),  0.35),
    ("$100M",  (24, 36),  0.27),
    ("$200M+", (30, 42),  0.15),
]

# --- Acquirer Analysis (strategic fit scoring) ---
# Dimensions scored 1-10. Weighted composite for ranking.
ACQUIRER_SCORES = {
    "Microsoft / GitHub": {
        "strategic_gap": 10, "distribution": 10, "precedent_acq": 9,
        "integration_ease": 9, "price_range_m": (500, 2000),
        "why": "Governance layer for Copilot + GitHub. 9/9 MECE fills biggest gap.",
    },
    "IBM": {
        "strategic_gap": 9, "distribution": 8, "precedent_acq": 10,
        "integration_ease": 8, "price_range_m": (300, 1000),
        "why": "Red Hat precedent ($34B open-core). Ishwar = internal champion. watsonx governance.",
    },
    "Palo Alto Networks": {
        "strategic_gap": 8, "distribution": 7, "precedent_acq": 10,
        "integration_ease": 7, "price_range_m": (200, 800),
        "why": "Acquired Bridgecrew + Cider Security. AI artifact governance extends code-to-cloud.",
    },
    "ServiceNow": {
        "strategic_gap": 8, "distribution": 8, "precedent_acq": 7,
        "integration_ease": 8, "price_range_m": (300, 1000),
        "why": "GRC platform needs artifact-level governance. Platform-of-platforms strategy.",
    },
    "CrowdStrike": {
        "strategic_gap": 6, "distribution": 7, "precedent_acq": 7,
        "integration_ease": 5, "price_range_m": (150, 500),
        "why": "AI security expansion. Weaker fit but possible if they enter GRC.",
    },
}
ACQUIRER_WEIGHTS = {"strategic_gap": 3, "distribution": 2, "precedent_acq": 2, "integration_ease": 1}

# Pre-seed valuation range (driven by narrative + signals, not revenue)
PRESEED_VALUATION = {
    "floor": 8e6,     # thesis + team (David/Igor/Kristen)
    "mid": 15e6,      # + working product (85%+ built) + demand signals
    "target": 20e6,   # + IBM warm + Z-Inspection + G7 interest
    "stretch": 30e6,  # + signed pilot + Z-Inspection report + standards reference
}

# --- Acquisition Price Model ---
# Strategic premium: acquirers pay above implied market valuation for category-defining assets.
# Benchmarks: GitHub at 37.5x rev ($7.5B/$200M ARR), Red Hat at 10x ($34B/$3.4B ARR),
# Bridgecrew at ~100x ($200M/~$2M ARR -- early stage, strategic).
STRATEGIC_PREMIUM = {
    "pre_revenue": 1.50,  # pre-revenue: pure strategic, 50% above market
    "early": 1.40,        # <$10M ARR: proven product, distribution value
    "growth": 1.30,       # $10-50M ARR: category leader momentum
    "scale": 1.20,        # $50M+ ARR: at scale, premium compresses
}

# What must be true for each acquirer to pay $1B+
BILLION_REQUIREMENTS = {
    "min_arr_floor": 30e6,          # $30M ARR absolute minimum
    "min_arr_strong": 50e6,         # $50M ARR for confident $1B+
    "min_yoy_growth": 0.50,         # 50%+ YoY
    "min_gross_margin": 0.85,       # 85%+
    "min_enterprise_logos": 5,      # 5+ enterprise reference customers
    "min_open_source_installs": 10000,  # 10K+ free GitHub Action installs
    "strategic_signals": [
        "9/9 MECE dimensions (nearest competitor: 3/9)",
        "Exclusive patent portfolio (MOU s5, cognitive attestation)",
        "G7/NIST standards reference (regulatory lock-in)",
        "Z-Inspection validation report (third-party credibility)",
        "IBM + Netflix logos (enterprise social proof)",
    ],
    "per_acquirer": {
        "Microsoft / GitHub": {
            "trigger_arr": 20e6,
            "likely_price_at_trigger": "800M-1.5B",
            "rationale": "GitHub Action adoption > 10K installs. Copilot Trust layer. Missing link in dev platform.",
        },
        "IBM": {
            "trigger_arr": 15e6,
            "likely_price_at_trigger": "500M-1.2B",
            "rationale": "Ishwar internal champion. watsonx governance gap. Red Hat playbook (open-core -> enterprise).",
        },
        "ServiceNow": {
            "trigger_arr": 25e6,
            "likely_price_at_trigger": "600M-1.2B",
            "rationale": "GRC platform-of-platforms. Artifact governance fills gap. Enterprise distribution.",
        },
        "Palo Alto Networks": {
            "trigger_arr": 30e6,
            "likely_price_at_trigger": "500M-1.0B",
            "rationale": "Code-to-cloud security. Bridgecrew + Cider precedent. AI artifact extension.",
        },
    },
}

# Target exit for the $1B+ analysis
TARGET_ACQUISITION = 1.0e9

# --- Fee Cascade ---
# k values calibrated from KPMG: 14% reduction at t=12 with logistic adoption curve.
# sum(A(s), s=0..12) = 2.0186, so k_audit = -ln(0.86)/2.0186 = 0.0747
# Other sectors scaled proportionally to preserve relative ordering.
CASCADE_SECTORS = {
    "Audit":          {"k": 0.0747, "lag": 0},
    "Legal":          {"k": 0.0534, "lag": 6},
    "Consulting":     {"k": 0.0427, "lag": 12},
    "Implementation": {"k": 0.0320, "lag": 18},
    "Design":         {"k": 0.0213, "lag": 24},
}
ADOPTION_RATE = 0.15
INFLECTION_MONTH = 18

# --- Governance Premium ---
G0 = 0.02
GMAX = 0.20
LAMBDA_GOV = 0.08

# --- Competitive Scoring (9-dimension MECE + 5-dimension moat) ---
MECE_MATRIX = {
    "GuardSpine":       {"code": 1, "docs": 1, "sheets": 1, "images": 1, "ai_prov": 1, "risk_gate": 1, "evidence": 1, "diffs": 1, "stop_line": 1},
    "Vanta/Secureframe": {"code": 0, "docs": 0, "sheets": 0, "images": 0, "ai_prov": 0, "risk_gate": 0.5, "evidence": 0.5, "diffs": 0, "stop_line": 0},
    "GitHub/GitLab":    {"code": 1, "docs": 0, "sheets": 0, "images": 0, "ai_prov": 0, "risk_gate": 0.5, "evidence": 0, "diffs": 1, "stop_line": 0.5},
    "Codebat":          {"code": 0.5, "docs": 0, "sheets": 0, "images": 0, "ai_prov": 0, "risk_gate": 0, "evidence": 0.5, "diffs": 0, "stop_line": 0},
    "SonarQube":        {"code": 1, "docs": 0, "sheets": 0, "images": 0, "ai_prov": 0, "risk_gate": 0.5, "evidence": 0, "diffs": 1, "stop_line": 0.5},
    "AI Observability": {"code": 0, "docs": 0, "sheets": 0, "images": 0, "ai_prov": 1, "risk_gate": 0, "evidence": 0, "diffs": 0, "stop_line": 0},
    "Manual Review":    {"code": 0.5, "docs": 0.5, "sheets": 0.5, "images": 0.5, "ai_prov": 0, "risk_gate": 0.5, "evidence": 0, "diffs": 0, "stop_line": 0.5},
}

MOAT_SCORES = {
    "GuardSpine":    {"data_moat": 9, "network": 7, "regulatory": 9, "tech_diff": 9, "switching": 8},
    "GitHub":        {"data_moat": 7, "network": 9, "regulatory": 4, "tech_diff": 5, "switching": 6},
    "SonarQube":     {"data_moat": 5, "network": 6, "regulatory": 3, "tech_diff": 6, "switching": 4},
    "Codebat":       {"data_moat": 6, "network": 3, "regulatory": 7, "tech_diff": 8, "switching": 5},
    "Manual Review": {"data_moat": 2, "network": 2, "regulatory": 6, "tech_diff": 2, "switching": 3},
}
MOAT_WEIGHTS = {"data_moat": 3, "network": 1, "regulatory": 2, "tech_diff": 1, "switching": 1}

# --- Speed-to-$100M Execution Timeline ---
# Every action sequenced for maximum speed. Month 0 = angel closed.
# Goal: $15-20M ARR -> $305M exit -> $100M post-tax in 18-24 months.
SPEED_TIMELINE = [
    {"phase": "Parallel Launch",   "months": (0, 2),  "actions": [
        "Close $1M angel via Kristen's investor network",
        "Ship TEAM tier (1-2 weeks remaining): full-text search, drift alerts",
        "File provisional patent ($2K, cognitive attestation + evidence bundle)",
        "CodeGuard Action already live on GitHub Marketplace",
    ], "arr": 0, "controlled": True},
    {"phase": "First Revenue",     "months": (1, 3),  "actions": [
        "Approach Ishwar at IBM (Z-Inspection evaluation, warm from Feb 13 meeting)",
        "Jacob intro to Canada gov/enterprise contact (Platform One angle)",
        "Inbound from GitHub Action installs -> 2-3 mid-market prospects",
        "Ship ORG tier (2-3 weeks): board packets, policy simulation",
    ], "arr": 0, "controlled": True},
    {"phase": "Anchor Customers",  "months": (2, 5),  "actions": [
        "Close 2-3 enterprise pilots at $200-500K ACV (IBM, gov, fintech)",
        "Close 2-3 mid-market at $100-200K ACV (regulated: healthtech/govtech)",
        "Publish case study (decision reduction 1000->10)",
    ], "arr": 1.5e6, "controlled": True},
    {"phase": "Revenue Ramp",      "months": (4, 9), "actions": [
        "Hit $1-5M ARR (IBM pilot + gov deals + mid-market)",
        "Ship ENTERPRISE tier: SSO federation, custom SLAs, dedicated support",
        "G7/NIST reference through Jacob (Triangle Strategy)",
        "Cash-flow positive at 87-91% margins and $40K/mo burn",
    ], "arr": 5e6, "controlled": True},
    {"phase": "Acquisition Zone",  "months": (9, 15), "actions": [
        "Hit $5-12M ARR, 200%+ YoY growth",
        "5+ enterprise logos, 5K+ GitHub Action installs",
        "Start acquirer conversations (warm intros via Chris/Kristen/angel network)",
        "Run Microsoft + IBM + ServiceNow in parallel (competition = speed + price)",
    ], "arr": 12e6, "controlled": True},
    {"phase": "Close the Exit",    "months": (15, 22), "actions": [
        "Hit $12-20M ARR (trigger zone for IBM $15M, MSFT $20M)",
        "20-35x revenue = $240-700M acquisition price",
        "Close at $316M+ -> David: $100M+ post-tax",
    ], "arr": 20e6, "controlled": False},
]

# The 5 things that add 6-12 months each if you let them happen.
SPEED_KILLERS = [
    {"name": "Underpricing",
     "description": "$50K ACV when you should charge $500K",
     "cost_months": 6,
     "fix": "Anchor on Netflix ACV. Never discount >10%. First customer sets the floor."},
    {"name": "Raising VC",
     "description": "$5M seed = 3-4 months founder time + 25% dilution",
     "cost_months": 4,
     "fix": "Bootstrap. $1M angel only. Revenue funds growth at 87-91% margins."},
    {"name": "Building before validating",
     "description": "PDFGuard before anyone asks = wasted months",
     "cost_months": 3,
     "fix": "Only build what a paying customer needs. Phase 0 gates everything."},
    {"name": "Small customer addiction",
     "description": "$10K/yr customer takes same effort as $500K/yr",
     "cost_months": 6,
     "fix": "Enterprise only. SMB comes free from open source. Min ACV $100K."},
    {"name": "Waiting for perfection",
     "description": "Delaying pilot until FP <1% when <5% is sufficient",
     "cost_months": 3,
     "fix": "Ship at FP <5%. Pilot IS the product-market fit test."},
]

SPEED_BEST_CASE_MONTHS = 15
SPEED_REALISTIC_MONTHS = 22

# --- AI Trajectory Adjustment (sourced data, Feb 2026) ---
# As AI competence and execution length grow, 5 effects change the math.
# All data points sourced from METR, Epoch AI, EU official gazette, industry reports.

AI_TRAJECTORY = {
    # Source: METR (Epoch AI affiliate), Feb 2026
    # Task completion horizon doubles every ~7 months.
    # Current: 50-min tasks. Projection: week-long tasks by late 2026-2027.
    "metr_doubling_months": 7,
    "current_task_horizon_min": 50,
    "projected_weeklong_year": 2027,

    # Source: Epoch AI index, token cost per million output tokens
    # 10x decline per year: $20 (2022) -> $2 (2023) -> $0.40 (2025) -> ~$0.04 (2026)
    "token_cost_decline_per_year": 10,
    "cost_per_m_tokens_2025": 0.40,
    "cost_per_m_tokens_2026e": 0.04,

    # Source: SWE-bench Verified (Feb 2026), SWE-bench Pro (Jan 2026)
    # Verified: ~75% solve rate (best models). Pro: 45.89% (Claude Opus 4.5).
    "swebench_verified_pct": 0.75,
    "swebench_pro_pct": 0.4589,

    # Source: GitHub (Jan 2026), Stack Overflow survey
    # 41% of code AI-generated. 20-30% real productivity gain.
    "ai_code_share": 0.41,
    "productivity_gain_low": 0.20,
    "productivity_gain_high": 0.30,

    # Source: Capgemini Research Institute (Feb 2026)
    # 79% of orgs adopted agentic AI, 340% YoY surge, 40% of apps by EOY 2026.
    "enterprise_adoption_pct": 0.79,
    "enterprise_agentic_surge": 3.40,
    "apps_agentic_by_eoy2026": 0.40,

    # Source: EU AI Act, Official Journal of the EU
    # Fully enforceable August 2, 2026. Fines up to 7% of global turnover.
    "eu_ai_act_date": "2026-08-02",
    "eu_ai_act_fine_pct": 0.07,
}

# --- 5 Effects on GuardSpine Math ---
AI_EFFECTS = [
    {
        "effect": "TAM Expands (artifact volume 2-5x)",
        "baseline": "TAM CAGR 35-40%",
        "adjusted": "TAM CAGR 50-70%",
        "mechanism": "More AI-generated artifacts = more governance surface. "
                     "41% of code already AI-generated. PDFs, spreadsheets, images next. "
                     "Each artifact needs evidence bundles. Volume scales linearly with generation.",
        "tam_cagr_baseline": 0.38,
        "tam_cagr_adjusted": 0.60,
    },
    {
        "effect": "Build Costs Collapse (lane delivery 3-4 weeks, not 2-3 months)",
        "baseline": "Lane delivery: 2-3 months",
        "adjusted": "Lane delivery: 3-4 weeks",
        "mechanism": "SWE-bench 75% solve rate + week-long agent tasks = "
                     "AI handles 60-80% of implementation. 2 founders + AI = 5-person team -> 8-person team. "
                     "Saves 2-3 months on 4-lane roadmap.",
        "timeline_savings_months": 3,
    },
    {
        "effect": "Competition Easier to Enter, Harder to Win",
        "baseline": "12-18 month head start",
        "adjusted": "6-9 month head start (but moat deepens)",
        "mechanism": "Anyone can build a governance tool faster. But evidence chain pedigree, "
                     "hash-linked bundles, and cognitive attestation are not code -- they are data structures "
                     "and trust networks. First mover with real enterprise data wins permanently.",
    },
    {
        "effect": "Customer Willingness to Pay Increases",
        "baseline": "ACV $200K (mid-market), $500K (enterprise)",
        "adjusted": "ACV $300-500K (mid-market), $500K-1M (enterprise)",
        "mechanism": "EU AI Act (Aug 2026, 7% turnover fines) creates compliance buyers. "
                     "79% of orgs adopted agentic AI but governance lags. "
                     "Budget shifts from 'nice to have' to 'regulatory mandate'.",
        "acv_multiplier": 1.5,
    },
    {
        "effect": "Acquirer Urgency Increases",
        "baseline": "Revenue multiple 15-25x",
        "adjusted": "Revenue multiple 20-35x",
        "mechanism": "EU AI Act deadline forces platform vendors to acquire governance capabilities "
                     "before Aug 2026. Compressed timeline = premium multiples. "
                     "Every major cloud vendor needs this; build-vs-buy tilts to buy.",
        "multiple_baseline_low": 15, "multiple_baseline_high": 25,
        "multiple_adjusted_low": 20, "multiple_adjusted_high": 35,
    },
]

# --- AI-Adjusted Speed Timeline ---
# Original: 18 best / 24 realistic.
# AI-adjusted: faster lane delivery (-3mo), higher ACV (fewer customers needed),
# higher multiples ($305M achievable at lower ARR).
AI_SPEED_BEST_CASE_MONTHS = 12
AI_SPEED_REALISTIC_MONTHS = 18

AI_SPEED_TIMELINE = [
    {"phase": "Parallel Launch",   "months": (0, 1),  "arr": 0,
     "note": "Compressed: TEAM tier ships in 1-2 weeks. Angel closing via Kristen."},
    {"phase": "First Revenue",     "months": (1, 3),  "arr": 0,
     "note": "Ishwar + Jacob already warm. AI accelerates pilot setup and demo prep."},
    {"phase": "Anchor Customers",  "months": (2, 4),  "arr": 2.0e6,
     "note": "Higher ACV ($300-500K) = $2M ARR from 4-5 customers."},
    {"phase": "Revenue Ramp",      "months": (3, 7),  "arr": 5e6,
     "note": "ORG tier already 85%+ built. Hit $5M ARR faster."},
    {"phase": "Acquisition Zone",  "months": (7, 12), "arr": 12e6,
     "note": "EU AI Act deadline (Aug 2026) compresses acquirer timelines."},
    {"phase": "Close the Exit",    "months": (12, 18), "arr": 12e6,
     "note": "$12M ARR x 25x = $300M+. $316M achievable at $9-12M ARR at 25-35x."},
]

# --- AI-Adjusted Revenue Scenarios ---
# Higher ACV, higher multiples, faster growth (AI productivity compounds).
AI_SCENARIOS = {
    "Bear (AI-adjusted)":  {"c0": 5,  "growth": 0.10, "avg_acv": 25000,
                             "arr_y1": 1.5e6,  "multiple_low": 20, "multiple_high": 30},
    "Base (AI-adjusted)":  {"c0": 8,  "growth": 0.18, "avg_acv": 45000,
                             "arr_y1": 5.0e6,  "multiple_low": 25, "multiple_high": 35},
    "Bull (AI-adjusted)":  {"c0": 15, "growth": 0.22, "avg_acv": 75000,
                             "arr_y1": 15.0e6, "multiple_low": 30, "multiple_high": 40},
}

# Key insight: $305M exit is achievable at $9-12M ARR (not $15-20M) at AI-adjusted multiples.
AI_EXIT_ARR_LOW = 9e6    # $9M ARR at 35x = $315M
AI_EXIT_ARR_HIGH = 12e6  # $12M ARR at 25x = $300M

# --- AI-Adjusted Unicorn Path ---
# Same model as UNICORN_SCENARIOS but with AI tailwind parameters.
# Changes: 1.5x ACV, higher NRR (compliance stickiness), faster growth,
# larger enterprise deals (EU AI Act urgency), mult_boost for higher multiples.
AI_UNICORN_SCENARIOS = {
    "Bear (AI)": {
        "c0": 5, "g0": 0.10, "acv": 27000, "nrr": 1.15, "decay": 0.12,
        "enterprise": {1: 75000, 2: 150000, 3: 250000, 4: 400000, 5: 600000,
                       6: 800000, 7: 1000000},
        "mult_boost": 1.30,
    },
    "Base (AI)": {
        "c0": 8, "g0": 0.18, "acv": 45000, "nrr": 1.25, "decay": 0.15,
        "enterprise": {1: 500000, 2: 1800000, 3: 3500000, 4: 6000000,
                       5: 9000000, 6: 12000000, 7: 15000000},
        "mult_boost": 1.35,
    },
    "Bull (AI)": {
        "c0": 15, "g0": 0.22, "acv": 75000, "nrr": 1.35, "decay": 0.12,
        "enterprise": {1: 1500000, 2: 5000000, 3: 13000000, 4: 28000000,
                       5: 50000000},
        "mult_boost": 1.40,
    },
}

# AI-adjusted unicorn probability
# Baseline: 27% (Triangle Strategy). AI adjusts 6 risk factors:
AI_RISK_ADJUSTMENTS = {
    "Technical Execution": -0.02,   # AI makes building easier (0.08 -> 0.06)
    "Product-Market Fit":  -0.03,   # EU AI Act = regulatory mandate (0.09 -> 0.06)
    "GTM / Sales":         -0.02,   # Higher urgency = faster close (0.12 -> 0.10)
    "Competitive Moat":    +0.03,   # Shorter head start (0.06 -> 0.09) -- ONLY NEGATIVE
    "Capital Access":      -0.01,   # Higher multiples (0.05 -> 0.04)
    "Scaling":             -0.02,   # AI handles more (0.15 -> 0.13)
}
def compute_ai_unicorn_probability(
    base_prob=UNICORN_PROBABILITIES["Current (Feb 2026)"],
    adjustments=None,
    leverage=2.0,
):
    """Compute AI-adjusted unicorn probability from risk model.

    Method: Sum the net risk adjustment across all factors. Negative values
    mean AI reduces risk. Apply a leverage multiplier (default 2x) to convert
    risk-percentage-point changes into probability-percentage-point changes.
    This accounts for the fact that reducing a bottleneck risk has outsized
    impact on overall success probability.

    Example with defaults:
      Net adjustment = -0.02 -0.03 -0.02 +0.03 -0.01 -0.02 = -0.07
      Probability lift = abs(-0.07) * 2.0 = 0.14
      AI probability = 0.27 + 0.14 = 0.41
    """
    if adjustments is None:
        adjustments = AI_RISK_ADJUSTMENTS
    net_risk_change = sum(adjustments.values())
    prob_lift = abs(net_risk_change) * leverage
    return min(base_prob + prob_lift, 1.0)

AI_UNICORN_PROBABILITY = compute_ai_unicorn_probability()

# --- Output Paths ---
BASE_DIR = Path(__file__).resolve().parent
FIGURES_DIR = BASE_DIR / "figures"
EXCEL_PATH = BASE_DIR / "guardspine-market-model.xlsx"
PDF_PATH = BASE_DIR / "GuardSpine-Market-Analysis-2026-Q1.pdf"


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def adoption_curve(t, r=ADOPTION_RATE, t0=INFLECTION_MONTH):
    return 1.0 / (1.0 + math.exp(-r * (t - t0)))

def fee_level(t, k, lag):
    if t <= lag:
        return 1.0
    cumulative = sum(adoption_curve(s) for s in range(lag, t + 1))
    return math.exp(-k * cumulative)

def governance_premium(t):
    return G0 + (GMAX - G0) * (1.0 - math.exp(-LAMBDA_GOV * t))

def customer_count(t, c0, growth):
    return c0 * ((1 + growth) ** t)

def blended_metric(dist, values):
    return sum(dist[k] * values[k] for k in dist)

def weighted_moat_score(scores):
    total_w = sum(MOAT_WEIGHTS.values())
    return sum(scores[d] * MOAT_WEIGHTS[d] for d in MOAT_WEIGHTS) / total_w

def tier_total_cogs(tier_data):
    # BYOK: no API costs. COGS = hosting + support + cognitive licensing only.
    return tier_data["hosting"] + tier_data["support"] + tier_data["cognitive_license"]


def compute_unicorn_path(params, max_years=7, mult_boost=1.0):
    """Compute annual ARR and implied valuation for unicorn path.
    mult_boost: multiplier applied to valuation multiples (e.g. 1.35 for AI-adjusted).
    """
    customers = params["c0"]
    arr = customers * params["acv"]
    results = [{"year": 0, "customers": customers, "arr": arr,
                "yoy": 0, "mult": 0, "valuation": 0}]
    for y in range(1, max_years + 1):
        g = params["g0"] * ((1 - params["decay"]) ** (y - 1))
        g = max(g, 0.02)
        prev_custs = customers
        customers = customers * ((1 + g) ** 12)
        new_logos = customers - prev_custs
        prev_arr = arr
        ent = params["enterprise"].get(y, 0)
        arr = prev_arr * params["nrr"] + new_logos * params["acv"] + ent
        yoy = (arr / prev_arr - 1) if prev_arr > 0 else 10
        # Valuation multiple
        if arr < 5e6:
            base = 20
        elif arr < 20e6:
            base = 15
        elif arr < 50e6:
            base = 12
        elif arr < 100e6:
            base = 10
        else:
            base = 8
        gp = 1.5 if yoy > 1.0 else (1.25 if yoy > 0.5 else 1.0)
        mult = base * gp * 1.15 * 1.10 * mult_boost
        val = arr * mult
        results.append({"year": y, "customers": round(customers), "arr": arr,
                        "yoy": yoy, "mult": round(mult, 1), "valuation": val,
                        "g_mo": g, "enterprise": ent})
    return results


def compute_personal_liquidity(equity=DAVID_EQUITY, rounds=None, tax=TAX_RATE,
                                exit_val=None, exit_label=None):
    """Compute founder equity value after dilution at each funding round."""
    if rounds is None:
        rounds = FUNDRAISING
    if exit_val is None:
        exit_val = EXIT_VALUATION
    if exit_label is None:
        exit_label = f"Exit (${exit_val/1e6:.0f}M)"
    ownership = equity
    results = []
    for r in rounds:
        ownership *= (1 - r["dilution"])
        pre_tax = ownership * r["valuation"]
        post_tax = pre_tax * (1 - tax)
        results.append({
            "round": r["round"], "month": r["month"],
            "ownership_pct": ownership * 100,
            "pre_tax": pre_tax, "post_tax": post_tax,
            "valuation": r["valuation"],
        })
    # Exit event
    final_ownership = ownership
    exit_pre = final_ownership * exit_val
    exit_post = exit_pre * (1 - tax)
    results.append({
        "round": exit_label, "month": EXIT_MONTH_RANGE[0],
        "ownership_pct": final_ownership * 100,
        "pre_tax": exit_pre, "post_tax": exit_post,
        "valuation": exit_val,
    })
    return results


def compute_path_comparison():
    """Compare bootstrap vs VC paths at various exit valuations."""
    exits = [296e6, 400e6, 600e6, 1e9, 1.5e9]
    paths = {
        "Bootstrap (angel only)": {"rounds": BOOTSTRAP_PATH, "color": "#43A047"},
        "VC 2-round (pre-seed+seed)": {"rounds": VC_PATH[:2], "color": "#1E88E5"},
        "VC 4-round (full)": {"rounds": VC_PATH, "color": "#E53935"},
    }
    results = {}
    for pname, pdata in paths.items():
        ownership = DAVID_EQUITY
        for r in pdata["rounds"]:
            ownership *= (1 - r["dilution"])
        outcomes = []
        for ev in exits:
            post_tax = ownership * ev * (1 - TAX_RATE)
            outcomes.append({"exit": ev, "post_tax": post_tax, "ownership": ownership})
        results[pname] = {"outcomes": outcomes, "ownership": ownership,
                          "color": pdata["color"],
                          "total_raised": sum(r["raise"] for r in pdata["rounds"])}
    return results, exits


def compute_dilution_waterfall(rounds=None, target_exits=None):
    """Compute round-by-round dilution cost at multiple exit valuations."""
    if rounds is None:
        rounds = VC_PATH  # show full VC path for comparison
    if target_exits is None:
        target_exits = [300e6, 600e6, 1.0e9, 2.0e9]
    ownership = DAVID_EQUITY
    results = []
    cumulative_raised = 0

    for r in rounds:
        pre_ownership = ownership
        ownership *= (1 - r["dilution"])
        equity_lost = pre_ownership - ownership
        cumulative_raised += r["raise"]

        costs = {}
        for ex in target_exits:
            pre_tax = equity_lost * ex
            post_tax = pre_tax * (1 - TAX_RATE)
            costs[ex] = {"pre_tax": pre_tax, "post_tax": post_tax}

        results.append({
            "round": r["round"],
            "month": r["month"],
            "raise": r["raise"],
            "pre_money": r["valuation"],
            "post_money": r["valuation"] + r["raise"],
            "dilution_pct": r["dilution"] * 100,
            "david_before": pre_ownership * 100,
            "david_after": ownership * 100,
            "equity_lost_pct": equity_lost * 100,
            "cumulative_raised": cumulative_raised,
            "costs": costs,
        })

    return results, ownership


def compute_acquisition_trajectory():
    """Compute acquisition price at each year, with strategic premium, for all scenarios."""
    all_results = {}
    for name, params in UNICORN_SCENARIOS.items():
        path = compute_unicorn_path(params)
        acq_path = []
        for r in path:
            if r["year"] == 0:
                acq_path.append({"year": 0, "arr": 0, "market_val": 0,
                                 "premium": 1.0, "acq_price": 0})
                continue
            arr = r["arr"]
            market_val = r["valuation"]
            if arr < 1e6:
                prem = STRATEGIC_PREMIUM["pre_revenue"]
            elif arr < 10e6:
                prem = STRATEGIC_PREMIUM["early"]
            elif arr < 50e6:
                prem = STRATEGIC_PREMIUM["growth"]
            else:
                prem = STRATEGIC_PREMIUM["scale"]
            acq_path.append({
                "year": r["year"],
                "arr": arr,
                "market_val": market_val,
                "premium": prem,
                "acq_price": market_val * prem,
                "yoy": r["yoy"],
            })
        all_results[name] = acq_path
    return all_results


def compute_ai_acquisition_trajectory():
    """Compute acquisition trajectory for AI-adjusted unicorn scenarios."""
    all_results = {}
    for name, params in AI_UNICORN_SCENARIOS.items():
        mb = params.get("mult_boost", 1.0)
        path = compute_unicorn_path(params, mult_boost=mb)
        acq_path = []
        for r in path:
            if r["year"] == 0:
                acq_path.append({"year": 0, "arr": 0, "market_val": 0,
                                 "premium": 1.0, "acq_price": 0})
                continue
            arr = r["arr"]
            market_val = r["valuation"]
            if arr < 1e6:
                prem = STRATEGIC_PREMIUM["pre_revenue"]
            elif arr < 10e6:
                prem = STRATEGIC_PREMIUM["early"]
            elif arr < 50e6:
                prem = STRATEGIC_PREMIUM["growth"]
            else:
                prem = STRATEGIC_PREMIUM["scale"]
            acq_path.append({
                "year": r["year"],
                "arr": arr,
                "market_val": market_val,
                "premium": prem,
                "acq_price": market_val * prem,
                "yoy": r["yoy"],
            })
        all_results[name] = acq_path
    return all_results


def find_billion_year(acq_trajectory, scenario="Base"):
    """Find the first year acquisition price exceeds $1B for a given scenario."""
    for r in acq_trajectory[scenario]:
        if r["acq_price"] >= TARGET_ACQUISITION:
            return r["year"], r["arr"], r["acq_price"]
    return None, None, None


# ============================================================
# FIGURE GENERATION
# ============================================================

def make_figures():
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"figure.dpi": 300, "savefig.dpi": 300, "font.size": 10,
                         "axes.titlesize": 13, "axes.labelsize": 11})

    # --- Figure 1: TAM/SAM/SOM with multi-artifact expansion ---
    fig, ax = plt.subplots(figsize=(8, 8))
    gov_layer = MULTI_ARTIFACT_TAM * GOVERNANCE_LAYER_PCT_HIGH
    circles = [
        (MULTI_ARTIFACT_TAM, "#90CAF9", f"Multi-Artifact TAM\n${MULTI_ARTIFACT_TAM/1e9:.1f}B"),
        (CODE_ONLY_TAM, "#2196F3", f"Code-Only TAM\n${CODE_ONLY_TAM/1e9:.1f}B"),
        (gov_layer, "#1565C0", f"Governance Layer\n${gov_layer/1e9:.1f}B"),
        (SAM_TOTAL, "#4CAF50", f"SAM\n${SAM_TOTAL/1e6:.0f}M"),
        (SAM_TOTAL * SOM_Y3_PCT, "#FF9800", f"SOM Y3\n${SAM_TOTAL*SOM_Y3_PCT/1e6:.0f}M"),
    ]
    max_r = 3.5
    offsets = [0, 0, 0, -0.3, -0.6]
    for i, (val, color, label) in enumerate(circles):
        r = max_r * math.sqrt(val / MULTI_ARTIFACT_TAM)
        circle = plt.Circle((0, offsets[i]), r, facecolor=color, alpha=0.25,
                             linewidth=2, edgecolor=color)
        ax.add_patch(circle)
        y_pos = offsets[i] + r * 0.4 if i < 3 else offsets[i]
        ax.text(0, y_pos, label, ha="center", va="center", fontsize=9, fontweight="bold")
    ax.set_xlim(-4, 4)
    ax.set_ylim(-4, 4)
    ax.set_aspect("equal")
    ax.set_title("GuardSpine Market Sizing\nMulti-Artifact AI Office", fontsize=14, fontweight="bold")
    ax.axis("off")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "tam_sam_som.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 2: Fee Cascade Curves ---
    months = list(range(37))
    fig, ax = plt.subplots(figsize=(10, 6))
    cmap = {"Audit": "#E53935", "Legal": "#FB8C00", "Consulting": "#FDD835",
            "Implementation": "#43A047", "Design": "#1E88E5"}
    for name, params in CASCADE_SECTORS.items():
        fees = [fee_level(t, params["k"], params["lag"]) for t in months]
        ax.plot(months, fees, label=f'{name} (k={params["k"]}, lag={params["lag"]}mo)',
                linewidth=2, color=cmap[name])
    ax.set_xlabel("Months from Now")
    ax.set_ylabel("Fee Level (1.0 = today)")
    ax.set_title("Professional Service Fee Cascade Under AI Adoption", fontsize=14, fontweight="bold")
    ax.legend(loc="lower left")
    ax.set_ylim(0, 1.05)
    ax.grid(True, alpha=0.3)
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(1.0))
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "fee_cascade.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 3: Governance Premium Growth ---
    fig, ax = plt.subplots(figsize=(10, 5))
    gov = [governance_premium(t) * 100 for t in months]
    ax.fill_between(months, gov, alpha=0.3, color="#7B1FA2")
    ax.plot(months, gov, linewidth=2.5, color="#7B1FA2")
    ax.axhline(y=G0 * 100, color="gray", linestyle="--", alpha=0.5, label=f"Current ({G0*100:.0f}%)")
    ax.axhline(y=GMAX * 100, color="gray", linestyle=":", alpha=0.5, label=f"Ceiling ({GMAX*100:.0f}%)")
    ax.set_xlabel("Months from Now")
    ax.set_ylabel("Governance Share of Software Spend (%)")
    ax.set_title("Governance Premium Growth Curve", fontsize=14, fontweight="bold")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "governance_premium.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 4: Revenue Scenario Fan Chart ---
    fig, ax = plt.subplots(figsize=(10, 6))
    scenario_colors = {"Bear": "#E53935", "Base": "#1E88E5", "Bull": "#43A047"}
    for name, params in SCENARIOS.items():
        custs = [customer_count(t, params["c0"], params["growth"]) for t in months]
        revenue = [c * params["avg_acv"] / 12 / 1e6 for c in custs]
        ax.plot(months, revenue, label=f'{name} (g={params["growth"]*100:.0f}%/mo, ACV=${params["avg_acv"]/1000:.0f}K)',
                linewidth=2.5, color=scenario_colors[name])
        ax.fill_between(months, revenue, alpha=0.1, color=scenario_colors[name])
    ax.set_xlabel("Months from Now")
    ax.set_ylabel("Monthly Revenue ($M)")
    ax.set_title("Revenue Scenarios: Bear / Base / Bull", fontsize=14, fontweight="bold")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "revenue_scenarios.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 5: Competitive Radar ---
    dims = list(MOAT_WEIGHTS.keys())
    dim_labels = ["Data Moat\n(3x)", "Network\nEffects (1x)", "Regulatory\nLock-in (2x)",
                  "Tech Diff\n(1x)", "Switching\nCosts (1x)"]
    n = len(dims)
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist() + [0]
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    comp_colors = {"GuardSpine": "#1E88E5", "GitHub": "#43A047", "SonarQube": "#FB8C00",
                   "Codebat": "#E53935", "Manual Review": "#9E9E9E"}
    for comp, scores in MOAT_SCORES.items():
        vals = [scores[d] for d in dims] + [scores[dims[0]]]
        ax.plot(angles, vals, "o-", linewidth=2, label=comp, color=comp_colors[comp])
        ax.fill(angles, vals, alpha=0.1, color=comp_colors[comp])
    ax.set_thetagrids([a * 180 / np.pi for a in angles[:-1]], dim_labels)
    ax.set_ylim(0, 10)
    ax.set_title("Competitive Moat Scorecard", fontsize=14, fontweight="bold", y=1.08)
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "competitive_radar.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 6: L0-L4 Risk Tier Distribution ---
    fig, ax = plt.subplots(figsize=(9, 5))
    tiers = list(TIER_DIST.keys())
    pcts = [TIER_DIST[t] * 100 for t in tiers]
    tier_colors = ["#81C784", "#AED581", "#FDD835", "#FB8C00", "#E53935"]
    labels = [f"{t}\n{TIER_NAMES[t]}" for t in tiers]
    bars = ax.bar(labels, pcts, color=tier_colors, edgecolor="white", linewidth=1.5, width=0.6)
    for bar, pct in zip(bars, pcts):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                f"{pct:.0f}%", ha="center", va="bottom", fontweight="bold")
    ax.set_ylabel("Share of Artifact Changes (%)")
    ax.set_xlabel("Risk Tier")
    ax.set_title("L0-L4 Risk Tier Distribution (Typical Enterprise)", fontsize=14, fontweight="bold")
    ax.set_ylim(0, 65)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "tier_distribution.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 7: MECE 9-Dimension Heatmap ---
    fig, ax = plt.subplots(figsize=(12, 5))
    competitors = list(MECE_MATRIX.keys())
    dimensions = list(MECE_MATRIX["GuardSpine"].keys())
    dim_display = ["Code", "Docs", "Sheets", "Images", "AI\nProvenance", "Risk\nGating",
                   "Evidence\nBundles", "Artifact\nDiffs", "Stop-\nthe-Line"]
    data = np.array([[MECE_MATRIX[c][d] for d in dimensions] for c in competitors])
    cmap_hm = plt.cm.RdYlGn
    im = ax.imshow(data, cmap=cmap_hm, aspect="auto", vmin=0, vmax=1)
    ax.set_xticks(range(len(dimensions)))
    ax.set_xticklabels(dim_display, fontsize=9)
    ax.set_yticks(range(len(competitors)))
    ax.set_yticklabels(competitors, fontsize=10)
    for i in range(len(competitors)):
        for j in range(len(dimensions)):
            v = data[i, j]
            txt = "Yes" if v == 1 else ("Partial" if v == 0.5 else "No")
            color = "white" if v < 0.3 else "black"
            ax.text(j, i, txt, ha="center", va="center", fontsize=8, fontweight="bold", color=color)
    ax.set_title("MECE Competitive Matrix: 9 Governance Dimensions", fontsize=14, fontweight="bold")
    fig.colorbar(im, ax=ax, shrink=0.6, label="Coverage")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "mece_heatmap.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 8: Four Guard Lanes Architecture ---
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.axis("off")
    lane_data = [
        ("CodeGuard\n(PRs, Commits)", "#1E88E5", "SARIF + Diffs"),
        ("PDFGuard\n(Contracts, Policy)", "#7B1FA2", "Redline + Clauses"),
        ("SheetGuard\n(Models, KPIs)", "#43A047", "Heatmap + Formulas"),
        ("ImageGuard\n(Screenshots, Charts)", "#FB8C00", "Overlay + Tags"),
    ]
    for i, (name, color, output) in enumerate(lane_data):
        x = 0.12 + i * 0.22
        rect = plt.Rectangle((x, 0.35), 0.18, 0.35, facecolor=color, alpha=0.3,
                              edgecolor=color, linewidth=2, transform=ax.transAxes)
        ax.add_patch(rect)
        ax.text(x + 0.09, 0.58, name, ha="center", va="center", fontsize=10,
                fontweight="bold", transform=ax.transAxes)
        ax.text(x + 0.09, 0.40, output, ha="center", va="center", fontsize=8,
                color="gray", transform=ax.transAxes)
    # Bottom bar: Evidence Bundle
    rect_bottom = plt.Rectangle((0.08, 0.12), 0.84, 0.15, facecolor="#E53935", alpha=0.2,
                                edgecolor="#E53935", linewidth=2, transform=ax.transAxes)
    ax.add_patch(rect_bottom)
    ax.text(0.50, 0.195, "Evidence Bundle (Hash-Chained, Offline-Verifiable)", ha="center",
            va="center", fontsize=11, fontweight="bold", transform=ax.transAxes)
    # Top bar: YAML Rubrics
    rect_top = plt.Rectangle((0.08, 0.78), 0.84, 0.12, facecolor="#607D8B", alpha=0.2,
                              edgecolor="#607D8B", linewidth=2, transform=ax.transAxes)
    ax.add_patch(rect_top)
    ax.text(0.50, 0.84, "YAML Rubric Engine (11 packs, 106+ rules, swappable per domain)",
            ha="center", va="center", fontsize=10, fontweight="bold", transform=ax.transAxes)
    # Side labels
    ax.text(0.03, 0.52, "PII-Shield\n(sanitization)", ha="center", va="center", fontsize=8,
            color="#E53935", rotation=90, transform=ax.transAxes)
    ax.text(0.97, 0.52, "Proprioceptive AI\n(cognitive probes)", ha="center", va="center",
            fontsize=8, color="#7B1FA2", rotation=-90, transform=ax.transAxes)
    ax.set_title("GuardSpine: Four Guard Lanes Architecture", fontsize=14, fontweight="bold")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "four_lanes.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 9: Unicorn Path ---
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
    scenario_colors = {"Bear": "#E53935", "Base": "#1E88E5", "Bull": "#43A047"}
    max_y = 7

    for name, params in UNICORN_SCENARIOS.items():
        path = compute_unicorn_path(params, max_years=max_y)
        years = [r["year"] for r in path]
        arr_vals = [r["arr"] / 1e6 for r in path]
        val_vals = [r["valuation"] / 1e6 for r in path]
        color = scenario_colors[name]
        ax1.plot(years, arr_vals, "o-", linewidth=2.5, label=name, color=color, markersize=5)
        ax2.plot(years, val_vals, "o-", linewidth=2.5, label=name, color=color, markersize=5)
        # Mark unicorn crossing
        for r in path:
            if r["valuation"] >= 1e9:
                ax2.plot(r["year"], r["valuation"] / 1e6, "*", color=color,
                         markersize=18, markeredgecolor="black", markeredgewidth=0.5, zorder=5)
                break

    ax1.set_xlabel("Year")
    ax1.set_ylabel("ARR ($M)")
    ax1.set_title("ARR Growth Path", fontsize=13, fontweight="bold")
    ax1.set_yscale("log")
    ax1.set_ylim(0.05, 500)
    ax1.legend()
    ax1.grid(True, alpha=0.3, which="both")
    ax1.set_xticks(range(max_y + 1))

    ax2.axhline(y=1000, color="gold", linestyle="--", linewidth=2.5, alpha=0.8, label="$1B Unicorn")
    ax2.set_xlabel("Year")
    ax2.set_ylabel("Implied Valuation ($M)")
    ax2.set_title("Path to Unicorn", fontsize=13, fontweight="bold")
    ax2.set_yscale("log")
    ax2.set_ylim(5, 5000)
    ax2.legend()
    ax2.grid(True, alpha=0.3, which="both")
    ax2.set_xticks(range(max_y + 1))

    fig.suptitle("GuardSpine: Unicorn Path Analysis (with Netflix + IBM catalysts)",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "unicorn_path.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 10: Probability Risk Waterfall ---
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 7))

    # Left panel: stacked risk factors per team configuration
    configs = ["solo", "igor", "igor_chris", "pre_mortem", "triangle", "current"]
    config_labels = ["Solo\n(David)", "+ Igor\n(CTO)", "+ Igor\n+ Chris",
                     "Post\nPre-Mortem", "+ Triangle\nStrategy", "Current\n(Feb 2026)"]
    factors = list(RISK_FACTORS.keys())
    x = np.arange(len(configs))
    bar_width = 0.6
    bottom_vals = np.zeros(len(configs))
    risk_colors = ["#E53935", "#FB8C00", "#FDD835", "#66BB6A", "#42A5F5",
                   "#7E57C2", "#8D6E63", "#78909C"]
    for fi, factor in enumerate(factors):
        vals = [RISK_FACTORS[factor][c] * 100 for c in configs]
        ax1.bar(x, vals, bar_width, bottom=bottom_vals, label=factor,
                color=risk_colors[fi], edgecolor="white", linewidth=0.5)
        bottom_vals += np.array(vals)
    ax1.set_xticks(x)
    ax1.set_xticklabels(config_labels, fontsize=8)
    ax1.set_ylabel("Cumulative Risk (%)")
    ax1.set_title("Risk Factor Decomposition by Team + Strategy", fontsize=12, fontweight="bold")
    ax1.legend(loc="upper right", fontsize=6, ncol=2)
    ax1.set_ylim(0, 300)
    ax1.grid(axis="y", alpha=0.2)

    # Right panel: unicorn probability progression
    probs = list(UNICORN_PROBABILITIES.values())
    labels = list(UNICORN_PROBABILITIES.keys())
    colors_prob = ["#E53935", "#FB8C00", "#42A5F5", "#43A047", "#7B1FA2", "#00BFA5"]
    bars = ax2.barh(range(len(probs)), [p * 100 for p in probs], color=colors_prob,
                     edgecolor="white", linewidth=1.5, height=0.6)
    for i, (bar, p) in enumerate(zip(bars, probs)):
        ax2.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                 f"{p*100:.1f}%", va="center", fontweight="bold", fontsize=11)
    ax2.set_yticks(range(len(labels)))
    ax2.set_yticklabels(labels, fontsize=9)
    ax2.set_xlabel("Unicorn Probability (%)")
    ax2.set_title("Cumulative Unicorn Probability", fontsize=12, fontweight="bold")
    ax2.set_xlim(0, 42)
    ax2.invert_yaxis()
    ax2.grid(axis="x", alpha=0.3)
    # Reference lines
    ax2.axvline(x=5, color="gray", linestyle="--", alpha=0.4, linewidth=1)
    ax2.text(5.2, 4.7, "Avg Series A\n(~5%)", fontsize=7, color="gray")

    fig.suptitle("GuardSpine: Pre-Mortem Risk & Probability Analysis (with Triangle Strategy)",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "probability_waterfall.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 11: Bootstrap vs VC Path Comparison ---
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 7))

    # Left panel: David's post-tax at various exit valuations for each path
    comparison, exit_vals = compute_path_comparison()
    exit_labels = [f"${v/1e6:.0f}M" for v in exit_vals]
    x = np.arange(len(exit_vals))
    bar_w = 0.25
    for i, (pname, pdata) in enumerate(comparison.items()):
        vals = [o["post_tax"] / 1e6 for o in pdata["outcomes"]]
        bars = ax1.bar(x + i * bar_w, vals, bar_w, label=f'{pname}\n({pdata["ownership"]*100:.1f}%, ${pdata["total_raised"]/1e3:.0f}K raised)',
                color=pdata["color"], edgecolor="white", linewidth=0.5)
        for bar, v in zip(bars, vals):
            if v > 5:
                ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 2,
                         f"${v:.0f}M", ha="center", va="bottom", fontsize=6, fontweight="bold")

    ax1.axhline(y=100, color="gold", linestyle="--", linewidth=2, alpha=0.8, label="$100M target")
    ax1.set_xticks(x + bar_w)
    ax1.set_xticklabels(exit_labels, fontsize=9)
    ax1.set_xlabel("Exit Valuation")
    ax1.set_ylabel("David Post-Tax ($M)")
    ax1.set_title("David's Take-Home by Exit & Capital Path", fontsize=12, fontweight="bold")
    ax1.legend(loc="upper left", fontsize=6)
    ax1.grid(axis="y", alpha=0.3)

    # Right panel: Bootstrap liquidity timeline
    # Show bootstrap path with exit scenarios
    boot_exits = [316e6, 400e6, 600e6]
    boot_labels = ["$316M\n($100M target)", "$400M", "$600M"]
    boot_colors = ["#E53935", "#FB8C00", "#43A047"]
    for ev, elabel, ecolor in zip(boot_exits, boot_labels, boot_colors):
        liq = compute_personal_liquidity(rounds=BOOTSTRAP_PATH, exit_val=ev,
                                          exit_label=f"Exit (${ev/1e6:.0f}M)")
        months_liq = [0] + [r["month"] for r in liq]
        values_liq = [0] + [r["post_tax"] / 1e6 for r in liq]
        ax2.plot(months_liq, values_liq, "o-", linewidth=2.5, color=ecolor, markersize=6,
                 label=f"Exit @ {elabel}", zorder=5)
        # Annotate exit point
        final = liq[-1]
        ax2.annotate(f'${final["post_tax"]/1e6:.0f}M\n({final["ownership_pct"]:.1f}%)',
                     xy=(final["month"], final["post_tax"] / 1e6),
                     xytext=(8, 0), textcoords="offset points",
                     ha="left", fontsize=8, fontweight="bold", color=ecolor)

    # VC comparison (faded)
    vc_liq = compute_personal_liquidity(rounds=VC_PATH, exit_val=600e6,
                                         exit_label="Exit ($600M)")
    months_vc = [0] + [r["month"] for r in vc_liq]
    values_vc = [0] + [r["post_tax"] / 1e6 for r in vc_liq]
    ax2.plot(months_vc, values_vc, "s--", linewidth=1.5, color="#9E9E9E", markersize=4,
             alpha=0.5, label="VC 4-round @ $600M (19%)", zorder=3)

    ax2.axhline(y=100, color="gold", linestyle="--", linewidth=2, alpha=0.8)
    ax2.text(1, 105, "$100M target", fontsize=8, color="goldenrod", fontweight="bold")
    ax2.set_xlabel("Month")
    ax2.set_ylabel("Post-Tax Equity Value ($M)")
    ax2.set_title("Bootstrap Timeline: David's Liquidity Path\n(42.19% ownership, $1M angel)",
                  fontsize=11, fontweight="bold")
    ax2.set_xlim(0, 42)
    ax2.set_ylim(0, 220)
    ax2.legend(loc="upper left", fontsize=7)
    ax2.grid(True, alpha=0.3)

    fig.suptitle("GuardSpine: Bootstrap vs VC -- Capital Efficiency Analysis",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "personal_liquidity.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 12: Acquirer Strategic Fit Scorecard ---
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

    acquirers = list(ACQUIRER_SCORES.keys())
    dims = list(ACQUIRER_WEIGHTS.keys())
    dim_labels = ["Strategic\nGap (3x)", "Distribution\n(2x)", "Precedent\nAcq. (2x)",
                  "Integration\nEase (1x)"]
    acq_colors = ["#1E88E5", "#43A047", "#FB8C00", "#7B1FA2", "#E53935"]

    # Left panel: grouped bar chart
    x = np.arange(len(dims))
    bar_w = 0.15
    for i, acq in enumerate(acquirers):
        vals = [ACQUIRER_SCORES[acq][d] for d in dims]
        ax1.bar(x + i * bar_w, vals, bar_w, label=acq.split("/")[0].strip(),
                color=acq_colors[i], edgecolor="white", linewidth=0.5)
    ax1.set_xticks(x + bar_w * 2)
    ax1.set_xticklabels(dim_labels, fontsize=8)
    ax1.set_ylabel("Score (1-10)")
    ax1.set_ylim(0, 12)
    ax1.set_title("Acquirer Fit by Dimension", fontsize=12, fontweight="bold")
    ax1.legend(loc="upper right", fontsize=7)
    ax1.grid(axis="y", alpha=0.2)

    # Right panel: weighted composite + price range
    total_w = sum(ACQUIRER_WEIGHTS.values())
    composites = []
    for acq in acquirers:
        score = sum(ACQUIRER_SCORES[acq][d] * ACQUIRER_WEIGHTS[d] for d in dims) / total_w
        composites.append(score)
    sorted_idx = np.argsort(composites)[::-1]
    sorted_acqs = [acquirers[i] for i in sorted_idx]
    sorted_scores = [composites[i] for i in sorted_idx]
    sorted_colors = [acq_colors[i] for i in sorted_idx]
    sorted_ranges = [ACQUIRER_SCORES[a]["price_range_m"] for a in sorted_acqs]

    bars = ax2.barh(range(len(sorted_acqs)), sorted_scores, color=sorted_colors,
                     edgecolor="white", linewidth=1.5, height=0.6)
    def fmt_price(m):
        return f"${m/1000:.1f}B" if m >= 1000 else f"${m:.0f}M"
    for i, (bar, score, (lo, hi)) in enumerate(zip(bars, sorted_scores, sorted_ranges)):
        ax2.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                 f"{score:.1f}  ({fmt_price(lo)}-{fmt_price(hi)})",
                 va="center", fontweight="bold", fontsize=9)
    ax2.set_yticks(range(len(sorted_acqs)))
    ax2.set_yticklabels([a.split("/")[0].strip() if "/" not in a else a for a in sorted_acqs],
                         fontsize=9)
    ax2.set_xlabel("Weighted Composite Score")
    ax2.set_title("Ranked Acquirers (score + price range)", fontsize=12, fontweight="bold")
    ax2.set_xlim(0, 12)
    ax2.invert_yaxis()
    ax2.grid(axis="x", alpha=0.3)

    fig.suptitle("GuardSpine: Strategic Acquirer Analysis",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "acquirer_scorecard.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 13: Acquisition Price Trajectory ---
    acq_traj = compute_acquisition_trajectory()
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 7))
    scenario_colors = {"Bear": "#E53935", "Base": "#1E88E5", "Bull": "#43A047"}

    # Left panel: All 3 scenarios with $1B line
    for name in ["Bear", "Base", "Bull"]:
        years = [r["year"] for r in acq_traj[name] if r["year"] > 0]
        prices = [r["acq_price"] / 1e6 for r in acq_traj[name] if r["year"] > 0]
        ax1.plot(years, prices, "o-", linewidth=2.5, color=scenario_colors[name],
                 label=name, markersize=6)
        # Mark $1B crossing
        for r in acq_traj[name]:
            if r["acq_price"] >= TARGET_ACQUISITION:
                ax1.plot(r["year"], r["acq_price"] / 1e6, "*", color=scenario_colors[name],
                         markersize=18, markeredgecolor="black", markeredgewidth=0.5, zorder=5)
                break

    ax1.axhline(y=1000, color="gold", linestyle="--", linewidth=2.5, alpha=0.8, label="$1B Target")
    ax1.axhline(y=BOOTSTRAP_EXIT_NEEDED / 1e6, color="#43A047", linestyle=":", linewidth=2,
                alpha=0.8, label=f"Bootstrap $100M target (${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M)")
    ax1.set_xlabel("Year")
    ax1.set_ylabel("Acquisition Price ($M)")
    ax1.set_title("Acquisition Price Trajectory\n(Market Valuation + Strategic Premium)", fontsize=12, fontweight="bold")
    ax1.set_yscale("log")
    ax1.set_ylim(30, 5000)
    ax1.legend(loc="upper left", fontsize=8)
    ax1.grid(True, alpha=0.3, which="both")
    ax1.set_xticks(range(1, 8))

    # Right panel: Base case with acquirer price range bands
    base_years = [r["year"] for r in acq_traj["Base"] if r["year"] > 0]
    base_prices = [r["acq_price"] / 1e6 for r in acq_traj["Base"] if r["year"] > 0]
    ax2.plot(base_years, base_prices, "o-", linewidth=3, color="#1E88E5",
             label="Base Case Acq. Price", markersize=7, zorder=5)

    # Acquirer bands
    band_colors = ["#90CAF9", "#A5D6A7", "#FFCC80", "#CE93D8", "#EF9A9A"]
    acq_names_short = ["MSFT/GH", "IBM", "PANW", "ServiceNow", "CrowdStrike"]
    acq_list = list(ACQUIRER_SCORES.items())
    for i, (name, scores) in enumerate(acq_list):
        lo, hi = scores["price_range_m"]
        ax2.axhspan(lo, hi, alpha=0.15, color=band_colors[i], label=f"{acq_names_short[i]} ({fmt_price(lo)}-{fmt_price(hi)})")

    ax2.axhline(y=1000, color="gold", linestyle="--", linewidth=2, alpha=0.8)
    ax2.axhline(y=BOOTSTRAP_EXIT_NEEDED / 1e6, color="#43A047", linestyle=":", linewidth=2,
                alpha=0.8, label=f"Bootstrap target (${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M)")
    ax2.set_xlabel("Year")
    ax2.set_ylabel("Price ($M)")
    ax2.set_title("Base Case vs Acquirer Price Ranges", fontsize=12, fontweight="bold")
    ax2.set_ylim(0, 2200)
    ax2.legend(loc="upper left", fontsize=7)
    ax2.grid(True, alpha=0.3)
    ax2.set_xticks(range(1, 8))

    fig.suptitle("GuardSpine: Path to $1B+ Acquisition",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "acquisition_trajectory.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 14: Bootstrap vs VC Dilution Comparison ---
    # Left panel: ownership comparison across paths
    # Right panel: David's take-home at key exit values
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 7))

    # Compute all three paths
    boot_results, boot_own = compute_dilution_waterfall(rounds=BOOTSTRAP_PATH,
                                                         target_exits=[316e6, 400e6, 600e6, 1e9])
    vc2_results, vc2_own = compute_dilution_waterfall(rounds=VC_PATH[:2],
                                                       target_exits=[305e6, 400e6, 600e6, 1e9])
    vc4_results, vc4_own = compute_dilution_waterfall(rounds=VC_PATH,
                                                       target_exits=[305e6, 400e6, 600e6, 1e9])

    # Left panel: ownership bars for each path
    path_names = ["Bootstrap\n(angel $1M)", "VC 2-round\n($6M)", "VC 4-round\n($76M)"]
    path_owns = [boot_own * 100, vc2_own * 100, vc4_own * 100]
    path_raised = [1e6, 6e6, 76e6]
    path_colors = ["#43A047", "#1E88E5", "#E53935"]

    bars = ax1.bar(path_names, path_owns, color=path_colors, edgecolor="white", linewidth=2, width=0.55)
    for bar, own, raised in zip(bars, path_owns, path_raised):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.8,
                 f"{own:.1f}%\n(raised ${raised/1e6:.1f}M)" if raised >= 1e6
                 else f"{own:.1f}%\n(raised ${raised/1e3:.0f}K)",
                 ha="center", va="bottom", fontweight="bold", fontsize=10)

    # Show equity lost as red portion
    start_pct = DAVID_EQUITY * 100
    for i, (bar, own) in enumerate(zip(bars, path_owns)):
        lost = start_pct - own
        ax1.bar(path_names[i], lost, bottom=own, color="#FFCDD2", edgecolor="white",
                linewidth=1, width=0.55, alpha=0.6)

    ax1.axhline(y=start_pct, color="black", linestyle=":", linewidth=1, alpha=0.4)
    ax1.text(2.4, start_pct + 0.5, f"Start: {start_pct:.0f}%", fontsize=8, color="gray")
    ax1.set_ylabel("David's Ownership (%)")
    ax1.set_title("Ownership After Dilution\n(Bootstrap vs VC paths)", fontsize=12, fontweight="bold")
    ax1.set_ylim(0, 58)
    ax1.grid(axis="y", alpha=0.3)

    # Right panel: David's post-tax take at 4 exit valuations
    exit_vals = [305e6, 400e6, 600e6, 1e9]
    exit_labels = ["$305M\n(bootstrap\ntarget)", "$400M", "$600M", "$1B"]
    x = np.arange(len(exit_vals))
    bar_w = 0.25

    for pi, (pname, own, pcolor) in enumerate(zip(
        ["Bootstrap", "VC 2-round", "VC 4-round"],
        [boot_own, vc2_own, vc4_own],
        path_colors
    )):
        takehomes = [own * ev * (1 - TAX_RATE) / 1e6 for ev in exit_vals]
        b = ax2.bar(x + pi * bar_w, takehomes, bar_w, label=pname,
                    color=pcolor, edgecolor="white", linewidth=0.5)
        for bar, val in zip(b, takehomes):
            if val > 5:
                ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 2,
                         f"${val:.0f}M", ha="center", fontsize=7, fontweight="bold",
                         color=pcolor)

    ax2.axhline(y=100, color="gold", linestyle="--", linewidth=2, alpha=0.8)
    ax2.text(0.0, 103, "$100M target", fontsize=8, color="goldenrod", fontweight="bold")
    ax2.set_xticks(x + bar_w)
    ax2.set_xticklabels(exit_labels, fontsize=8)
    ax2.set_ylabel("David's Post-Tax Take ($M)")
    ax2.set_title("Personal Liquidity at Exit\n(post-tax, 25% rate)", fontsize=12, fontweight="bold")
    ax2.legend(loc="upper left", fontsize=8)
    ax2.grid(axis="y", alpha=0.3)

    fig.suptitle("GuardSpine: Why Bootstrap Wins -- Dilution Cost Analysis",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "dilution_waterfall.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 15: Speed-to-$100M Execution Timeline ---
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10), gridspec_kw={"height_ratios": [3, 1.2]})

    # Top panel: Gantt-style phase timeline with ARR overlay
    phase_colors = ["#1565C0", "#2E7D32", "#E65100", "#6A1B9A", "#C62828", "#00695C"]
    n_phases = len(SPEED_TIMELINE)

    for i, phase in enumerate(SPEED_TIMELINE):
        m0, m1 = phase["months"]
        color = phase_colors[i % len(phase_colors)]
        # Phase bar
        ax1.barh(n_phases - 1 - i, m1 - m0, left=m0, height=0.6,
                 color=color, alpha=0.85, edgecolor="white", linewidth=1.5)
        # Phase label inside bar
        mid = (m0 + m1) / 2
        ax1.text(mid, n_phases - 1 - i, phase["phase"],
                 ha="center", va="center", fontweight="bold", fontsize=9, color="white")
        # Action bullets to the right
        actions_text = " | ".join(phase["actions"][:2])
        if len(phase["actions"]) > 2:
            actions_text += f" (+{len(phase['actions'])-2} more)"
        ax1.text(m1 + 0.3, n_phases - 1 - i, actions_text,
                 ha="left", va="center", fontsize=7, color=color, style="italic",
                 clip_on=True)

    # ARR trajectory overlay (secondary axis)
    ax1_arr = ax1.twinx()
    arr_months = [0]
    arr_values = [0]
    for phase in SPEED_TIMELINE:
        arr_months.append(phase["months"][1])
        arr_values.append(phase["arr"] / 1e6)
    ax1_arr.plot(arr_months, arr_values, "D-", color="#FF6F00", linewidth=2.5,
                 markersize=7, zorder=10, label="ARR ($M)")
    ax1_arr.set_ylabel("ARR ($M)", color="#FF6F00", fontsize=10)
    ax1_arr.tick_params(axis="y", labelcolor="#FF6F00")
    ax1_arr.set_ylim(0, 25)

    # Target lines
    ax1.axvline(x=SPEED_BEST_CASE_MONTHS, color="#43A047", linestyle="--", linewidth=2, alpha=0.7)
    ax1.text(SPEED_BEST_CASE_MONTHS, n_phases - 0.3, f"Best case: {SPEED_BEST_CASE_MONTHS}mo",
             fontsize=8, color="#43A047", fontweight="bold", ha="center")
    ax1.axvline(x=SPEED_REALISTIC_MONTHS, color="#1565C0", linestyle="--", linewidth=2, alpha=0.7)
    ax1.text(SPEED_REALISTIC_MONTHS, n_phases - 0.3, f"Realistic: {SPEED_REALISTIC_MONTHS}mo",
             fontsize=8, color="#1565C0", fontweight="bold", ha="center")

    ax1.set_xlim(-0.5, 30)
    ax1.set_xlabel("Month")
    ax1.set_yticks(range(n_phases))
    ax1.set_yticklabels([""] * n_phases)
    ax1.set_title("Speed-to-\\$100M: Execution Timeline\n(\\$305M exit at 43.65% ownership = \\$100M post-tax)",
                  fontsize=13, fontweight="bold")
    ax1.grid(axis="x", alpha=0.3)

    # Bottom panel: Speed killers
    killer_names = [k["name"] for k in SPEED_KILLERS]
    killer_costs = [k["cost_months"] for k in SPEED_KILLERS]
    killer_colors = ["#E53935", "#FF7043", "#FFA726", "#FFCA28", "#66BB6A"]

    bars = ax2.barh(range(len(SPEED_KILLERS)), killer_costs,
                     color=killer_colors, edgecolor="white", linewidth=1.5, height=0.55)
    for i, (bar, killer) in enumerate(zip(bars, SPEED_KILLERS)):
        ax2.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                 f"+{killer['cost_months']}mo: {killer['description']}",
                 va="center", fontsize=8, color="#424242")
    ax2.set_yticks(range(len(SPEED_KILLERS)))
    ax2.set_yticklabels(killer_names, fontsize=9, fontweight="bold")
    ax2.set_xlabel("Months Added to Timeline")
    ax2.set_title("Speed Killers: Avoid These (each adds months to exit)",
                  fontsize=12, fontweight="bold", color="#C62828")
    ax2.set_xlim(0, 10)
    ax2.grid(axis="x", alpha=0.3)
    ax2.invert_yaxis()

    fig.suptitle("GuardSpine: Speed Execution Plan",
                 fontsize=15, fontweight="bold", y=1.01)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "speed_timeline.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 16: AI-Adjusted Math Comparison ---
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))

    # Panel A (top-left): Timeline compression -- baseline vs AI-adjusted Gantt
    ax_tl = axes[0, 0]
    # Baseline phases (from SPEED_TIMELINE)
    baseline_phases = [(p["phase"], p["months"]) for p in SPEED_TIMELINE]
    ai_phases = [(p["phase"], p["months"]) for p in AI_SPEED_TIMELINE]

    for i, (name, (m0, m1)) in enumerate(baseline_phases):
        ax_tl.barh(len(baseline_phases) - 1 - i + 0.2, m1 - m0, left=m0, height=0.35,
                    color="#90A4AE", alpha=0.7, edgecolor="white")
    for i, (name, (m0, m1)) in enumerate(ai_phases):
        ax_tl.barh(len(ai_phases) - 1 - i - 0.2, m1 - m0, left=m0, height=0.35,
                    color="#1565C0", alpha=0.9, edgecolor="white")

    ax_tl.axvline(x=SPEED_BEST_CASE_MONTHS, color="#90A4AE", linestyle="--", linewidth=1.5, alpha=0.6)
    ax_tl.axvline(x=AI_SPEED_BEST_CASE_MONTHS, color="#1565C0", linestyle="--", linewidth=2)
    ax_tl.axvline(x=SPEED_REALISTIC_MONTHS, color="#90A4AE", linestyle=":", linewidth=1.5, alpha=0.6)
    ax_tl.axvline(x=AI_SPEED_REALISTIC_MONTHS, color="#1565C0", linestyle=":", linewidth=2)

    ax_tl.set_yticks(range(len(baseline_phases)))
    ax_tl.set_yticklabels([p[0] for p in reversed(baseline_phases)], fontsize=8)
    ax_tl.set_xlabel("Month")
    ax_tl.set_title("Timeline Compression", fontsize=11, fontweight="bold")
    ax_tl.legend(["Baseline", "AI-Adjusted"], loc="lower right", fontsize=8)

    # Panel B (top-right): Revenue multiple expansion
    ax_tr = axes[0, 1]
    scenarios = ["Bear", "Base", "Bull"]
    baseline_low = [15, 15, 15]
    baseline_high = [25, 25, 25]
    adjusted_low = [AI_SCENARIOS[f"{s} (AI-adjusted)"]["multiple_low"] for s in scenarios]
    adjusted_high = [AI_SCENARIOS[f"{s} (AI-adjusted)"]["multiple_high"] for s in scenarios]

    x_pos = np.arange(len(scenarios))
    width = 0.35
    # Baseline range bars
    ax_tr.bar(x_pos - width/2, baseline_high, width, color="#90A4AE", alpha=0.7, label="Baseline high")
    ax_tr.bar(x_pos - width/2, baseline_low, width, color="#B0BEC5", alpha=0.5)
    # AI-adjusted range bars
    ax_tr.bar(x_pos + width/2, adjusted_high, width, color="#1565C0", alpha=0.9, label="AI-Adjusted high")
    ax_tr.bar(x_pos + width/2, adjusted_low, width, color="#42A5F5", alpha=0.5)

    ax_tr.set_xticks(x_pos)
    ax_tr.set_xticklabels(scenarios)
    ax_tr.set_ylabel("Revenue Multiple (x)")
    ax_tr.set_title("Revenue Multiple Expansion", fontsize=11, fontweight="bold")
    ax_tr.legend(fontsize=7)

    # Panel C (bottom-left): ARR needed for $305M exit
    ax_bl = axes[1, 0]
    multiples = [15, 20, 25, 30, 35]
    arr_needed = [BOOTSTRAP_EXIT_NEEDED / m / 1e6 for m in multiples]
    colors_arr = ["#E53935" if a > 15 else "#FF9800" if a > 10 else "#43A047" for a in arr_needed]

    bars_arr = ax_bl.bar(range(len(multiples)), arr_needed, color=colors_arr, edgecolor="white", linewidth=1.5)
    ax_bl.set_xticks(range(len(multiples)))
    ax_bl.set_xticklabels([f"{m}x" for m in multiples])
    for bar, val in zip(bars_arr, arr_needed):
        ax_bl.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                   f"\\${val:.1f}M", ha="center", fontsize=9, fontweight="bold")
    ax_bl.axhline(y=12, color="#1565C0", linestyle="--", linewidth=1.5, alpha=0.7,
                  label="AI-adjusted target ($9-12M)")
    ax_bl.axhspan(9, 12, alpha=0.1, color="#1565C0")
    ax_bl.set_ylabel("ARR Needed (\\$M)")
    ax_bl.set_xlabel("Revenue Multiple")
    ax_bl.set_title("ARR Required for \\$305M Exit", fontsize=11, fontweight="bold")
    ax_bl.legend(fontsize=8)

    # Panel D (bottom-right): 5 effects summary (horizontal bars showing magnitude)
    ax_br = axes[1, 1]
    effect_names = [
        "TAM CAGR\n(38% -> 60%)",
        "Lane delivery\n(3mo -> 1mo)",
        "Head start\n(18mo -> 9mo)",
        "ACV\n(1.5x increase)",
        "Multiple\n(15-25x -> 20-35x)",
    ]
    effect_magnitudes = [
        (0.60 - 0.38) / 0.38 * 100,   # +58% TAM CAGR increase
        (3 - 1) / 3 * 100,              # 67% time saved on lane delivery
        (18 - 9) / 18 * 100,            # 50% head start compressed (risk)
        50,                              # 50% ACV increase
        (27.5 - 20) / 20 * 100,         # 37.5% multiple expansion (midpoint)
    ]
    effect_colors = ["#43A047", "#43A047", "#E53935", "#43A047", "#43A047"]

    bars_eff = ax_br.barh(range(len(effect_names)), effect_magnitudes,
                           color=effect_colors, edgecolor="white", linewidth=1.5, height=0.55)
    for bar, val, color in zip(bars_eff, effect_magnitudes, effect_colors):
        label = f"+{val:.0f}%" if color == "#43A047" else f"-{val:.0f}%"
        ax_br.text(bar.get_width() + 1, bar.get_y() + bar.get_height()/2,
                   label, va="center", fontsize=9, fontweight="bold", color=color)
    ax_br.set_yticks(range(len(effect_names)))
    ax_br.set_yticklabels(effect_names, fontsize=8)
    ax_br.set_xlabel("Impact (%)")
    ax_br.set_title("AI Trajectory: 5 Effects on GuardSpine", fontsize=11, fontweight="bold")
    ax_br.invert_yaxis()

    fig.suptitle("GuardSpine: AI-Adjusted Math (Sourced Data, Feb 2026)",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "ai_adjusted_math.png", bbox_inches="tight")
    plt.close(fig)

    # --- Figure 17: AI-Adjusted Unicorn Path ---
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))

    # Left panel: Acquisition price trajectories (baseline vs AI-adjusted, base case)
    baseline_traj = compute_acquisition_trajectory()
    ai_traj = compute_ai_acquisition_trajectory()

    colors_base = {"Bear": "#90CAF9", "Base": "#1565C0", "Bull": "#0D47A1"}
    colors_ai = {"Bear (AI)": "#A5D6A7", "Base (AI)": "#2E7D32", "Bull (AI)": "#1B5E20"}

    # Baseline (dashed)
    for name, path in baseline_traj.items():
        years = [r["year"] for r in path[1:]]
        acqs = [r["acq_price"] / 1e6 for r in path[1:]]
        ax1.plot(years, acqs, "--", color=colors_base[name], linewidth=1.5,
                 alpha=0.6, label=f"{name} (baseline)")

    # AI-adjusted (solid)
    for name, path in ai_traj.items():
        years = [r["year"] for r in path[1:]]
        acqs = [min(r["acq_price"] / 1e6, 5000) for r in path[1:]]  # cap at $5B for readability
        ax1.plot(years, acqs, "-", color=colors_ai[name], linewidth=2.5,
                 marker="o", markersize=5, label=f"{name}")

    # Target lines
    ax1.axhline(y=305, color="#FF6F00", linestyle="--", linewidth=2,
                label="\\$305M (\\$100M post-tax)")
    ax1.axhline(y=1000, color="#C62828", linestyle="--", linewidth=2,
                label="\\$1B (unicorn)")
    ax1.set_xlabel("Year")
    ax1.set_ylabel("Acquisition Price (\\$M)")
    ax1.set_title("Path to Unicorn: Baseline vs AI-Adjusted", fontsize=12, fontweight="bold")
    ax1.legend(fontsize=7, loc="upper left")
    ax1.set_ylim(0, 5000)
    ax1.set_xlim(0, 7)
    ax1.grid(alpha=0.3)

    # Right panel: Unicorn probability waterfall (solo -> team -> triangle -> AI-adjusted)
    prob_stages = [
        ("Solo\n(David)", 0.051),
        ("+ Igor\n(CTO)", 0.08),
        ("+ Chris\n(CCO)", 0.175),
        ("Post\nPre-Mortem", 0.21),
        ("+ Triangle\nStrategy", 0.27),
        ("+ AI\nTailwinds", AI_UNICORN_PROBABILITY),
    ]
    stage_names = [s[0] for s in prob_stages]
    stage_probs = [s[1] * 100 for s in prob_stages]
    stage_colors = ["#E0E0E0", "#BDBDBD", "#90A4AE", "#42A5F5", "#1565C0", "#2E7D32"]

    bars = ax2.bar(range(len(stage_names)), stage_probs, color=stage_colors,
                    edgecolor="white", linewidth=1.5)
    for bar, val in zip(bars, stage_probs):
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                 f"{val:.0f}%", ha="center", fontsize=10, fontweight="bold")
    ax2.set_xticks(range(len(stage_names)))
    ax2.set_xticklabels(stage_names, fontsize=8)
    ax2.set_ylabel("Unicorn Probability (%)")
    ax2.set_title("Unicorn Probability Build-Up", fontsize=12, fontweight="bold")
    ax2.set_ylim(0, 55)
    # Add baseline average line
    ax2.axhline(y=5, color="#E53935", linestyle=":", linewidth=1.5, alpha=0.7)
    ax2.text(0.1, 6, "Series A avg: 5%", fontsize=8, color="#E53935")
    ax2.grid(axis="y", alpha=0.3)

    fig.suptitle("GuardSpine: AI-Adjusted Unicorn Analysis",
                 fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "ai_unicorn_path.png", bbox_inches="tight")
    plt.close(fig)

    print(f"  Generated 17 figures in {FIGURES_DIR}")


# ============================================================
# EXCEL WORKBOOK
# ============================================================

def h_font():
    return Font(bold=True, color="FFFFFF", size=11)

def h_fill():
    return PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

def apply_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = h_font()
        cell.fill = h_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

CUR = '$#,##0'
PCT = '0.0%'


def make_excel():
    wb = Workbook()

    # ========== Tab 1: TAM/SAM/SOM ==========
    ws1 = wb.active
    ws1.title = "TAM-SAM-SOM"
    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 40

    headers = ["Market Segment", "Value ($)", "Source", "Notes"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=1, column=i, value=h)
    apply_header(ws1, 1, 4)

    # Code-only TAM
    ws1.cell(row=2, column=1, value="CODE-ONLY TAM").font = Font(bold=True, size=12)
    code_items = [
        ("DevSecOps TAM 2026", DEVSECOPS_TAM, "Precedence Research"),
        ("AI Governance TAM 2026", AI_GOVERNANCE_TAM, "Precedence/Grand View"),
        ("Code Review Tools TAM 2026", CODE_REVIEW_TAM, "Estimate"),
    ]
    for ri, (label, val, src) in enumerate(code_items, 3):
        ws1.cell(row=ri, column=1, value=label)
        ws1.cell(row=ri, column=2, value=val).number_format = CUR
        ws1.cell(row=ri, column=3, value=src)
    ws1.cell(row=6, column=1, value="Code-Only TAM").font = Font(bold=True)
    ws1.cell(row=6, column=2, value="=SUM(B3:B5)").number_format = CUR

    # Multi-artifact expansion
    ws1.cell(row=8, column=1, value="MULTI-ARTIFACT EXPANSION").font = Font(bold=True, size=12)
    ws1.cell(row=8, column=4, value="AI Office thesis: code + docs + sheets + images")
    expand_items = [
        ("GRC Software TAM 2026", GRC_TAM, "Mordor Intelligence"),
        ("Document Management TAM 2026", DOC_MGMT_TAM, "Market estimate"),
        ("Digital Asset Management TAM 2026", DIGITAL_ASSET_TAM, "Market estimate"),
    ]
    for ri, (label, val, src) in enumerate(expand_items, 9):
        ws1.cell(row=ri, column=1, value=label)
        ws1.cell(row=ri, column=2, value=val).number_format = CUR
        ws1.cell(row=ri, column=3, value=src)
    ws1.cell(row=12, column=1, value="Multi-Artifact TAM").font = Font(bold=True)
    ws1.cell(row=12, column=2, value="=B6+SUM(B9:B11)").number_format = CUR

    ws1.cell(row=13, column=1, value="Governance Layer (5%)")
    ws1.cell(row=13, column=2, value="=B12*0.05").number_format = CUR
    ws1.cell(row=14, column=1, value="Governance Layer (15%)")
    ws1.cell(row=14, column=2, value="=B12*0.15").number_format = CUR

    # Bottom-up
    ws1.cell(row=16, column=1, value="BOTTOM-UP TAM").font = Font(bold=True, size=12)
    bu = [
        ("Total Developers", TOTAL_DEVS, "SlashData 2025"),
        ("PRs per Dev per Month", PRS_PER_DEV_MONTH, "minware/GitClear"),
        ("Months per Year", 12, ""),
        ("Governance Price per PR ($)", GOVERNANCE_PRICE_PER_PR, ""),
    ]
    for ri, (label, val, src) in enumerate(bu, 17):
        ws1.cell(row=ri, column=1, value=label)
        ws1.cell(row=ri, column=2, value=val)
        ws1.cell(row=ri, column=3, value=src)
    ws1.cell(row=21, column=1, value="Bottom-Up Code TAM").font = Font(bold=True)
    ws1.cell(row=21, column=2, value="=B17*B18*B19*B20").number_format = CUR

    # SAM (3-segment)
    ws1.cell(row=23, column=1, value="SERVICEABLE ADDRESSABLE MARKET (SAM)").font = Font(bold=True, size=12)
    sam_items = [
        ("GitHub Enterprise Regulated Orgs", SAM_CODE_ORGS, SAM_CODE_ACV),
        ("Enterprise Document Governance Orgs", SAM_DOC_GOV_ORGS, SAM_DOC_GOV_ACV),
        ("Cross-Artifact (Code + Docs) Orgs", SAM_CROSS_ORGS, SAM_CROSS_ACV),
    ]
    ws1.cell(row=24, column=1, value="Segment")
    ws1.cell(row=24, column=2, value="Org Count")
    ws1.cell(row=24, column=3, value="Avg ACV")
    ws1.cell(row=24, column=4, value="Segment SAM")
    apply_header(ws1, 24, 4)
    for ri, (label, orgs, acv) in enumerate(sam_items, 25):
        ws1.cell(row=ri, column=1, value=label)
        ws1.cell(row=ri, column=2, value=orgs)
        ws1.cell(row=ri, column=3, value=acv).number_format = CUR
        ws1.cell(row=ri, column=4, value=f"=B{ri}*C{ri}").number_format = CUR
    ws1.cell(row=28, column=1, value="Total SAM").font = Font(bold=True)
    ws1.cell(row=28, column=4, value="=SUM(D25:D27)").number_format = CUR

    # SOM
    ws1.cell(row=30, column=1, value="SOM PROJECTIONS").font = Font(bold=True, size=12)
    som_items = [("Year 1", SOM_Y1_PCT), ("Year 2", SOM_Y2_PCT), ("Year 3", SOM_Y3_PCT)]
    for ri, (label, pct) in enumerate(som_items, 31):
        ws1.cell(row=ri, column=1, value=label)
        ws1.cell(row=ri, column=2, value=pct).number_format = PCT
        ws1.cell(row=ri, column=3, value=f"=D28*B{ri}").number_format = CUR
        ws1.cell(row=ri, column=4, value="SOM")

    # ========== Tab 2: Unit Economics ==========
    ws2 = wb.create_sheet("Unit Economics")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 22

    # L0-L4 Risk Tier Model
    ws2.cell(row=1, column=1, value="L0-L4 RISK TIER MODEL").font = Font(bold=True, size=12)
    ws2.cell(row=1, column=4, value="Risk tiers = governance behavior, not just cost")
    tier_headers = ["Tier", "Share of Changes", "COGS/Change ($)", "Revenue/Change ($)"]
    for i, h in enumerate(tier_headers, 1):
        ws2.cell(row=2, column=i, value=h)
    apply_header(ws2, 2, 4)
    for ri, tier in enumerate(TIER_DIST.keys(), 3):
        ws2.cell(row=ri, column=1, value=f"{tier}: {TIER_NAMES[tier]}")
        ws2.cell(row=ri, column=2, value=TIER_DIST[tier]).number_format = PCT
        ws2.cell(row=ri, column=3, value=TIER_COGS[tier]).number_format = '$#,##0.00'
        ws2.cell(row=ri, column=4, value=TIER_REV[tier]).number_format = '$#,##0.00'

    r = 8
    ws2.cell(row=r, column=1, value="Blended COGS/Change").font = Font(bold=True)
    ws2.cell(row=r, column=3, value="=SUMPRODUCT(B3:B7,C3:C7)").number_format = '$#,##0.000'
    ws2.cell(row=r+1, column=1, value="Blended Revenue/Change").font = Font(bold=True)
    ws2.cell(row=r+1, column=4, value="=SUMPRODUCT(B3:B7,D3:D7)").number_format = '$#,##0.000'
    ws2.cell(row=r+2, column=1, value="Per-Change Gross Margin").font = Font(bold=True)
    ws2.cell(row=r+2, column=2, value=f"=(D{r+1}-C{r})/(D{r+1})").number_format = PCT

    # Open-Core Model
    r = 12
    ws2.cell(row=r, column=1, value="OPEN-CORE MODEL (Linux vs Red Hat)").font = Font(bold=True, size=12)
    ws2.cell(row=r+1, column=1, value="BYOK: Users bring their own LLM API keys. GuardSpine pays ZERO API costs.")
    ws2.cell(row=r+2, column=1, value="Free: Spec + verifier + CodeGuard Action + PII-Shield + YAML rubrics")
    ws2.cell(row=r+3, column=1, value="Paid: UI + management coordination + enterprise features + cognitive attestation")

    # Funnel metrics
    r2 = r + 5
    ws2.cell(row=r2, column=1, value="CONVERSION FUNNEL").font = Font(bold=True, size=12)
    ws2.cell(row=r2+1, column=1, value="Free-to-Pro Conversion Rate")
    ws2.cell(row=r2+1, column=2, value=FREE_TO_PRO_CONVERSION).number_format = PCT
    ws2.cell(row=r2+1, column=3, value="Industry: 1-5% for open-core")
    ws2.cell(row=r2+2, column=1, value="Pro-to-Business Expansion (12mo)")
    ws2.cell(row=r2+2, column=2, value=PRO_TO_BUSINESS_EXPANSION).number_format = PCT
    ws2.cell(row=r2+3, column=1, value="Business-to-Enterprise (18mo)")
    ws2.cell(row=r2+3, column=2, value=BUSINESS_TO_ENTERPRISE).number_format = PCT

    # Per-Customer Model (4 tiers)
    r3 = r2 + 5
    ws2.cell(row=r3, column=1, value="PER-CUSTOMER MODEL (MONTHLY, BYOK)").font = Font(bold=True, size=12)
    # Only show paid tiers in the detailed model (Free has $0 everything)
    paid_tiers = {k: v for k, v in TIERS.items() if v["monthly"] > 0}
    tier_names_list = list(paid_tiers.keys())
    ncols = len(tier_names_list) + 1
    cust_headers = ["Metric"] + tier_names_list
    for i, h in enumerate(cust_headers, 1):
        ws2.cell(row=r3+1, column=i, value=h)
    apply_header(ws2, r3+1, ncols)

    metrics = [
        ("Monthly Subscription ($)", "monthly"),
        ("Changes per Month", "changes_mo"),
        ("Hosting/Infrastructure ($)", "hosting"),
        ("Cognitive Attestation License ($)", "cognitive_license"),
        ("Support Allocation ($)", "support"),
    ]
    start_r = r3 + 2
    for mi, (label, key) in enumerate(metrics):
        row = start_r + mi
        ws2.cell(row=row, column=1, value=label)
        for ti, tname in enumerate(tier_names_list, 2):
            ws2.cell(row=row, column=ti, value=paid_tiers[tname][key]).number_format = CUR

    # Total COGS formula (only hosting + cognitive + support, NO API costs)
    cogs_r = start_r + len(metrics)
    ws2.cell(row=cogs_r, column=1, value="Total COGS ($) [BYOK: zero API cost]").font = Font(bold=True)
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        # Sum hosting + cognitive + support (rows start_r+2 through start_r+4)
        ws2.cell(row=cogs_r, column=ti,
                 value=f"=SUM({col}{start_r+2}:{col}{cogs_r-1})").number_format = CUR

    # Gross profit
    gp_r = cogs_r + 1
    ws2.cell(row=gp_r, column=1, value="Gross Profit ($)").font = Font(bold=True)
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        ws2.cell(row=gp_r, column=ti,
                 value=f"={col}{start_r}-{col}{cogs_r}").number_format = CUR

    # Gross margin
    gm_r = gp_r + 1
    ws2.cell(row=gm_r, column=1, value="Gross Margin (%)").font = Font(bold=True)
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        ws2.cell(row=gm_r, column=ti,
                 value=f"=IF({col}{start_r}>0,{col}{gp_r}/{col}{start_r},0)").number_format = PCT

    # LTV
    ltv_r = gm_r + 2
    ws2.cell(row=ltv_r, column=1, value="LIFETIME VALUE").font = Font(bold=True, size=12)
    ws2.cell(row=ltv_r+1, column=1, value="Customer Lifetime (months)")
    for ti in range(2, ncols + 1):
        ws2.cell(row=ltv_r+1, column=ti, value=CUSTOMER_LIFETIME_MONTHS)
    ws2.cell(row=ltv_r+2, column=1, value="Net Revenue Retention")
    for ti, tname in enumerate(tier_names_list, 2):
        ws2.cell(row=ltv_r+2, column=ti, value=paid_tiers[tname]["nrr"]).number_format = PCT
    ws2.cell(row=ltv_r+3, column=1, value="LTV ($)").font = Font(bold=True)
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        ws2.cell(row=ltv_r+3, column=ti,
                 value=f"={col}{start_r}*{col}{gm_r}*{col}{ltv_r+1}*{col}{ltv_r+2}").number_format = CUR

    # CAC
    cac_r = ltv_r + 5
    ws2.cell(row=cac_r, column=1, value="CAC & PAYBACK").font = Font(bold=True, size=12)
    ws2.cell(row=cac_r+1, column=1, value="Target CAC ($)")
    for ti, tname in enumerate(tier_names_list, 2):
        ws2.cell(row=cac_r+1, column=ti, value=paid_tiers[tname]["cac"]).number_format = CUR
    ws2.cell(row=cac_r+2, column=1, value="LTV/CAC Ratio").font = Font(bold=True)
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        ws2.cell(row=cac_r+2, column=ti,
                 value=f"={col}{ltv_r+3}/{col}{cac_r+1}").number_format = "0.0"
    ws2.cell(row=cac_r+3, column=1, value="CAC Payback (months)")
    for ti in range(2, ncols + 1):
        col = get_column_letter(ti)
        ws2.cell(row=cac_r+3, column=ti,
                 value=f"={col}{cac_r+1}/{col}{gp_r}").number_format = "0.0"

    # ========== Tab 3: Revenue Scenarios ==========
    ws3 = wb.create_sheet("Revenue Scenarios")
    ws3.column_dimensions["A"].width = 22
    for c in range(2, 8):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    ws3.cell(row=1, column=1, value="Scenario Parameters").font = Font(bold=True, size=12)
    for i, h in enumerate(["Parameter", "Bear", "Base", "Bull"], 1):
        ws3.cell(row=2, column=i, value=h)
    apply_header(ws3, 2, 4)

    param_rows = [
        ("Starting Customers", 5, 8, 15),
        ("Monthly Growth Rate", 0.08, 0.15, 0.20),
        ("Avg ACV ($)", 18000, 30000, 50000),
    ]
    for ri, (label, *vals) in enumerate(param_rows, 3):
        ws3.cell(row=ri, column=1, value=label)
        for ci, v in enumerate(vals, 2):
            ws3.cell(row=ri, column=ci, value=v)
            if "Rate" in label:
                ws3.cell(row=ri, column=ci).number_format = PCT
            elif "$" in label:
                ws3.cell(row=ri, column=ci).number_format = CUR

    # Monthly projection
    r = 7
    ws3.cell(row=r, column=1, value="Monthly Projection").font = Font(bold=True, size=12)
    proj_h = ["Month", "Bear Cust", "Bear MRR ($)", "Base Cust", "Base MRR ($)",
              "Bull Cust", "Bull MRR ($)"]
    for i, h in enumerate(proj_h, 1):
        ws3.cell(row=r+1, column=i, value=h)
    apply_header(ws3, r+1, 7)

    ps = r + 2  # projection start
    for m in range(0, 37):
        row = ps + m
        ws3.cell(row=row, column=1, value=m)
        if m == 0:
            ws3.cell(row=row, column=2, value="=B3")
            ws3.cell(row=row, column=3, value=f"=B{row}*B5/12")
            ws3.cell(row=row, column=4, value="=C3")
            ws3.cell(row=row, column=5, value=f"=D{row}*C5/12")
            ws3.cell(row=row, column=6, value="=D3")
            ws3.cell(row=row, column=7, value=f"=F{row}*D5/12")
        else:
            p = row - 1
            ws3.cell(row=row, column=2, value=f"=B{p}*(1+B4)")
            ws3.cell(row=row, column=3, value=f"=B{row}*B5/12")
            ws3.cell(row=row, column=4, value=f"=D{p}*(1+C4)")
            ws3.cell(row=row, column=5, value=f"=D{row}*C5/12")
            ws3.cell(row=row, column=6, value=f"=F{p}*(1+D4)")
            ws3.cell(row=row, column=7, value=f"=F{row}*D5/12")
        for c in [2, 4, 6]:
            ws3.cell(row=row, column=c).number_format = "0.0"
        for c in [3, 5, 7]:
            ws3.cell(row=row, column=c).number_format = CUR

    # ARR summary
    sr = ps + 37 + 1
    ws3.cell(row=sr, column=1, value="Year 1 ARR (MRR*12)").font = Font(bold=True)
    y1 = ps + 12
    ws3.cell(row=sr, column=2, value=f"=C{y1}*12").number_format = CUR
    ws3.cell(row=sr, column=3, value="Bear")
    ws3.cell(row=sr, column=4, value=f"=E{y1}*12").number_format = CUR
    ws3.cell(row=sr, column=5, value="Base")
    ws3.cell(row=sr, column=6, value=f"=G{y1}*12").number_format = CUR
    ws3.cell(row=sr, column=7, value="Bull")

    # ========== Tab 4: Fee Cascade ==========
    ws4 = wb.create_sheet("Fee Cascade")
    ws4.column_dimensions["A"].width = 12
    for c in range(2, 10):
        ws4.column_dimensions[get_column_letter(c)].width = 16

    ws4.cell(row=1, column=1, value="Fee Cascade Model").font = Font(bold=True, size=12)
    ws4.cell(row=2, column=1, value="Adoption rate (r)")
    ws4.cell(row=2, column=2, value=ADOPTION_RATE)
    ws4.cell(row=3, column=1, value="Inflection (t0)")
    ws4.cell(row=3, column=2, value=INFLECTION_MONTH)
    ws4.cell(row=4, column=1, value="Calibration")
    ws4.cell(row=4, column=2, value="KPMG 14% in ~12mo => k=-ln(0.86)/cumsum(A,0..12)=0.0747")

    sector_names = list(CASCADE_SECTORS.keys())
    ws4.cell(row=6, column=1, value="Sector")
    ws4.cell(row=6, column=2, value="k (decay)")
    ws4.cell(row=6, column=3, value="Lag (months)")
    apply_header(ws4, 6, 3)
    for i, name in enumerate(sector_names):
        ws4.cell(row=7 + i, column=1, value=name)
        ws4.cell(row=7 + i, column=2, value=CASCADE_SECTORS[name]["k"])
        ws4.cell(row=7 + i, column=3, value=CASCADE_SECTORS[name]["lag"])

    ws4.cell(row=13, column=1, value="Governance Premium").font = Font(bold=True, size=12)
    ws4.cell(row=14, column=1, value="G0 (current)")
    ws4.cell(row=14, column=2, value=G0).number_format = PCT
    ws4.cell(row=15, column=1, value="Gmax (ceiling)")
    ws4.cell(row=15, column=2, value=GMAX).number_format = PCT
    ws4.cell(row=16, column=1, value="Lambda")
    ws4.cell(row=16, column=2, value=LAMBDA_GOV)

    # Monthly table (computed -- logistic cumsum not possible in basic Excel)
    r = 18
    ws4.cell(row=r, column=1, value="Monthly Fee Levels & Governance Premium").font = Font(bold=True, size=12)
    fee_h = ["Month", "A(t)"] + sector_names + ["Gov Premium %"]
    for i, h in enumerate(fee_h, 1):
        ws4.cell(row=r+1, column=i, value=h)
    apply_header(ws4, r+1, len(fee_h))

    for m in range(0, 37):
        row = r + 2 + m
        ws4.cell(row=row, column=1, value=m)
        ws4.cell(row=row, column=2, value=round(adoption_curve(m), 4))
        for si, name in enumerate(sector_names):
            p = CASCADE_SECTORS[name]
            ws4.cell(row=row, column=3+si, value=round(fee_level(m, p["k"], p["lag"]), 4)).number_format = PCT
        ws4.cell(row=row, column=3+len(sector_names),
                 value=round(governance_premium(m), 4)).number_format = PCT

    nr = r + 39
    ws4.cell(row=nr, column=1, value="Formulas: P(t)=exp(-k*cumsum(A(s))), A(t)=1/(1+exp(-r*(t-t0))), G(t)=G0+(Gmax-G0)*(1-exp(-lambda*t))")

    # ========== Tab 5: Competitive Scorecard ==========
    ws5 = wb.create_sheet("Competitive Scorecard")
    ws5.column_dimensions["A"].width = 22
    for c in range(2, 9):
        ws5.column_dimensions[get_column_letter(c)].width = 16

    # Moat scorecard
    ws5.cell(row=1, column=1, value="MOAT SCORECARD (Weighted)").font = Font(bold=True, size=12)
    moat_h = ["Competitor", "Data Moat (3x)", "Network (1x)", "Regulatory (2x)",
              "Tech Diff (1x)", "Switching (1x)", "Weighted Score"]
    for i, h in enumerate(moat_h, 1):
        ws5.cell(row=2, column=i, value=h)
    apply_header(ws5, 2, 7)

    for ri, (comp, scores) in enumerate(MOAT_SCORES.items(), 3):
        ws5.cell(row=ri, column=1, value=comp)
        dims = ["data_moat", "network", "regulatory", "tech_diff", "switching"]
        for ci, d in enumerate(dims, 2):
            ws5.cell(row=ri, column=ci, value=scores[d])
        ws5.cell(row=ri, column=7,
                 value=f"=(B{ri}*3+C{ri}*1+D{ri}*2+E{ri}*1+F{ri}*1)/8").number_format = "0.00"
    for c in range(1, 8):
        ws5.cell(row=3, column=c).font = Font(bold=True, color="1E88E5")

    # MECE 9-dimension matrix
    mr = len(MOAT_SCORES) + 5
    ws5.cell(row=mr, column=1, value="MECE 9-DIMENSION MATRIX").font = Font(bold=True, size=12)
    ws5.cell(row=mr+1, column=1, value="1=Yes, 0.5=Partial, 0=No")
    mece_dims = ["code", "docs", "sheets", "images", "ai_prov", "risk_gate", "evidence", "diffs", "stop_line"]
    mece_labels = ["Code", "Docs", "Sheets", "Images", "AI Prov", "Risk Gate", "Evidence", "Diffs", "Stop-Line"]
    mece_h2 = ["Competitor"] + mece_labels + ["Total"]
    for i, h in enumerate(mece_h2, 1):
        ws5.cell(row=mr+2, column=i, value=h)
    apply_header(ws5, mr+2, len(mece_h2))

    for ri, (comp, scores) in enumerate(MECE_MATRIX.items(), mr+3):
        ws5.cell(row=ri, column=1, value=comp)
        for ci, d in enumerate(mece_dims, 2):
            ws5.cell(row=ri, column=ci, value=scores[d])
        # Total
        ws5.cell(row=ri, column=len(mece_dims)+2,
                 value=f"=SUM(B{ri}:{get_column_letter(len(mece_dims)+1)}{ri})").number_format = "0.0"
    for c in range(1, len(mece_dims)+3):
        ws5.cell(row=mr+3, column=c).font = Font(bold=True, color="1E88E5")

    # ========== Tab 6: Acquisition & Dilution ==========
    ws6 = wb.create_sheet("Acquisition & Dilution")
    ws6.column_dimensions["A"].width = 28
    for c in range(2, 10):
        ws6.column_dimensions[get_column_letter(c)].width = 18

    # Section 0: Bootstrap vs VC Path Comparison (PRIMARY)
    ws6.cell(row=1, column=1, value="CAPITAL STRATEGY: BOOTSTRAP vs VC").font = Font(bold=True, size=14)
    ws6.cell(row=1, column=5, value="Bootstrap = PRIMARY. VC = comparison only.")

    pc_h = ["Metric", "Bootstrap (Angel)", "VC 2-Round", "VC 4-Round"]
    for i, h in enumerate(pc_h, 1):
        ws6.cell(row=2, column=i, value=h)
    apply_header(ws6, 2, 4)

    path_comp, _ = compute_path_comparison()
    path_list = list(path_comp.items())
    # Row 3: Total raised
    ws6.cell(row=3, column=1, value="Total Raised")
    for ci, (pname, pdata) in enumerate(path_list, 2):
        ws6.cell(row=3, column=ci, value=pdata["total_raised"]).number_format = CUR
    # Row 4: David ownership
    ws6.cell(row=4, column=1, value="David Ownership")
    for ci, (pname, pdata) in enumerate(path_list, 2):
        ws6.cell(row=4, column=ci, value=pdata["ownership"]).number_format = PCT
    # Row 5: Exit needed for $100M
    ws6.cell(row=5, column=1, value="Exit for $100M post-tax")
    for ci, (pname, pdata) in enumerate(path_list, 2):
        needed = PERSONAL_TARGET / (pdata["ownership"] * (1 - TAX_RATE))
        ws6.cell(row=5, column=ci, value=needed).number_format = CUR
    # Rows 6-9: Take-home at exit valuations
    for ri, ev in enumerate([305e6, 400e6, 600e6, 1e9], 6):
        elabel = f"${ev/1e6:.0f}M" if ev < 1e9 else "$1B"
        ws6.cell(row=ri, column=1, value=f"Take-home @ {elabel} exit")
        for ci, (pname, pdata) in enumerate(path_list, 2):
            ws6.cell(row=ri, column=ci, value=pdata["ownership"] * ev * (1-TAX_RATE)).number_format = CUR
    # Highlight bootstrap column
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    for r in range(3, 10):
        ws6.cell(row=r, column=2).fill = green_fill

    # Section 1: VC Dilution Detail (reference only)
    sr_start = 12
    ws6.cell(row=sr_start, column=1, value="VC 4-ROUND DILUTION DETAIL (reference)").font = Font(bold=True, size=12)
    ws6.cell(row=sr_start, column=5, value="Shown for comparison -- NOT the plan")
    dil_h = ["Round", "Raise", "Pre-Money Val.", "Dilution %", "David Before",
             "David After", "Equity Lost", "Cost @ $400M Exit", "Cost @ $1B Exit"]
    for i, h in enumerate(dil_h, 1):
        ws6.cell(row=sr_start+1, column=i, value=h)
    apply_header(ws6, sr_start+1, 9)

    dil_results, final_own = compute_dilution_waterfall(
        rounds=VC_PATH, target_exits=[400e6, 600e6, 1.0e9, 2.0e9])
    for ri, r in enumerate(dil_results, sr_start+2):
        ws6.cell(row=ri, column=1, value=r["round"])
        ws6.cell(row=ri, column=2, value=r["raise"]).number_format = CUR
        ws6.cell(row=ri, column=3, value=r["pre_money"]).number_format = CUR
        ws6.cell(row=ri, column=4, value=r["dilution_pct"] / 100).number_format = PCT
        ws6.cell(row=ri, column=5, value=r["david_before"] / 100).number_format = PCT
        ws6.cell(row=ri, column=6, value=r["david_after"] / 100).number_format = PCT
        ws6.cell(row=ri, column=7, value=r["equity_lost_pct"] / 100).number_format = PCT
        ws6.cell(row=ri, column=8, value=r["costs"][400e6]["post_tax"]).number_format = CUR
        ws6.cell(row=ri, column=9, value=r["costs"][1.0e9]["post_tax"]).number_format = CUR

    sr = sr_start + 2 + len(dil_results)
    ws6.cell(row=sr, column=1, value="TOTAL").font = Font(bold=True)
    ws6.cell(row=sr, column=2, value=f"=SUM(B{sr_start+2}:B{sr-1})").number_format = CUR
    ws6.cell(row=sr, column=7, value=(DAVID_EQUITY - final_own)).number_format = PCT
    total_400 = sum(r["costs"][400e6]["post_tax"] for r in dil_results)
    total_1b = sum(r["costs"][1.0e9]["post_tax"] for r in dil_results)
    ws6.cell(row=sr, column=8, value=total_400).number_format = CUR
    ws6.cell(row=sr, column=9, value=total_1b).number_format = CUR
    ws6.cell(row=sr+1, column=1, value="Final ownership (David)")
    ws6.cell(row=sr+1, column=6, value=final_own).number_format = PCT
    ws6.cell(row=sr+2, column=1, value="David post-tax at $1B exit (VC path)")
    ws6.cell(row=sr+2, column=8, value=final_own * 1e9 * (1 - TAX_RATE)).number_format = CUR

    # Section 2: Acquisition Price Trajectory
    ar = sr + 4
    ws6.cell(row=ar, column=1, value="ACQUISITION PRICE TRAJECTORY").font = Font(bold=True, size=12)
    ws6.cell(row=ar, column=4, value="Market val * strategic premium (1.2-1.5x)")
    acq_h = ["Year", "Bear ARR", "Bear Acq Price", "Base ARR", "Base Acq Price",
             "Bull ARR", "Bull Acq Price"]
    for i, h in enumerate(acq_h, 1):
        ws6.cell(row=ar+1, column=i, value=h)
    apply_header(ws6, ar+1, 7)

    acq_traj = compute_acquisition_trajectory()
    for y in range(1, 8):
        row = ar + 1 + y
        ws6.cell(row=row, column=1, value=f"Year {y}")
        for si, name in enumerate(["Bear", "Base", "Bull"]):
            if y < len(acq_traj[name]):
                r = acq_traj[name][y]
                ws6.cell(row=row, column=2+si*2, value=r["arr"]).number_format = CUR
                ws6.cell(row=row, column=3+si*2, value=r["acq_price"]).number_format = CUR

    # Section 3: $1B Requirements
    br = ar + 10
    ws6.cell(row=br, column=1, value="$1B+ ACQUISITION REQUIREMENTS").font = Font(bold=True, size=12)
    ws6.cell(row=br+1, column=1, value="Minimum ARR")
    ws6.cell(row=br+1, column=2, value=BILLION_REQUIREMENTS["min_arr_floor"]).number_format = CUR
    ws6.cell(row=br+1, column=3, value="Base case: Year 4")
    ws6.cell(row=br+2, column=1, value="Strong-confidence ARR")
    ws6.cell(row=br+2, column=2, value=BILLION_REQUIREMENTS["min_arr_strong"]).number_format = CUR
    ws6.cell(row=br+2, column=3, value="Base case: Year 5")
    ws6.cell(row=br+3, column=1, value="Min YoY Growth")
    ws6.cell(row=br+3, column=2, value=BILLION_REQUIREMENTS["min_yoy_growth"]).number_format = PCT
    ws6.cell(row=br+4, column=1, value="Min Gross Margin")
    ws6.cell(row=br+4, column=2, value=BILLION_REQUIREMENTS["min_gross_margin"]).number_format = PCT
    ws6.cell(row=br+5, column=1, value="Min Enterprise Logos")
    ws6.cell(row=br+5, column=2, value=BILLION_REQUIREMENTS["min_enterprise_logos"])

    # Per-acquirer triggers
    pr = br + 7
    ws6.cell(row=pr, column=1, value="PER-ACQUIRER $1B TRIGGERS").font = Font(bold=True, size=12)
    pa_h = ["Acquirer", "Trigger ARR", "Likely Price", "Rationale"]
    for i, h in enumerate(pa_h, 1):
        ws6.cell(row=pr+1, column=i, value=h)
    apply_header(ws6, pr+1, 4)
    for ri, (acq, data) in enumerate(BILLION_REQUIREMENTS["per_acquirer"].items(), pr+2):
        ws6.cell(row=ri, column=1, value=acq)
        ws6.cell(row=ri, column=2, value=data["trigger_arr"]).number_format = CUR
        ws6.cell(row=ri, column=3, value=data["likely_price_at_trigger"])
        ws6.cell(row=ri, column=4, value=data["rationale"])

    # ========== Tab 7: Speed Execution Plan ==========
    ws7 = wb.create_sheet("Speed Plan")
    ws7.column_dimensions["A"].width = 22
    for c in range(2, 8):
        ws7.column_dimensions[get_column_letter(c)].width = 20

    # Section 1: Timeline
    ws7.cell(row=1, column=1, value="SPEED-TO-$100M EXECUTION PLAN").font = Font(bold=True, size=14)
    ws7.cell(row=1, column=4, value=f"Target: ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit in {SPEED_BEST_CASE_MONTHS}-{SPEED_REALISTIC_MONTHS} months")
    sp_h = ["Phase", "Months", "Key Actions", "Target ARR", "Controlled?"]
    for i, h in enumerate(sp_h, 1):
        ws7.cell(row=2, column=i, value=h)
    apply_header(ws7, 2, 5)

    for ri, phase in enumerate(SPEED_TIMELINE, 3):
        ws7.cell(row=ri, column=1, value=phase["phase"]).font = Font(bold=True)
        ws7.cell(row=ri, column=2, value=f'{phase["months"][0]}-{phase["months"][1]}')
        ws7.cell(row=ri, column=3, value=" | ".join(phase["actions"][:3]))
        ws7.cell(row=ri, column=4, value=phase["arr"]).number_format = CUR
        ws7.cell(row=ri, column=5, value="Yes" if phase["controlled"] else "Partly")

    # Section 2: Speed Killers
    kr = 3 + len(SPEED_TIMELINE) + 2
    ws7.cell(row=kr, column=1, value="SPEED KILLERS (avoid these)").font = Font(bold=True, size=12,
                                                                                  color="C62828")
    kh = ["Killer", "Cost (months)", "Description", "Fix"]
    for i, h in enumerate(kh, 1):
        ws7.cell(row=kr+1, column=i, value=h)
    apply_header(ws7, kr+1, 4)

    for ki, killer in enumerate(SPEED_KILLERS, kr+2):
        ws7.cell(row=ki, column=1, value=killer["name"]).font = Font(bold=True, color="C62828")
        ws7.cell(row=ki, column=2, value=killer["cost_months"])
        ws7.cell(row=ki, column=3, value=killer["description"])
        ws7.cell(row=ki, column=4, value=killer["fix"])

    # Section 3: Controllable vs Uncontrollable
    cr = kr + 2 + len(SPEED_KILLERS) + 2
    ws7.cell(row=cr, column=1, value="WHAT YOU CONTROL vs DON'T").font = Font(bold=True, size=12)
    ctrl_h = ["You Control", "You Don't Control"]
    for i, h in enumerate(ctrl_h, 1):
        ws7.cell(row=cr+1, column=i, value=h)
    apply_header(ws7, cr+1, 2)

    controls = [
        ("Product quality (FP/FN rates)", "Enterprise procurement speed"),
        ("Pricing discipline (anchor high)", "Acquirer timing appetite"),
        ("Which customers to approach first", "Market sentiment at exit"),
        ("Open-source distribution speed", "Competitor responses"),
        ("Patent filing timing", "Regulatory timeline"),
        ("When to start acquirer conversations", "How fast acquirers move"),
        ("Whether to accept sub-$305M offers", "Whether $305M+ offer arrives"),
    ]
    for ci, (ctrl, nctrl) in enumerate(controls, cr+2):
        ws7.cell(row=ci, column=1, value=ctrl)
        ws7.cell(row=ci, column=2, value=nctrl)
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws7.cell(row=ci, column=1).fill = green_fill
        ws7.cell(row=ci, column=2).fill = red_fill

    # ========== Tab 8: AI-Adjusted Projections ==========
    ws8 = wb.create_sheet("AI-Adjusted Math")
    ws8.column_dimensions["A"].width = 30
    for c in range(2, 8):
        ws8.column_dimensions[get_column_letter(c)].width = 20

    # Section 1: AI Trajectory Data (sourced)
    ws8.cell(row=1, column=1, value="AI TRAJECTORY: SOURCED DATA (Feb 2026)").font = Font(bold=True, size=14)
    ai_data_h = ["Data Point", "Value", "Source"]
    for i, h in enumerate(ai_data_h, 1):
        ws8.cell(row=2, column=i, value=h)
    apply_header(ws8, 2, 3)

    ai_data_rows = [
        ("METR task doubling time", "7 months", "METR / Epoch AI (Feb 2026)"),
        ("Current task horizon", "50 minutes", "METR benchmark"),
        ("Week-long task projection", "Late 2026-2027", "METR extrapolation"),
        ("Token cost decline", "10x per year", "Epoch AI cost index"),
        ("Cost per M tokens (2025)", "$0.40", "Epoch AI"),
        ("Cost per M tokens (2026e)", "$0.04", "Epoch AI extrapolation"),
        ("SWE-bench Verified solve rate", "75%", "SWE-bench (Feb 2026)"),
        ("SWE-bench Pro (Claude Opus 4.5)", "45.89%", "SWE-bench Pro (Jan 2026)"),
        ("AI-generated code share", "41%", "GitHub / Stack Overflow (Jan 2026)"),
        ("Productivity gain (real)", "20-30%", "Multiple studies, GitHub Copilot"),
        ("Enterprise agentic AI adoption", "79%", "Capgemini Research (Feb 2026)"),
        ("Agentic AI surge (YoY)", "340%", "Capgemini Research"),
        ("Apps agentic by EOY 2026", "40%", "Capgemini forecast"),
        ("EU AI Act enforcement", "August 2, 2026", "Official Journal of the EU"),
        ("EU AI Act max fine", "7% of global turnover", "EU AI Act Article 99"),
    ]
    for ri, (dp, val, src) in enumerate(ai_data_rows, 3):
        ws8.cell(row=ri, column=1, value=dp)
        ws8.cell(row=ri, column=2, value=val).font = Font(bold=True)
        ws8.cell(row=ri, column=3, value=src)

    # Section 2: 5 Effects on GuardSpine
    eff_start = 3 + len(ai_data_rows) + 2
    ws8.cell(row=eff_start, column=1, value="5 EFFECTS ON GUARDSPINE MATH").font = Font(bold=True, size=12)
    eff_h = ["Effect", "Baseline", "AI-Adjusted", "Mechanism"]
    for i, h in enumerate(eff_h, 1):
        ws8.cell(row=eff_start+1, column=i, value=h)
    apply_header(ws8, eff_start+1, 4)

    for ei, eff in enumerate(AI_EFFECTS, eff_start+2):
        ws8.cell(row=ei, column=1, value=eff["effect"]).font = Font(bold=True)
        ws8.cell(row=ei, column=2, value=eff.get("baseline", ""))
        ws8.cell(row=ei, column=3, value=eff.get("adjusted", "")).font = Font(bold=True, color="1565C0")
        ws8.cell(row=ei, column=4, value=eff["mechanism"])

    # Section 3: AI-Adjusted Timeline Comparison
    tl_start = eff_start + 2 + len(AI_EFFECTS) + 2
    ws8.cell(row=tl_start, column=1, value="AI-ADJUSTED TIMELINE").font = Font(bold=True, size=12)
    ws8.cell(row=tl_start, column=3,
             value=f"Best: {AI_SPEED_BEST_CASE_MONTHS}mo (was {SPEED_BEST_CASE_MONTHS}mo) | "
                   f"Realistic: {AI_SPEED_REALISTIC_MONTHS}mo (was {SPEED_REALISTIC_MONTHS}mo)")
    tl_h = ["Phase", "Baseline Months", "AI-Adjusted Months", "ARR (AI)", "Note"]
    for i, h in enumerate(tl_h, 1):
        ws8.cell(row=tl_start+1, column=i, value=h)
    apply_header(ws8, tl_start+1, 5)

    for ti, (ai_phase, base_phase) in enumerate(zip(AI_SPEED_TIMELINE, SPEED_TIMELINE), tl_start+2):
        ws8.cell(row=ti, column=1, value=ai_phase["phase"]).font = Font(bold=True)
        ws8.cell(row=ti, column=2, value=f'{base_phase["months"][0]}-{base_phase["months"][1]}')
        ws8.cell(row=ti, column=3, value=f'{ai_phase["months"][0]}-{ai_phase["months"][1]}')
        ws8.cell(row=ti, column=3).font = Font(bold=True, color="1565C0")
        ws8.cell(row=ti, column=4, value=ai_phase["arr"]).number_format = CUR
        ws8.cell(row=ti, column=5, value=ai_phase["note"])

    # Section 4: Exit math at different multiples
    ex_start = tl_start + 2 + len(AI_SPEED_TIMELINE) + 2
    ws8.cell(row=ex_start, column=1,
             value=f"ARR NEEDED FOR ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M EXIT").font = Font(bold=True, size=12)
    ex_h = ["Revenue Multiple", "ARR Needed", "Achievable?", "Zone"]
    for i, h in enumerate(ex_h, 1):
        ws8.cell(row=ex_start+1, column=i, value=h)
    apply_header(ws8, ex_start+1, 4)

    for mi, mult in enumerate([15, 20, 25, 30, 35], ex_start+2):
        arr_req = BOOTSTRAP_EXIT_NEEDED / mult
        zone = "GREEN" if arr_req <= 12e6 else "YELLOW" if arr_req <= 16e6 else "RED"
        ws8.cell(row=mi, column=1, value=f"{mult}x")
        ws8.cell(row=mi, column=2, value=arr_req).number_format = CUR
        ws8.cell(row=mi, column=3, value="Yes" if arr_req <= 15e6 else "Stretch")
        ws8.cell(row=mi, column=4, value=zone)
        fill_color = "E8F5E9" if zone == "GREEN" else "FFF8E1" if zone == "YELLOW" else "FFEBEE"
        for col in range(1, 5):
            ws8.cell(row=mi, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Section 5: Key insight box
    ins_start = ex_start + 2 + 5 + 2
    ws8.cell(row=ins_start, column=1, value="KEY INSIGHT").font = Font(bold=True, size=12, color="1565C0")
    ws8.cell(row=ins_start+1, column=1,
             value=f"At AI-adjusted multiples (25-35x), ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit "
                   f"requires only $9-12M ARR instead of $15-20M.")
    ws8.cell(row=ins_start+2, column=1,
             value=f"This compresses the timeline by 4-6 months: {AI_SPEED_BEST_CASE_MONTHS} months "
                   f"best case (was {SPEED_BEST_CASE_MONTHS}), "
                   f"{AI_SPEED_REALISTIC_MONTHS} realistic (was {SPEED_REALISTIC_MONTHS}).")
    ws8.cell(row=ins_start+3, column=1,
             value="EU AI Act (Aug 2, 2026) creates a hard deadline forcing acquirers to move. "
                   "GuardSpine must be in market BEFORE this date.")
    for r in range(ins_start+1, ins_start+4):
        ws8.cell(row=r, column=1).font = Font(bold=True)

    # ========== Tab 9: AI-Adjusted Unicorn Path ==========
    ws9 = wb.create_sheet("AI Unicorn Path")
    ws9.column_dimensions["A"].width = 22
    for c in range(2, 10):
        ws9.column_dimensions[get_column_letter(c)].width = 16

    ws9.cell(row=1, column=1, value="AI-ADJUSTED PATH TO UNICORN").font = Font(bold=True, size=14)
    ws9.cell(row=1, column=4,
             value=f"Unicorn probability: {AI_UNICORN_PROBABILITY*100:.0f}% "
                   f"(up from 27% baseline)")

    # Section 1: Baseline vs AI-adjusted unicorn timelines
    ai_traj = compute_ai_acquisition_trajectory()
    baseline_traj = compute_acquisition_trajectory()

    row = 3
    for scenario_pair in [("Bear", "Bear (AI)"), ("Base", "Base (AI)"), ("Bull", "Bull (AI)")]:
        bname, aname = scenario_pair
        ws9.cell(row=row, column=1,
                 value=f"{bname}: BASELINE vs AI-ADJUSTED").font = Font(bold=True, size=11)
        row += 1
        uh = ["Year", "Baseline ARR", "Baseline Acq", "AI ARR", "AI Acq", "AI David Post-Tax"]
        for i, h in enumerate(uh, 1):
            ws9.cell(row=row, column=i, value=h)
        apply_header(ws9, row, len(uh))
        row += 1

        bpath = baseline_traj[bname]
        apath = ai_traj[aname]
        for yr in range(1, 8):
            b_entry = bpath[yr] if yr < len(bpath) else {"arr": 0, "acq_price": 0}
            a_entry = apath[yr] if yr < len(apath) else {"arr": 0, "acq_price": 0}
            david_take = BOOTSTRAP_OWNERSHIP * a_entry["acq_price"] * (1 - TAX_RATE)
            ws9.cell(row=row, column=1, value=f"Y{yr}")
            ws9.cell(row=row, column=2, value=b_entry["arr"]).number_format = CUR
            ws9.cell(row=row, column=3, value=b_entry["acq_price"]).number_format = CUR
            ws9.cell(row=row, column=4, value=a_entry["arr"]).number_format = CUR
            ws9.cell(row=row, column=4).font = Font(bold=True, color="1565C0")
            ws9.cell(row=row, column=5, value=a_entry["acq_price"]).number_format = CUR
            ws9.cell(row=row, column=5).font = Font(bold=True, color="1565C0")
            ws9.cell(row=row, column=6, value=david_take).number_format = CUR
            # Highlight unicorn and $305M rows
            if a_entry["acq_price"] >= 1e9:
                for c in range(1, 7):
                    ws9.cell(row=row, column=c).fill = PatternFill(
                        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            elif a_entry["acq_price"] >= 305e6:
                for c in range(1, 7):
                    ws9.cell(row=row, column=c).fill = PatternFill(
                        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
            row += 1
        row += 1

    # Section 2: Milestone summary
    ws9.cell(row=row, column=1, value="MILESTONE SUMMARY").font = Font(bold=True, size=12)
    row += 1
    ms_h = ["Scenario", "$305M Exit Year", "Unicorn Year", "David at Unicorn"]
    for i, h in enumerate(ms_h, 1):
        ws9.cell(row=row, column=i, value=h)
    apply_header(ws9, row, 4)
    row += 1

    for label, scenarios, traj_func in [
        ("Baseline", UNICORN_SCENARIOS, baseline_traj),
        ("AI-Adjusted", AI_UNICORN_SCENARIOS, ai_traj),
    ]:
        for sname, path_data in traj_func.items():
            y305 = "-"
            yunicorn = "-"
            david_at_unicorn = 0
            for r in path_data[1:]:
                if r["acq_price"] >= 305e6 and y305 == "-":
                    y305 = f"Year {r['year']}"
                if r["acq_price"] >= 1e9 and yunicorn == "-":
                    yunicorn = f"Year {r['year']}"
                    david_at_unicorn = BOOTSTRAP_OWNERSHIP * r["acq_price"] * (1 - TAX_RATE)
            ws9.cell(row=row, column=1, value=f"{sname}")
            ws9.cell(row=row, column=2, value=y305)
            ws9.cell(row=row, column=3, value=yunicorn)
            ws9.cell(row=row, column=4, value=david_at_unicorn).number_format = CUR
            row += 1

    # Section 3: Probability waterfall
    row += 1
    ws9.cell(row=row, column=1, value="UNICORN PROBABILITY BUILD-UP").font = Font(bold=True, size=12)
    row += 1
    prob_stages = [
        ("Solo (David)", 0.051), ("+ Igor (CTO)", 0.08),
        ("+ Chris (CCO)", 0.175), ("Post Pre-Mortem", 0.21),
        ("+ Triangle Strategy", 0.27), ("+ AI Tailwinds", AI_UNICORN_PROBABILITY),
    ]
    pw_h = ["Stage", "Probability"]
    for i, h in enumerate(pw_h, 1):
        ws9.cell(row=row, column=i, value=h)
    apply_header(ws9, row, 2)
    row += 1
    for stage, prob in prob_stages:
        ws9.cell(row=row, column=1, value=stage)
        ws9.cell(row=row, column=2, value=f"{prob*100:.1f}%")
        if prob >= 0.27:
            ws9.cell(row=row, column=1).font = Font(bold=True, color="2E7D32")
            ws9.cell(row=row, column=2).font = Font(bold=True, color="2E7D32")
        row += 1

    # Risk adjustments
    row += 1
    ws9.cell(row=row, column=1, value="AI RISK ADJUSTMENTS").font = Font(bold=True, size=11)
    row += 1
    for factor, adj in AI_RISK_ADJUSTMENTS.items():
        ws9.cell(row=row, column=1, value=factor)
        ws9.cell(row=row, column=2, value=f"{adj*100:+.0f}pp")
        color = "2E7D32" if adj < 0 else "C62828"
        ws9.cell(row=row, column=2).font = Font(bold=True, color=color)
        row += 1

    wb.save(EXCEL_PATH)
    print(f"  Generated Excel workbook: {EXCEL_PATH}")


# ============================================================
# PDF REPORT
# ============================================================

def make_pdf():
    doc = SimpleDocTemplate(str(PDF_PATH), pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Title2", parent=styles["Title"], fontSize=22,
                              spaceAfter=20, textColor=colors.HexColor("#1E3A5F")))
    styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"],
                              textColor=colors.HexColor("#1E3A5F"), spaceBefore=16, spaceAfter=8))
    styles.add(ParagraphStyle(name="Body", parent=styles["BodyText"],
                              alignment=TA_JUSTIFY, fontSize=10, leading=14, spaceAfter=8))
    styles.add(ParagraphStyle(name="Sub", parent=styles["Normal"], fontSize=12,
                              textColor=colors.gray, alignment=TA_CENTER, spaceAfter=30))

    story = []

    # Cover
    story.append(Spacer(1, 1.5*inch))
    story.append(Paragraph("GuardSpine", styles["Title2"]))
    story.append(Paragraph("Market Analysis & Financial Model", styles["H2"]))
    story.append(Paragraph("Q1 2026 | Confidential", styles["Sub"]))
    story.append(Spacer(1, 0.3*inch))

    exec_text = """
    <b>Executive Summary</b><br/><br/>
    The $285B SaaS crash (Jan 30-31, 2026) validated a structural thesis: as AI generation costs
    collapse, the premium shifts to governance. GuardSpine is not a code review tool -- it is a
    multi-artifact governance spine for the AI office, covering code, PDFs, spreadsheets, and images
    through four Guard Lanes with swappable YAML rubrics.<br/><br/>
    <b>Open-core model (Linux/Red Hat):</b> Spec, verifier, CodeGuard Action, PII-Shield, and YAML
    rubrics are free and open source. Revenue from management UI, multi-lane coordination, enterprise
    features, and cognitive attestation licensing. <b>BYOK</b>: users bring their own LLM API keys --
    GuardSpine pays zero API costs.<br/><br/>
    <b>Unique advantages:</b><br/>
    &bull; <b>PII-Shield integration</b> (live): entropy-based secret detection + HMAC redaction in evidence bundles<br/>
    &bull; <b>Proprioceptive AI</b> (MOU pending, exclusive): cognitive probes reading model hidden states for
    deterministic confidence attestation -- the only product that proves both process AND cognition<br/>
    &bull; <b>Triangle Strategy</b>: Proprioceptive AI + Z-Inspection + G7/NIST = IBM as first enterprise client Q3 2026<br/>
    &bull; <b>Unicorn probability: 27%</b> (5.4x avg Series A) -- team + pre-mortem + Triangle Strategy<br/><br/>
    <b>Key numbers:</b><br/>
    &bull; Multi-artifact TAM: $51.3B (governance layer: $2.6-7.7B)<br/>
    &bull; SAM (3-segment): $850M<br/>
    &bull; Base case: $1.3M ARR Year 1, $36.8M Year 3<br/>
    &bull; Gross margins: 87-91% across paid tiers (BYOK: zero API COGS)<br/>
    &bull; LTV/CAC: 10.2-14.5x with 2.9-4.6 month payback<br/>
    &bull; 9/9 MECE dimensions covered (nearest competitor: 3/9)<br/>
    """
    story.append(Paragraph(exec_text, styles["Body"]))
    story.append(PageBreak())

    # Four Lanes Architecture
    story.append(Paragraph("1. The AI Office: Four Guard Lanes", styles["H2"]))
    lanes_fig = str(FIGURES_DIR / "four_lanes.png")
    if os.path.exists(lanes_fig):
        story.append(Image(lanes_fig, width=6.5*inch, height=3.9*inch))
    story.append(Spacer(1, 0.2*inch))
    lanes_text = """
    Every artifact in the AI office flows through the same governance spine: YAML rubric evaluation,
    L0-L4 risk tiering, approval routing, and hash-chained evidence bundle generation. The L0-L4
    risk tiers drive governance behavior:<br/><br/>
    &bull; <b>L0 (Auto-pass)</b>: Metadata logged, no review. 55% of changes.<br/>
    &bull; <b>L1 (Light review)</b>: AI summary, single reviewer. 25%.<br/>
    &bull; <b>L2 (Standard)</b>: Multi-model consensus, rubric eval, evidence bundle. 12%.<br/>
    &bull; <b>L3 (Elevated)</b>: Role-based approvers, stop-the-line gating. 6%.<br/>
    &bull; <b>L4 (Full audit)</b>: Cross-functional review, adversarial analysis, cognitive attestation. 2%.<br/>
    """
    story.append(Paragraph(lanes_text, styles["Body"]))
    tier_fig = str(FIGURES_DIR / "tier_distribution.png")
    if os.path.exists(tier_fig):
        story.append(Image(tier_fig, width=5.5*inch, height=3*inch))
    story.append(PageBreak())

    # Market Sizing
    story.append(Paragraph("2. Market Sizing", styles["H2"]))
    tam_fig = str(FIGURES_DIR / "tam_sam_som.png")
    if os.path.exists(tam_fig):
        story.append(Image(tam_fig, width=5*inch, height=5*inch))
    tam_text = f"""
    <b>Code-Only TAM:</b> ${CODE_ONLY_TAM/1e9:.2f}B (DevSecOps + AI Governance + Code Review)<br/>
    <b>Multi-Artifact TAM:</b> ${MULTI_ARTIFACT_TAM/1e9:.1f}B (adds GRC, Doc Mgmt, Digital Asset Mgmt)<br/>
    <b>Governance Layer:</b> ${MULTI_ARTIFACT_TAM*GOVERNANCE_LAYER_PCT_LOW/1e9:.1f}B - ${MULTI_ARTIFACT_TAM*GOVERNANCE_LAYER_PCT_HIGH/1e9:.1f}B (5-15% of multi-artifact TAM)<br/>
    <b>SAM:</b> ${SAM_TOTAL/1e6:.0f}M (3 segments: code regulated + doc governance + cross-artifact)<br/>
    <b>SOM Y1:</b> ${SAM_TOTAL*SOM_Y1_PCT/1e6:.1f}M | <b>Y2:</b> ${SAM_TOTAL*SOM_Y2_PCT/1e6:.0f}M |
    <b>Y3:</b> ${SAM_TOTAL*SOM_Y3_PCT/1e6:.0f}M
    """
    story.append(Paragraph(tam_text, styles["Body"]))
    story.append(PageBreak())

    # Unit Economics (Open-Core / BYOK)
    story.append(Paragraph("3. Unit Economics (Open-Core / BYOK)", styles["H2"]))
    oc_text = """
    <b>Open-Core Model (Linux/Red Hat):</b> The spec, verifier, CodeGuard GitHub Action,
    PII-Shield integration, and YAML rubric format are free and open source. Revenue comes from the
    management UI, multi-lane coordination, enterprise features, and cognitive attestation licensing.<br/><br/>
    <b>BYOK (Bring Your Own Keys):</b> Users provide their own LLM API keys. GuardSpine pays
    zero API costs. COGS = hosting/infrastructure + support + cognitive attestation licensing only.<br/><br/>
    <b>Conversion Funnel:</b> Free-to-Pro: 3% | Pro-to-Business: 20% (12mo) |
    Business-to-Enterprise: 10% (18mo)<br/><br/>
    <b>Open-Core Analogues:</b> Linux/RHEL ($3.4B ARR), GitLab CE/EE ($580M ARR),
    Docker CE/Desktop ($200M ARR), WordPress/Automattic ($700M+ ARR)
    """
    story.append(Paragraph(oc_text, styles["Body"]))

    # Only show paid tiers in the table (Community Free is $0 across the board)
    paid_tiers = {k: v for k, v in TIERS.items() if v["monthly"] > 0}
    ue_data = [["Metric"] + list(paid_tiers.keys())]
    for label, key in [("Monthly Subscription", "monthly"), ("Changes/Month", "changes_mo")]:
        ue_data.append([label] + [
            f'${paid_tiers[t][key]:,}' if key == "monthly" else str(paid_tiers[t][key])
            for t in paid_tiers
        ])
    # COGS breakdown rows
    for label, key in [("Hosting/Infra", "hosting"), ("Support", "support"),
                       ("Cognitive License", "cognitive_license")]:
        ue_data.append([label] + [f'${paid_tiers[t][key]:,}' for t in paid_tiers])
    # Computed rows
    cogs_row = ["Total COGS (BYOK)"]
    gp_row = ["Gross Profit"]
    gm_row = ["Gross Margin"]
    ltv_row = ["LTV (36mo)"]
    cac_row = ["Target CAC"]
    ratio_row = ["LTV/CAC"]
    payback_row = ["CAC Payback (mo)"]
    for t in paid_tiers:
        td = paid_tiers[t]
        cogs = tier_total_cogs(td)
        gp = td["monthly"] - cogs
        gm = gp / td["monthly"]
        ltv = td["monthly"] * gm * CUSTOMER_LIFETIME_MONTHS * td["nrr"]
        cac = td["cac"]
        cogs_row.append(f"${cogs:,}")
        gp_row.append(f"${gp:,}")
        gm_row.append(f"{gm*100:.1f}%")
        ltv_row.append(f"${ltv:,.0f}")
        cac_row.append(f"${cac:,}")
        ratio_row.append(f"{ltv/cac:.1f}x" if cac > 0 else "N/A")
        payback_row.append(f"{cac/gp:.1f}" if gp > 0 else "N/A")
    ue_data.extend([cogs_row, gp_row, gm_row, ltv_row, cac_row, ratio_row, payback_row])

    table = Table(ue_data, colWidths=[1.8*inch, 1.5*inch, 1.8*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    story.append(PageBreak())

    # Fee Cascade
    story.append(Paragraph("4. Fee Cascade & Governance Premium", styles["H2"]))
    cascade_text = """
    Professional service fees are compressing: KPMG -14% (Irish audit), Thomson Reuters -16%,
    LexisNexis -14%, LegalZoom -20%. Catalyst: Anthropic shipped a 200-line markdown prompt
    that displaced legal review workflows.<br/><br/>
    As fees compress, the governance vacuum grows. GuardSpine captures the inverse:
    governance share of software spend grows from 2% to ~19% over 36 months.
    """
    story.append(Paragraph(cascade_text, styles["Body"]))
    for fig_name in ["fee_cascade.png", "governance_premium.png"]:
        fig_path = str(FIGURES_DIR / fig_name)
        if os.path.exists(fig_path):
            story.append(Image(fig_path, width=6*inch, height=3.2*inch))
            story.append(Spacer(1, 0.1*inch))
    story.append(PageBreak())

    # Revenue Projections
    story.append(Paragraph("5. Revenue Projections", styles["H2"]))
    rev_data = [["Scenario", "Growth/mo", "Start Custs", "Avg ACV", "Y1 ARR"]]
    for name, p in SCENARIOS.items():
        c12 = customer_count(12, p["c0"], p["growth"])
        y1 = c12 * p["avg_acv"]
        rev_data.append([name, f'{p["growth"]*100:.0f}%', str(p["c0"]),
                         f'${p["avg_acv"]:,}', f'${y1:,.0f}'])
    t2 = Table(rev_data, colWidths=[1.2*inch]*5)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.2*inch))
    rev_fig = str(FIGURES_DIR / "revenue_scenarios.png")
    if os.path.exists(rev_fig):
        story.append(Image(rev_fig, width=6*inch, height=3.6*inch))
    story.append(PageBreak())

    # Competitive Positioning
    story.append(Paragraph("6. Competitive Positioning", styles["H2"]))
    mece_fig = str(FIGURES_DIR / "mece_heatmap.png")
    if os.path.exists(mece_fig):
        story.append(Image(mece_fig, width=6.5*inch, height=2.8*inch))
    story.append(Spacer(1, 0.2*inch))
    radar_fig = str(FIGURES_DIR / "competitive_radar.png")
    if os.path.exists(radar_fig):
        story.append(Image(radar_fig, width=5*inch, height=5*inch))
    story.append(PageBreak())

    # Unicorn Path
    story.append(Paragraph("7. Path to Unicorn", styles["H2"]))
    unicorn_intro = """
    <b>Model:</b> ARR(Y) = ARR(Y-1) * NRR + new_logos * ACV + enterprise_deals.
    Growth rate decays annually. Valuation = ARR * multiple (adjusted for growth rate,
    87-91% gross margins, and category-creator premium).<br/><br/>
    <b>Enterprise catalysts:</b> Netflix pilot (March 2026, pain-driven code review) and
    IBM (Triangle Strategy, Q3 2026). Netflix validates product; IBM validates enterprise
    governance. Both create reference-customer acceleration for subsequent deals.
    """
    story.append(Paragraph(unicorn_intro, styles["Body"]))

    # Unicorn path table
    up_data = [["Metric", "Bear", "Base", "Bull"]]
    unicorn_results = {}
    for name, params in UNICORN_SCENARIOS.items():
        unicorn_results[name] = compute_unicorn_path(params)
    # Find unicorn year
    for name in ["Bear", "Base", "Bull"]:
        uy = "Never"
        for r in unicorn_results[name]:
            if r["valuation"] >= 1e9:
                uy = f"Year {r['year']}"
                break
        up_data.append(["Unicorn Year", "", "", ""])
        break
    up_data = [["Metric", "Bear", "Base", "Bull"]]
    for year in [1, 2, 3, 4, 5]:
        row = [f"Year {year} ARR"]
        for name in ["Bear", "Base", "Bull"]:
            r = unicorn_results[name][year]
            row.append(f"${r['arr']/1e6:.1f}M")
        up_data.append(row)
    for year in [1, 2, 3, 4, 5]:
        row = [f"Year {year} Valuation"]
        for name in ["Bear", "Base", "Bull"]:
            r = unicorn_results[name][year]
            row.append(f"${r['valuation']/1e6:.0f}M")
        up_data.append(row)
    # Unicorn year
    uy_row = ["Unicorn Year"]
    for name in ["Bear", "Base", "Bull"]:
        found = False
        for r in unicorn_results[name]:
            if r["valuation"] >= 1e9:
                uy_row.append(f"Year {r['year']}")
                found = True
                break
        if not found:
            uy_row.append("Never (>7yr)")
    up_data.append(uy_row)

    t_up = Table(up_data, colWidths=[1.8*inch, 1.4*inch, 1.4*inch, 1.4*inch])
    t_up.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(t_up)
    story.append(Spacer(1, 0.2*inch))
    up_fig = str(FIGURES_DIR / "unicorn_path.png")
    if os.path.exists(up_fig):
        story.append(Image(up_fig, width=6.5*inch, height=3*inch))
    story.append(Spacer(1, 0.2*inch))

    unicorn_note = """
    <b>Base case (planning scenario):</b> Unicorn at Year 4 with ~$49M ARR at 23x multiple.
    Driven by 15%/mo initial growth (decaying to ~8%/mo by Y4), 120% NRR from lane expansion,
    and $3.3M in enterprise catalyst deals (Netflix full deployment + IBM pilot + reference-driven
    enterprise sales).<br/><br/>
    <b>Bull case:</b> Unicorn at Year 2-3. Requires viral open-core adoption (20%/mo initial growth)
    and rapid enterprise deal closure. Comparable to Wiz ($100M ARR in 18mo) or Snyk (~18%/mo growth
    in first 3 years).<br/><br/>
    <b>Bear case:</b> Does not reach unicorn within 7 years. Viable as a profitable niche business
    or acquisition target.
    """
    story.append(Paragraph(unicorn_note, styles["Body"]))
    story.append(PageBreak())

    # --- 7b: Pre-Mortem Probability Model ---
    story.append(Paragraph("7b. Pre-Mortem Probability Analysis", styles["H2"]))
    prob_intro = """
    A pre-mortem analysis identifies 8 independent risk factors that could prevent unicorn outcome.
    Team composition (Igor as CTO + Chris Hood as CCO) and pre-mortem validation gates (Phase 0:
    OSS repo validation before Netflix pilot) shift the cumulative probability from 5.1% baseline
    to 21%. The Triangle Strategy (Logan/Proprioceptive AI + Ishwar/Z-Inspection/IBM +
    Jacob/G7/NIST) independently reduces 6 of 8 risk factors, pushing the probability to
    27% -- placing GuardSpine in the top 1% of pre-revenue probability profiles, at 5.4x
    the average Series A startup.
    """
    story.append(Paragraph(prob_intro, styles["Body"]))
    prob_fig = str(FIGURES_DIR / "probability_waterfall.png")
    if os.path.exists(prob_fig):
        story.append(Image(prob_fig, width=6.5*inch, height=3*inch))
    story.append(Spacer(1, 0.1*inch))

    # Probability table
    prob_data = [["Team Configuration", "Unicorn Probability"]]
    for label, prob in UNICORN_PROBABILITIES.items():
        prob_data.append([label, f"{prob*100:.1f}%"])
    prob_data.append(["Avg Series A startup", "~5%"])
    t_prob = Table(prob_data, colWidths=[3*inch, 2*inch])
    t_prob.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Oblique"),
    ]))
    story.append(t_prob)
    story.append(Spacer(1, 0.15*inch))

    team_text = """
    <b>Key team multipliers:</b><br/>
    &bull; <b>Igor Malovitsa (CTO):</b> 13yr commercial engineering, Rust + cryptography + blockchain,
    MSc experimental nuclear physics, Stanford crypto coursework. Built ASTERIX radar codec in Rust
    in 3 weeks. Eliminates technical execution risk.<br/>
    &bull; <b>Chris Hood (CCO):</b> 7yr Google Head of Digital Strategy, Nomotic AI inventor (USPTO
    trademark), PhilArchive publication, 500+ enterprise relationships, 4M+ listener podcast,
    AAA game studio co-founder. Eliminates GTM/sales risk and accelerates IEEE standard submission.<br/>
    &bull; <b>Pre-mortem validation gates:</b> Phase 0 (OSS repo validation before Netflix pilot)
    reduces technical execution failure from 15% to 8-10%. Capital discipline ("raise minimum,
    stay profitable") reduces capital risk from 15-20% to 8-10%.<br/><br/>
    <b>Triangle Strategy multiplier (21% -> 27%):</b><br/>
    &bull; <b>Ishwar (Z-Inspection / IBM):</b> Internal champion at IBM. Z-Inspection assessment
    = independent third-party validation. Reduces GTM risk (warm sale vs cold outbound) and
    PMF risk (enterprise willingness to pilot = PMF signal).<br/>
    &bull; <b>Jacob (G7 / NIST):</b> Standards reference creates regulatory lock-in. When G7
    recommends "cognitive state attestation in evidence bundles," only one product implements
    the pattern. Reduces competitive moat risk from 11% to 6%.<br/>
    &bull; <b>Logan (Proprioceptive AI):</b> MOU Section 5 exclusive in governance vertical.
    55 provisional patents. No competitor can replicate cognitive attestation. Combined with
    G7 reference = strongest possible capital-raising narrative (9% to 5% capital risk).
    """
    story.append(Paragraph(team_text, styles["Body"]))
    story.append(PageBreak())

    # --- 7c: Capital Strategy & Founder Liquidity ---
    story.append(Paragraph("7c. Capital Strategy & Founder Liquidity (Bootstrap Primary)", styles["H2"]))

    bootstrap_intro = f"""
    <b>PRIMARY PATH: Bootstrap with $300K angel round.</b> BYOK model means zero API COGS.
    Two technical founders (David + Igor) plus AI coding assistance = equivalent of a 5-person team.
    Monthly burn: ~$15K. The $300K angel round at $10M valuation (3% dilution) provides ~20 months
    of runway insurance. David retains ~{BOOTSTRAP_OWNERSHIP*100:.1f}% ownership.<br/><br/>
    <b>Why bootstrap beats VC:</b> Every avoided round saves $35-74M in dilution at exit.
    At 87% gross margins, revenue self-funds growth after the first 2-3 enterprise customers.
    The open-core GitHub Action markets itself. Warm enterprise intros (Netflix via Dennis,
    IBM via Ishwar) eliminate the need for a paid sales team early.<br/><br/>
    <b>$100M personal target math:</b> At {BOOTSTRAP_OWNERSHIP*100:.1f}% ownership and 25% tax,
    David needs a ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit. This is NOT a unicorn -- it is a
    $15-20M ARR company with 87% margins in a 40% CAGR market, acquirable at 15-25x revenue.
    """
    story.append(Paragraph(bootstrap_intro, styles["Body"]))
    story.append(Spacer(1, 0.1*inch))

    # Path comparison table
    path_comp, exit_list = compute_path_comparison()
    pc_data = [["Path", "Raised", "David %", "Exit for $100M*", "Take @ $400M", "Take @ $1B"]]
    for pname, pdata in path_comp.items():
        exit_needed = PERSONAL_TARGET / (pdata["ownership"] * (1 - TAX_RATE))
        take_400 = pdata["ownership"] * 400e6 * (1 - TAX_RATE)
        take_1b = pdata["ownership"] * 1e9 * (1 - TAX_RATE)
        raised = pdata["total_raised"]
        raised_str = f"${raised/1e3:.0f}K" if raised < 1e6 else f"${raised/1e6:.0f}M"
        pc_data.append([
            pname.split("(")[0].strip(),
            raised_str,
            f'{pdata["ownership"]*100:.1f}%',
            f'${exit_needed/1e6:.0f}M',
            f'${take_400/1e6:.0f}M',
            f'${take_1b/1e6:.0f}M',
        ])
    t_pc = Table(pc_data, colWidths=[1.2*inch, 0.7*inch, 0.7*inch, 1.1*inch, 1.0*inch, 1.0*inch])
    t_pc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#E8F5E9")),
    ]))
    story.append(t_pc)
    story.append(Spacer(1, 0.08*inch))
    story.append(Paragraph("*Post-tax $100M target. Green row = primary path.", styles["Body"]))

    liq_fig = str(FIGURES_DIR / "personal_liquidity.png")
    if os.path.exists(liq_fig):
        story.append(Image(liq_fig, width=6.5*inch, height=3.2*inch))
    story.append(Spacer(1, 0.1*inch))

    liq_text = f"""
    <b>Key insight:</b> The bootstrap path halves the required exit valuation. At {BOOTSTRAP_OWNERSHIP*100:.1f}%
    ownership, David clears $100M post-tax at just ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M -- compared to $702M
    with full VC dilution (19.0% ownership). The VC path raises $76M but costs $195M in equity at a $1B exit.<br/><br/>
    <b>Revenue milestones (pre-mortem adjusted, base case):</b><br/>
    &bull; First customer: Month 5-6 (85% prob) | $1M ARR: Month 11-15 (60% prob)<br/>
    &bull; $10M ARR: Month 22-28 (30% prob) | $30M ARR: Month 30-38 (22% prob)<br/><br/>
    <b>Netflix kill criteria:</b> False positive rate &lt;5%, false negative rate &lt;2%,
    reduce 1,000 review decisions to 10. 116 engineers pushing code 4x faster than review capacity.
    ACV $200K-$2M (pricing anchor: saves one senior hire at $180-250K/yr).<br/><br/>
    <b>When to reconsider VC:</b> Only if (a) a strategic investor (IBM Ventures, Microsoft M12)
    offers distribution that accelerates revenue 3x+ faster than organic growth, AND the dilution
    cost at projected exit is less than the incremental revenue gained. Otherwise, bootstrap.
    """
    story.append(Paragraph(liq_text, styles["Body"]))
    story.append(PageBreak())

    # Appendix
    story.append(Paragraph("Appendix: Partnerships & Path to First Client", styles["H2"]))
    partner_text = """
    <b>PII-Shield (Live)</b>: Entropy-based secret detection + HMAC redaction. Integrated in
    codeguard-action Phase 1. Deterministic org-wide PII_SALT for cross-bundle correlation.
    7-phase rollout across 16 repos.<br/><br/>
    <b>Proprioceptive AI (MOU Pending)</b>: 55 provisional patents. Cognitive probes read LLM hidden
    states for deterministic confidence attestation. MOU Section 5: exclusive in governance vertical.
    Claims: 999x class separation, 0.003% overhead.<br/><br/>
    <b>Triangle Strategy</b>: Logan (tech) + Ishwar/Z-Inspection (validation) + Jacob/G7 (standards) =
    IBM as first enterprise client Q3 2026. The loop: integration -> assessment -> standards reference
    -> procurement pressure -> enterprise sale.<br/><br/>
    <b>Timeline</b>: MOU Feb -> Prototype Mar -> Z-Inspection Apr -> G7 Reference May -> IBM Pilot Q3 2026
    """
    story.append(Paragraph(partner_text, styles["Body"]))
    story.append(PageBreak())

    # --- Acquirer Analysis ---
    story.append(Paragraph("8. Strategic Acquirer Analysis", styles["H2"]))
    acq_intro = """
    <b>Dual-track strategy:</b> Raise the pre-seed at maximum valuation while building acquirer
    relationships organically. The pre-seed round funds the product + enterprise pilots. If an
    acquisition offer materializes at a multiple satisfying the founder liquidity model ($600M+),
    evaluate it. If not, continue building.<br/><br/>
    <b>Pre-seed valuation target: $15-25M.</b> At pre-seed, valuation = narrative + signals, not
    revenue. Each signal stacked before raising compresses dilution. The difference between $8M and
    $20M pre-seed is 7% equity saved = $42M at $600M exit.
    """
    story.append(Paragraph(acq_intro, styles["Body"]))
    acq_fig = str(FIGURES_DIR / "acquirer_scorecard.png")
    if os.path.exists(acq_fig):
        story.append(Image(acq_fig, width=6.5*inch, height=3*inch))
    story.append(Spacer(1, 0.15*inch))

    # Acquirer table
    total_w = sum(ACQUIRER_WEIGHTS.values())
    acq_data = [["Acquirer", "Score", "Price Range", "Key Rationale"]]
    sorted_acqs = sorted(ACQUIRER_SCORES.items(),
                          key=lambda x: sum(x[1][d] * ACQUIRER_WEIGHTS[d] for d in ACQUIRER_WEIGHTS) / total_w,
                          reverse=True)
    for name, scores in sorted_acqs:
        composite = sum(scores[d] * ACQUIRER_WEIGHTS[d] for d in ACQUIRER_WEIGHTS) / total_w
        lo, hi = scores["price_range_m"]
        lo_s = f"${lo/1000:.1f}B" if lo >= 1000 else f"${lo:.0f}M"
        hi_s = f"${hi/1000:.1f}B" if hi >= 1000 else f"${hi:.0f}M"
        acq_data.append([
            name, f"{composite:.1f}/10",
            f"{lo_s}-{hi_s}",
            scores["why"][:70] + "..." if len(scores["why"]) > 70 else scores["why"],
        ])
    t_acq = Table(acq_data, colWidths=[1.4*inch, 0.7*inch, 1.1*inch, 3.3*inch])
    t_acq.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (2, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t_acq)
    story.append(Spacer(1, 0.15*inch))

    acq_closing = """
    <b>Top 2 acquirers:</b><br/>
    &bull; <b>Microsoft/GitHub:</b> Build the best GitHub Action. Make CodeGuard the de facto governance
    action in the Marketplace. When they see 10K+ installs, they notice. They don't acquire products;
    they acquire adoption. GuardSpine = "Copilot Trust" or "GitHub Governance."<br/>
    &bull; <b>IBM:</b> Execute the Triangle Strategy. IBM doesn't need to be pitched -- they need to
    experience the product through Ishwar. Let Z-Inspection be the evaluation. IBM Ventures for
    strategic pre-seed investment = acquisition on-ramp.<br/><br/>
    <b>90-day signal stack (Feb-May 2026):</b> MOU signed -> Phase 0 validated -> Netflix bug test ->
    Z-Inspection in progress -> Jacob interest documented -> raise at $15-25M with full signal stack.
    """
    story.append(Paragraph(acq_closing, styles["Body"]))
    story.append(PageBreak())

    # --- Section 9: Path to Acquisition (Bootstrap Primary) ---
    story.append(Paragraph("9. Path to Acquisition: Bootstrap vs VC Exit Analysis", styles["H2"]))

    acq_traj = compute_acquisition_trajectory()
    billion_year, billion_arr, billion_price = find_billion_year(acq_traj, "Base")

    path_intro = f"""
    <b>Bootstrap target: ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M acquisition = $100M post-tax for David.</b>
    At {BOOTSTRAP_OWNERSHIP*100:.1f}% ownership (only 3% dilution from angel round), the exit bar is
    dramatically lower than the VC path ($702M needed at 19.0% ownership).<br/><br/>
    In the Base case, acquisition price crosses the ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M bootstrap target
    well before $1B. The $1B mark is reached in <b>Year {billion_year}</b> at ~${billion_arr/1e6:.0f}M ARR
    (implied acquisition price: ${billion_price/1e6:.0f}M) -- but this is upside, not the requirement.<br/><br/>
    <b>Strategic premium model:</b> Acquirers pay above implied market valuation for category-defining
    companies. Benchmarks: GitHub at 37.5x revenue ($7.5B at $200M ARR), Red Hat at 10x ($34B at $3.4B ARR),
    Bridgecrew at ~100x ($200M at ~$2M ARR -- early, strategic). Our model applies 20-50% premium
    depending on stage (pre-revenue: 1.5x, early: 1.4x, growth: 1.3x, scale: 1.2x).
    """
    story.append(Paragraph(path_intro, styles["Body"]))

    # Acquisition trajectory figure
    traj_fig = str(FIGURES_DIR / "acquisition_trajectory.png")
    if os.path.exists(traj_fig):
        story.append(Image(traj_fig, width=6.5*inch, height=3.2*inch))
    story.append(Spacer(1, 0.15*inch))

    # Acquisition trajectory table
    traj_data = [["Year", "Base ARR", "Market Val", "Premium", "Acq Price", "vs $1B"]]
    for r in acq_traj["Base"]:
        if r["year"] == 0:
            continue
        deficit = r["acq_price"] - TARGET_ACQUISITION
        status = f"+${deficit/1e6:.0f}M" if deficit >= 0 else f"-${abs(deficit)/1e6:.0f}M"
        traj_data.append([
            f"Year {r['year']}",
            f"${r['arr']/1e6:.1f}M",
            f"${r['market_val']/1e6:.0f}M",
            f"{r['premium']:.1f}x",
            f"${r['acq_price']/1e6:.0f}M",
            status,
        ])
    t_traj = Table(traj_data, colWidths=[0.8*inch, 1.0*inch, 1.0*inch, 0.8*inch, 1.0*inch, 1.0*inch])
    t_traj.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(t_traj)
    story.append(Spacer(1, 0.15*inch))
    story.append(PageBreak())

    # Dilution cost comparison section
    story.append(Paragraph("9b. Why Bootstrap Wins: Dilution Cost Comparison", styles["H2"]))

    # Compute both paths for comparison
    boot_dil, boot_own = compute_dilution_waterfall(rounds=BOOTSTRAP_PATH,
                                                     target_exits=[305e6, 400e6, 600e6, 1e9])
    vc4_dil, vc4_own = compute_dilution_waterfall(rounds=VC_PATH,
                                                   target_exits=[305e6, 400e6, 600e6, 1e9])

    dil_intro = f"""
    <b>Bootstrap (angel only):</b> David retains {boot_own*100:.1f}% after $300K angel round (3% dilution).
    At ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit = $100M post-tax. At $400M = ${boot_own * 400e6 * (1-TAX_RATE) / 1e6:.0f}M.
    At $1B = ${boot_own * 1e9 * (1-TAX_RATE) / 1e6:.0f}M.<br/><br/>
    <b>VC 4-round (comparison):</b> David dilutes to {vc4_own*100:.1f}% after raising $76M across 4 rounds.
    At $1B exit = ${vc4_own * 1e9 * (1-TAX_RATE) / 1e6:.0f}M post-tax. Must reach $702M just to clear $100M.<br/><br/>
    <b>The math is clear:</b> VC raises $76M but costs David ${(boot_own - vc4_own) * 1e9 * (1-TAX_RATE) / 1e6:.0f}M
    in equity at a $1B exit. That is a 2.5x cost of capital. Bootstrap wins unless VC distribution
    accelerates revenue 3x+ beyond organic growth.
    """
    story.append(Paragraph(dil_intro, styles["Body"]))

    # Dilution comparison figure
    dil_fig = str(FIGURES_DIR / "dilution_waterfall.png")
    if os.path.exists(dil_fig):
        story.append(Image(dil_fig, width=6.5*inch, height=3.2*inch))
    story.append(Spacer(1, 0.15*inch))

    # Side-by-side comparison table
    comp_data = [["Metric", "Bootstrap (Angel)", "VC 4-Round", "Difference"]]
    comp_data.append(["Total raised", "$300K", "$76M", "$75.7M more capital"])
    comp_data.append(["David ownership", f"{boot_own*100:.1f}%", f"{vc4_own*100:.1f}%",
                       f"-{(boot_own - vc4_own)*100:.1f}pp"])
    comp_data.append(["Exit for $100M", f"${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M",
                       f"${PERSONAL_TARGET / (vc4_own * (1-TAX_RATE)) / 1e6:.0f}M",
                       f"${(PERSONAL_TARGET / (vc4_own * (1-TAX_RATE)) - BOOTSTRAP_EXIT_NEEDED) / 1e6:.0f}M higher bar"])
    comp_data.append(["Take @ $400M exit",
                       f"${boot_own * 400e6 * (1-TAX_RATE) / 1e6:.0f}M",
                       f"${vc4_own * 400e6 * (1-TAX_RATE) / 1e6:.0f}M",
                       f"${(boot_own - vc4_own) * 400e6 * (1-TAX_RATE) / 1e6:.0f}M more"])
    comp_data.append(["Take @ $1B exit",
                       f"${boot_own * 1e9 * (1-TAX_RATE) / 1e6:.0f}M",
                       f"${vc4_own * 1e9 * (1-TAX_RATE) / 1e6:.0f}M",
                       f"${(boot_own - vc4_own) * 1e9 * (1-TAX_RATE) / 1e6:.0f}M more"])

    t_comp = Table(comp_data, colWidths=[1.3*inch, 1.4*inch, 1.3*inch, 1.5*inch])
    t_comp.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("BACKGROUND", (1, 1), (1, -1), colors.HexColor("#E8F5E9")),
    ]))
    story.append(t_comp)
    story.append(Spacer(1, 0.15*inch))

    # Acquisition requirements (still relevant for $1B upside)
    reqs_text = f"""
    <b>Bootstrap target requirements (${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit):</b><br/>
    &bull; <b>ARR:</b> $15-20M (at 15-25x revenue multiple with strategic premium)<br/>
    &bull; <b>Growth:</b> 50%+ YoY at time of acquisition<br/>
    &bull; <b>Margins:</b> 85%+ gross (we project 87-91% from BYOK model)<br/>
    &bull; <b>Logos:</b> 3-5 enterprise reference customers<br/>
    &bull; <b>Timeline:</b> Year 2-3 (Month 24-36)<br/><br/>
    <b>$1B+ upside requirements (if market allows continued growth):</b><br/>
    &bull; <b>ARR:</b> $30-50M minimum ($30M floor, $50M for confident pricing)<br/>
    &bull; <b>Logos:</b> 5+ enterprise reference customers (IBM + Netflix + 3)<br/>
    &bull; <b>Distribution:</b> 10K+ open-source installs (adoption wedge for acquirer)<br/>
    &bull; <b>Strategic signals:</b> 9/9 MECE, exclusive patents, G7/NIST reference, Z-Inspection report<br/><br/>
    <b>David's personal outcome (bootstrap):</b><br/>
    &bull; At ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M: $100M post-tax (the target)<br/>
    &bull; At $400M: ${boot_own * 400e6 * (1-TAX_RATE) / 1e6:.0f}M post-tax<br/>
    &bull; At $600M: ${boot_own * 600e6 * (1-TAX_RATE) / 1e6:.0f}M post-tax<br/>
    &bull; At $1B: ${boot_own * 1e9 * (1-TAX_RATE) / 1e6:.0f}M post-tax
    """
    story.append(Paragraph(reqs_text, styles["Body"]))
    story.append(PageBreak())

    # --- Section 10: Speed-to-$100M Execution Plan ---
    story.append(Paragraph("10. Speed-to-$100M: Execution Plan", styles["H2"]))

    speed_intro = f"""
    <b>The name of the game is speed.</b> Markets move fast. Governance is a timing play -- the $285B
    crash opened a window that will narrow as incumbents respond. Every month of delay costs optionality.<br/><br/>
    <b>Target:</b> ${BOOTSTRAP_EXIT_NEEDED/1e6:.0f}M exit at {BOOTSTRAP_OWNERSHIP*100:.1f}% ownership = $100M post-tax.
    $15-20M ARR at 15-25x revenue with strategic premium. Best case: {SPEED_BEST_CASE_MONTHS} months.
    Realistic: {SPEED_REALISTIC_MONTHS} months. <b>AI-adjusted (Section 11):</b> {AI_SPEED_BEST_CASE_MONTHS} months
    best case, {AI_SPEED_REALISTIC_MONTHS} months realistic. Everything before Month 14 is fully in David's hands.
    """
    story.append(Paragraph(speed_intro, styles["Body"]))

    # Speed timeline figure
    speed_fig = str(FIGURES_DIR / "speed_timeline.png")
    if os.path.exists(speed_fig):
        story.append(Image(speed_fig, width=6.5*inch, height=4.5*inch))
    story.append(Spacer(1, 0.15*inch))

    # Phase-by-phase table
    phase_data = [["Phase", "Months", "Target ARR", "Key Actions"]]
    for phase in SPEED_TIMELINE:
        phase_data.append([
            phase["phase"],
            f'{phase["months"][0]}-{phase["months"][1]}',
            f'${phase["arr"]/1e6:.0f}M' if phase["arr"] >= 1e6 else "$0",
            " | ".join(phase["actions"][:2]),
        ])
    t_speed = Table(phase_data, colWidths=[1.2*inch, 0.7*inch, 0.8*inch, 3.8*inch])
    t_speed.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (2, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(t_speed)
    story.append(Spacer(1, 0.15*inch))

    # Speed killers section
    story.append(Paragraph("10b. Speed Killers: Five Ways to Add 6-12 Months", styles["H2"]))

    killers_text = ""
    for k in SPEED_KILLERS:
        killers_text += f"""
        <b>{k['name']}</b> (+{k['cost_months']} months): {k['description']}<br/>
        <i>Fix: {k['fix']}</i><br/><br/>
        """
    story.append(Paragraph(killers_text, styles["Body"]))

    # Controllable vs not
    ctrl_text = """
    <b>What David Controls:</b> Product quality, pricing discipline, which customers to approach,
    open-source distribution, patent timing, when to start acquirer conversations, whether to
    accept sub-$305M offers.<br/><br/>
    <b>What David Doesn't Control:</b> Enterprise procurement speed, acquirer timing appetite,
    market sentiment at exit, competitor responses, regulatory timeline.<br/><br/>
    <b>The gap between 18 and 24 months is entirely determined by procurement cycles and acquirer
    competition.</b> Everything before Month 14 is execution. After that, it's execution quality
    meeting market timing.
    """
    story.append(Paragraph(ctrl_text, styles["Body"]))
    story.append(PageBreak())

    # --- Section 11: AI Trajectory Effects on the Math ---
    story.append(Paragraph("11. AI Trajectory: How Accelerating AI Changes the Math", styles["H2"]))

    ai_intro = f"""
    <b>The math changes if AI keeps getting better -- and the data says it will.</b><br/><br/>
    METR (Epoch AI affiliate) measures AI task completion horizons. The current doubling time is
    <b>7 months</b>: today's models handle 50-minute tasks autonomously; by late 2026-2027 they will
    handle week-long tasks. Token costs decline <b>10x per year</b> ($20/M in 2022, $0.40/M in 2025,
    $0.04/M projected 2026). SWE-bench Verified hit <b>75% solve rate</b>. 41% of code is already
    AI-generated. 79% of enterprises have adopted agentic AI (340% YoY surge).<br/><br/>
    The EU AI Act becomes fully enforceable <b>August 2, 2026</b> with fines up to <b>7% of global turnover</b>.
    This is not a prediction -- it is published law.<br/><br/>
    These forces change 5 things about GuardSpine's math:
    """
    story.append(Paragraph(ai_intro, styles["Body"]))

    # AI-adjusted figure
    ai_fig = str(FIGURES_DIR / "ai_adjusted_math.png")
    if os.path.exists(ai_fig):
        story.append(Image(ai_fig, width=6.5*inch, height=4.5*inch))
    story.append(Spacer(1, 0.15*inch))

    # 5 effects table
    effects_data = [["Effect", "Baseline", "AI-Adjusted"]]
    for eff in AI_EFFECTS:
        effects_data.append([
            eff["effect"],
            eff.get("baseline", ""),
            eff.get("adjusted", ""),
        ])
    t_effects = Table(effects_data, colWidths=[2.2*inch, 1.8*inch, 2.5*inch])
    t_effects.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#1565C0")),
        ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
    ]))
    story.append(t_effects)
    story.append(Spacer(1, 0.15*inch))

    # AI-adjusted timeline section
    story.append(Paragraph("11b. AI-Adjusted Timeline", styles["H2"]))

    ai_tl_text = f"""
    <b>Baseline:</b> {SPEED_BEST_CASE_MONTHS} months best case, {SPEED_REALISTIC_MONTHS} months realistic.<br/>
    <b>AI-Adjusted:</b> {AI_SPEED_BEST_CASE_MONTHS} months best case, {AI_SPEED_REALISTIC_MONTHS} months realistic.<br/><br/>
    <b>Three drivers of compression:</b><br/>
    &bull; <b>Lane delivery accelerates:</b> 2-3 months per lane becomes 3-4 weeks. AI handles 60-80% of implementation at SWE-bench 75%.<br/>
    &bull; <b>Higher ACV = fewer customers needed:</b> EU AI Act compliance pressure lifts mid-market ACV from $200K to $300-500K. Fewer deals to close for same ARR.<br/>
    &bull; <b>Higher multiples = lower ARR target:</b> At 25-35x (vs 15-25x), $305M exit requires $9-12M ARR instead of $15-20M.<br/><br/>
    <b>Critical calendar date:</b> EU AI Act enforcement (August 2, 2026) creates a hard deadline for acquirers.
    Every major cloud and DevSecOps vendor needs governance capabilities before this date. Build-vs-buy
    analysis tilts decisively to buy when timeline pressure is high.
    """
    story.append(Paragraph(ai_tl_text, styles["Body"]))

    # Exit math table
    story.append(Paragraph("11c. Exit Math at AI-Adjusted Multiples", styles["H2"]))

    exit_data = [["Revenue Multiple", "ARR Required for $305M", "Zone"]]
    for mult in [15, 20, 25, 30, 35]:
        arr_req = BOOTSTRAP_EXIT_NEEDED / mult
        zone = "Baseline" if mult < 20 else "AI-Adjusted" if mult <= 35 else ""
        exit_data.append([
            f"{mult}x",
            f"${arr_req/1e6:.1f}M",
            zone,
        ])
    t_exit = Table(exit_data, colWidths=[1.5*inch, 2.0*inch, 3.0*inch])
    t_exit.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(t_exit)
    story.append(Spacer(1, 0.1*inch))

    ai_conclusion = f"""
    <b>Bottom line:</b> At AI-adjusted multiples, David hits $100M post-tax at ~$9-12M ARR (Month 14-20)
    instead of $15-20M ARR (Month 18-24). The AI trajectory compresses the exit window by 4-6 months
    and raises the probability of a $305M+ offer by increasing acquirer urgency (EU AI Act deadline)
    and willingness to pay (governance is now regulatory mandate, not optional).<br/><br/>
    <b>The AI wave is a tailwind for GuardSpine, not a threat.</b> Every gain in AI capability increases
    artifact volume (more to govern), reduces build cost (faster to market), and increases buyer urgency
    (compliance deadlines approach faster than incumbents can build). The window to capture this is
    12-20 months. The optimal exit is in 2027.
    """
    story.append(Paragraph(ai_conclusion, styles["Body"]))
    story.append(PageBreak())

    # --- Section 12: AI-Adjusted Unicorn Path ---
    story.append(Paragraph("12. AI-Adjusted Path to Unicorn", styles["H2"]))

    unicorn_intro = f"""
    <b>How soon does GuardSpine reach $1B+ acquisition value with AI tailwinds?</b><br/><br/>
    The baseline model reaches unicorn at <b>Year 4</b> (Base case) or <b>Year 2</b> (Bull case).
    With AI tailwinds -- higher multiples, higher ACV, faster growth, regulatory urgency --
    the AI-adjusted model reaches unicorn at <b>Year 3</b> (Base) or <b>Year 2</b> (Bull).<br/><br/>
    <b>Unicorn probability rises from 27% to {AI_UNICORN_PROBABILITY*100:.0f}%</b> (8x the Series A average of 5%).
    The AI wave improves 5 of 6 risk factors; only competitive moat risk increases (shorter head start).
    """
    story.append(Paragraph(unicorn_intro, styles["Body"]))

    # Unicorn path figure
    unicorn_fig = str(FIGURES_DIR / "ai_unicorn_path.png")
    if os.path.exists(unicorn_fig):
        story.append(Image(unicorn_fig, width=6.5*inch, height=3.5*inch))
    story.append(Spacer(1, 0.15*inch))

    # Milestone table
    story.append(Paragraph("12b. Unicorn Milestone Comparison", styles["H2"]))

    ai_traj_pdf = compute_ai_acquisition_trajectory()
    baseline_traj_pdf = compute_acquisition_trajectory()

    milestone_data = [["Scenario", "$305M Exit", "Unicorn ($1B+)", "David at Unicorn"]]
    for traj_dict in [baseline_traj_pdf, ai_traj_pdf]:
        for sname, path_data in traj_dict.items():
            y305 = "-"
            yunicorn = "-"
            david_u = "$0"
            for r in path_data[1:]:
                if r["acq_price"] >= 305e6 and y305 == "-":
                    y305 = f"Year {r['year']}"
                if r["acq_price"] >= 1e9 and yunicorn == "-":
                    yunicorn = f"Year {r['year']}"
                    david_u = f"${BOOTSTRAP_OWNERSHIP * r['acq_price'] * (1-TAX_RATE) / 1e6:.0f}M"
            milestone_data.append([sname, y305, yunicorn, david_u])

    t_mile = Table(milestone_data, colWidths=[1.8*inch, 1.2*inch, 1.2*inch, 1.5*inch])
    t_mile.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(t_mile)
    story.append(Spacer(1, 0.15*inch))

    # Probability waterfall
    story.append(Paragraph("12c. Unicorn Probability: 5.1% to 41%", styles["H2"]))

    prob_text = f"""
    <b>The probability build-up:</b><br/>
    &bull; Solo (David): 5.1% (baseline Series A average)<br/>
    &bull; + Igor (CTO): 8.0% (technical de-risk: 13yr Rust/crypto, physics MSc)<br/>
    &bull; + Chris (CCO): 17.5% (GTM de-risk: ex-Google 7yr, Nomotic AI, USPTO)<br/>
    &bull; Post Pre-Mortem: 21.0% (risk identification + mitigation plan)<br/>
    &bull; + Triangle Strategy: 27.0% (IBM pathway: Logan + Ishwar + Jacob)<br/>
    &bull; <b>+ AI Tailwinds: {AI_UNICORN_PROBABILITY*100:.0f}%</b> (regulatory urgency + higher multiples + faster build)<br/><br/>
    <b>AI risk adjustments:</b><br/>
    &bull; Technical Execution: -2pp (AI handles 60-80% of implementation at SWE-bench 75%)<br/>
    &bull; Product-Market Fit: -3pp (EU AI Act converts governance from optional to mandate)<br/>
    &bull; GTM / Sales: -2pp (buyer urgency up, procurement faster under deadline)<br/>
    &bull; Competitive Moat: +3pp (ONLY NEGATIVE -- shorter head start, but moat is data not code)<br/>
    &bull; Capital Access: -1pp (higher multiples = easier to raise if needed)<br/>
    &bull; Scaling: -2pp (AI handles more of the work, fewer people needed)<br/><br/>
    <b>Net: -7pp risk reduction = +14pp probability lift.</b> 41% is 8x the Series A average.
    """
    story.append(Paragraph(prob_text, styles["Body"]))

    # Base case narrative
    story.append(Paragraph("12d. Base Case (AI): Year-by-Year to Unicorn", styles["H2"]))

    base_ai_path = ai_traj_pdf.get("Base (AI)", [])
    base_narrative = ""
    for r in base_ai_path[1:4]:  # Years 1-3
        david_take = BOOTSTRAP_OWNERSHIP * r["acq_price"] * (1 - TAX_RATE)
        base_narrative += (
            f"<b>Year {r['year']} (Month {r['year']*12}):</b> "
            f"ARR ${r['arr']/1e6:.1f}M | "
            f"Acquisition price ${r['acq_price']/1e6:.0f}M | "
            f"David post-tax ${david_take/1e6:.0f}M"
        )
        if r["acq_price"] >= 1e9:
            base_narrative += " <b>[UNICORN]</b>"
        elif r["acq_price"] >= 305e6:
            base_narrative += " <b>[$305M EXIT ZONE]</b>"
        base_narrative += "<br/>"

    base_narrative += """<br/>
    <b>The AI-adjusted base case reaches $305M exit (David's $100M) at Year 2 and
    unicorn ($1B+) at Year 3.</b> This is 1 year earlier than the baseline model for both milestones.
    At unicorn, David's post-tax take is $744M at 43.65% bootstrap ownership -- far above the
    $100M target and impossible at VC-diluted ownership levels.
    """
    story.append(Paragraph(base_narrative, styles["Body"]))

    doc.build(story)
    print(f"  Generated PDF report: {PDF_PATH}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("GuardSpine Market Analysis Generator v2")
    print("=" * 50)
    print("Thesis: Multi-artifact AI Office governance spine")
    print("Lanes: CodeGuard + PDFGuard + SheetGuard + ImageGuard")
    print("Partners: PII-Shield (live) + Proprioceptive AI (MOU pending)")

    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    print("\n[1/3] Generating figures...")
    make_figures()

    print("\n[2/3] Generating Excel workbook...")
    make_excel()

    print("\n[3/3] Generating PDF report...")
    make_pdf()

    print("\n" + "=" * 50)
    print("All artifacts generated successfully.")
    print(f"  Excel:   {EXCEL_PATH}")
    print(f"  PDF:     {PDF_PATH}")
    print(f"  Figures: {FIGURES_DIR}")


if __name__ == "__main__":
    main()
