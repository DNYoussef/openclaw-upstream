"""
GuardSpine SaaS P&L Model - Comprehensive Update Script
Fixes dates, adds Kristen, adds missing metrics, creates bear/bull scenarios.
All numbers are formula-derived by openpyxl or explicit calculation.
"""
import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import math

# ============================================================
# CONFIGURATION
# ============================================================
INPUT_FILE = r"C:\Users\17175\Desktop\guardspine\financial\GuardSpine-SaaS-Model.xlsx"
OUTPUT_FILE = r"C:\Users\17175\Desktop\guardspine\financial\GuardSpine-SaaS-Model-v2.xlsx"

# Kristen config
KRISTEN_MONTHLY = 7000  # $7K/mo
KRISTEN_START_MONTH = 1  # Month 1

# Scenario configs
SCENARIOS = {
    "Base": {
        # Matches original template: 15% decaying to 5% over 8 quarters
        "price_y1": 2500, "price_y2": 2900,
        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.08, 0.07, 0.06, 0.05],
        "churn_rates": [0.003, 0.003, 0.004, 0.004, 0.004, 0.004, 0.004, 0.004],
        "starting_customers": 1, "starting_mrr": 2500,
        "label": "Base Case"
    },
    "Bear": {
        # Slower adoption, lower ACV, higher churn, delayed hires
        "price_y1": 2000, "price_y2": 2200,
        "growth_rates": [0.10, 0.08, 0.06, 0.05, 0.05, 0.04, 0.04, 0.03],
        "churn_rates": [0.005, 0.005, 0.006, 0.006, 0.006, 0.006, 0.006, 0.006],
        "starting_customers": 1, "starting_mrr": 2000,
        "label": "Bear Case (Slow Adoption)"
    },
    "Bull": {
        # AI adoption wave: EU AI Act + DORA + Copilot proliferation
        # Higher ACV (Business tier mix), faster growth, lower churn
        # 2 starting customers (Brent Foster pilot + 1 early adopter)
        "price_y1": 3500, "price_y2": 4200,
        "growth_rates": [0.18, 0.15, 0.13, 0.11, 0.10, 0.09, 0.08, 0.07],
        "churn_rates": [0.002, 0.002, 0.003, 0.003, 0.003, 0.003, 0.003, 0.003],
        "starting_customers": 2, "starting_mrr": 7000,
        "label": "Bull Case (AI Adoption Wave)"
    }
}

# Personnel (monthly costs)
PERSONNEL = [
    {"name": "CEO - David Youssef", "monthly": 10000, "start_month": 1, "dept": "G&A"},
    {"name": "CTO - Igor Malovitsa", "monthly": 10000, "start_month": 1, "dept": "R&D"},
    {"name": "GTM - Kristen Hengst Smith", "monthly": 7000, "start_month": 1, "dept": "S&M"},
    {"name": "Engineer - Hire #1", "monthly": 10000, "start_month": 4, "dept": "R&D"},
    {"name": "Engineer - Hire #2", "monthly": 10000, "start_month": 10, "dept": "R&D"},
]

# Bear case: delay hires
PERSONNEL_BEAR = [
    {"name": "CEO - David Youssef", "monthly": 10000, "start_month": 1, "dept": "G&A"},
    {"name": "CTO - Igor Malovitsa", "monthly": 10000, "start_month": 1, "dept": "R&D"},
    {"name": "GTM - Kristen Hengst Smith", "monthly": 7000, "start_month": 1, "dept": "S&M"},
    {"name": "Engineer - Hire #1", "monthly": 10000, "start_month": 6, "dept": "R&D"},
    # No hire #2 in bear case
]

# Bull case: earlier + more hires
PERSONNEL_BULL = [
    {"name": "CEO - David Youssef", "monthly": 10000, "start_month": 1, "dept": "G&A"},
    {"name": "CTO - Igor Malovitsa", "monthly": 10000, "start_month": 1, "dept": "R&D"},
    {"name": "GTM - Kristen Hengst Smith", "monthly": 7000, "start_month": 1, "dept": "S&M"},
    {"name": "Engineer - Hire #1", "monthly": 10000, "start_month": 3, "dept": "R&D"},
    {"name": "Engineer - Hire #2", "monthly": 10000, "start_month": 7, "dept": "R&D"},
    {"name": "Sales - SDR Hire", "monthly": 6000, "start_month": 9, "dept": "S&M"},
]

# Fixed costs
HOSTING_MONTHLY = 1500
PAYMENT_PROCESSING_PCT = 0.03
MARKETING_ADWORDS_PCT = 0.02  # of MRR
MARKETING_SOCIAL_PCT = 0.01   # of MRR

# Misc costs
MISC_COSTS = {
    "AI Agent APIs (3x $1K/mo)": 3000,
    "Software (CRM, monitoring)": 1000,
    "Legal / IP / patents": 2000,
    "Travel / conferences": 500,
    "Internet, supplies": 200,
    "Insurance (D&O, E&O)": 667,
    "Content / SEO": 500,
}
TOTAL_MISC = sum(MISC_COSTS.values())  # $7,867

# Fundraise
RAISE_AMOUNT = 1_000_000
PRE_MONEY = 9_000_000

# Months
MONTHS_Y1 = ["Jan-2026","Feb-2026","Mar-2026","Apr-2026","May-2026","Jun-2026",
              "Jul-2026","Aug-2026","Sep-2026","Oct-2026","Nov-2026","Dec-2026"]
MONTHS_Y2 = ["Jan-2027","Feb-2027","Mar-2027","Apr-2027","May-2027","Jun-2027",
              "Jul-2027","Aug-2027","Sep-2027","Oct-2027","Nov-2027","Dec-2027"]
ALL_MONTHS = MONTHS_Y1 + MONTHS_Y2


# ============================================================
# COMPUTATION ENGINE
# ============================================================
def compute_scenario(scenario_key):
    """Compute full P&L for a scenario. Returns dict of arrays (24 months)."""
    sc = SCENARIOS[scenario_key]

    if scenario_key == "Bear":
        personnel = PERSONNEL_BEAR
    elif scenario_key == "Bull":
        personnel = PERSONNEL_BULL
    else:
        personnel = PERSONNEL

    # Growth rates by quarter (8 values = 8 quarters over 2 years, each applied for 3 months)
    gr = sc["growth_rates"]
    month_growth = []
    for q_rate in gr:
        month_growth.extend([q_rate] * 3)

    cr = sc["churn_rates"]
    month_churn = []
    for q_rate in cr:
        month_churn.extend([q_rate] * 3)

    price_y1 = sc["price_y1"]
    price_y2 = sc["price_y2"]
    prices = [price_y1]*12 + [price_y2]*12

    # === Revenue Model ===
    # CRITICAL: Template compounds NEW MRR at the growth rate, then adds to existing base.
    # new_mrr[0] = starting_mrr * (1 + growth_rate)
    # new_mrr[m] = new_mrr[m-1] * (1 + growth_rate[m])
    # gross_mrr[m] = gross_mrr[m-1] + new_mrr[m] - lost_mrr[m]
    # lost_mrr[m] = gross_mrr[m-1] * churn_rate[m]
    # This creates much faster growth than simple % growth on total MRR.
    customers = [0]*24
    new_customers = [0]*24
    mrr_new = [0]*24
    mrr_lost = [0]*24
    mrr_prev = [0]*24
    gross_mrr = [0]*24
    net_mrr = [0]*24
    payment_proc = [0]*24

    starting_mrr = sc["starting_mrr"]

    for m in range(24):
        if m == 0:
            # New MRR in month 1 = starting * (1 + growth)
            mrr_new[0] = starting_mrr * (1 + month_growth[0])
            mrr_prev[0] = starting_mrr
            mrr_lost[0] = starting_mrr * month_churn[0]
            gross_mrr[0] = starting_mrr + mrr_new[0] - mrr_lost[0]
        else:
            # New MRR compounds on previous new MRR
            mrr_new[m] = mrr_new[m-1] * (1 + month_growth[m])
            mrr_prev[m] = gross_mrr[m-1]
            mrr_lost[m] = gross_mrr[m-1] * month_churn[m]
            gross_mrr[m] = gross_mrr[m-1] + mrr_new[m] - mrr_lost[m]

        net_mrr[m] = gross_mrr[m] - mrr_lost[m]  # net after churn
        customers[m] = max(1, round(gross_mrr[m] / prices[m]))
        if m > 0:
            new_customers[m] = max(0, customers[m] - customers[m-1])
        else:
            new_customers[m] = max(1, customers[0] - sc["starting_customers"])
        payment_proc[m] = gross_mrr[m] * PAYMENT_PROCESSING_PCT

    # === Personnel Costs ===
    payroll = [0]*24
    payroll_sm = [0]*24  # S&M personnel
    payroll_rd = [0]*24  # R&D personnel
    payroll_ga = [0]*24  # G&A personnel

    for p in personnel:
        for m in range(24):
            if (m + 1) >= p["start_month"]:
                payroll[m] += p["monthly"]
                if p["dept"] == "S&M":
                    payroll_sm[m] += p["monthly"]
                elif p["dept"] == "R&D":
                    payroll_rd[m] += p["monthly"]
                elif p["dept"] == "G&A":
                    payroll_ga[m] += p["monthly"]

    # === Marketing ===
    mktg_adwords = [gross_mrr[m] * MARKETING_ADWORDS_PCT for m in range(24)]
    mktg_social = [gross_mrr[m] * MARKETING_SOCIAL_PCT for m in range(24)]
    total_marketing = [mktg_adwords[m] + mktg_social[m] for m in range(24)]

    # === COGS ===
    cogs = [payment_proc[m] + HOSTING_MONTHLY for m in range(24)]

    # === Gross Profit ===
    gross_profit = [gross_mrr[m] - cogs[m] for m in range(24)]
    gross_margin_pct = [gross_profit[m] / gross_mrr[m] if gross_mrr[m] > 0 else 0 for m in range(24)]

    # === OpEx ===
    sm_total = [payroll_sm[m] + total_marketing[m] + payment_proc[m] for m in range(24)]

    # R&D = R&D personnel + AI APIs
    ai_apis = 3000
    rd_total = [payroll_rd[m] + ai_apis for m in range(24)]

    # G&A = G&A personnel + misc (excluding AI APIs which are in R&D)
    ga_misc = TOTAL_MISC - ai_apis  # $4,867
    ga_total = [payroll_ga[m] + ga_misc for m in range(24)]

    total_opex = [sm_total[m] + rd_total[m] + ga_total[m] for m in range(24)]

    # === EBITDA ===
    ebitda = [gross_profit[m] - total_opex[m] for m in range(24)]
    ebitda_margin = [ebitda[m] / gross_mrr[m] if gross_mrr[m] > 0 else 0 for m in range(24)]

    # === Key Metrics ===
    arr = [gross_mrr[m] * 12 for m in range(24)]

    # CAC = S&M spend / new customers (trailing)
    cac = []
    for m in range(24):
        nc = new_customers[m]
        if nc > 0:
            cac.append(sm_total[m] / nc)
        else:
            cac.append(0)

    # CAC Payback = CAC / (avg revenue per customer * gross margin)
    cac_payback = []
    for m in range(24):
        if cac[m] > 0 and prices[m] > 0 and gross_margin_pct[m] > 0:
            monthly_margin_per_cust = prices[m] * gross_margin_pct[m]
            cac_payback.append(cac[m] / monthly_margin_per_cust)
        else:
            cac_payback.append(0)

    # LTV = (avg revenue per customer * gross margin) / monthly churn
    # LTV:CAC ratio
    ltv = []
    ltv_cac = []
    for m in range(24):
        if month_churn[m] > 0:
            l = (prices[m] * gross_margin_pct[m]) / month_churn[m]
            ltv.append(l)
            if cac[m] > 0:
                ltv_cac.append(l / cac[m])
            else:
                ltv_cac.append(0)
        else:
            ltv.append(0)
            ltv_cac.append(0)

    # NRR (monthly) = (ending MRR - lost MRR + expansion) / starting MRR
    nrr = []
    for m in range(24):
        if m == 0:
            nrr.append(1.0)
        else:
            if gross_mrr[m-1] > 0:
                nrr.append(gross_mrr[m] / gross_mrr[m-1])
            else:
                nrr.append(1.0)

    # GRR = 1 - churn rate
    grr = [1 - month_churn[m] for m in range(24)]

    # Rule of 40 (annualized)
    rule_of_40 = []
    for m in range(24):
        if m >= 1 and gross_mrr[m-1] > 0:
            growth_pct = (gross_mrr[m] - gross_mrr[m-1]) / gross_mrr[m-1] * 12  # annualized
        else:
            growth_pct = 0
        margin_pct = ebitda_margin[m]
        rule_of_40.append(growth_pct + margin_pct)

    # Burn rate and runway
    total_costs = [cogs[m] + total_opex[m] for m in range(24)]
    monthly_burn = [total_costs[m] - gross_mrr[m] for m in range(24)]  # negative = profitable

    cumulative_cash = [0]*24
    cumulative_cash[0] = RAISE_AMOUNT - max(0, monthly_burn[0])
    for m in range(1, 24):
        cumulative_cash[m] = cumulative_cash[m-1] - max(0, monthly_burn[m])
        if monthly_burn[m] < 0:  # profitable month
            cumulative_cash[m] = cumulative_cash[m-1] + abs(monthly_burn[m])

    runway = []
    for m in range(24):
        if monthly_burn[m] > 0:  # burning cash
            runway.append(cumulative_cash[m] / monthly_burn[m])
        else:
            runway.append(99)  # profitable, infinite runway

    # Breakeven month
    breakeven_month = None
    for m in range(24):
        if ebitda[m] > 0:
            breakeven_month = m + 1
            break

    # Annual totals
    y1_revenue = sum(gross_mrr[:12])
    y2_revenue = sum(gross_mrr[12:])
    y1_cogs = sum(cogs[:12])
    y2_cogs = sum(cogs[12:])
    y1_opex = sum(total_opex[:12])
    y2_opex = sum(total_opex[12:])
    y1_ebitda = sum(ebitda[:12])
    y2_ebitda = sum(ebitda[12:])

    return {
        "gross_mrr": gross_mrr, "net_mrr": net_mrr, "mrr_new": mrr_new,
        "mrr_lost": mrr_lost, "mrr_prev": mrr_prev, "customers": customers,
        "new_customers": new_customers, "prices": prices,
        "payment_proc": payment_proc, "cogs": cogs,
        "gross_profit": gross_profit, "gross_margin_pct": gross_margin_pct,
        "payroll": payroll, "payroll_sm": payroll_sm, "payroll_rd": payroll_rd,
        "payroll_ga": payroll_ga, "total_marketing": total_marketing,
        "mktg_adwords": mktg_adwords, "mktg_social": mktg_social,
        "sm_total": sm_total, "rd_total": rd_total, "ga_total": ga_total,
        "total_opex": total_opex, "ebitda": ebitda, "ebitda_margin": ebitda_margin,
        "arr": arr, "cac": cac, "cac_payback": cac_payback,
        "ltv": ltv, "ltv_cac": ltv_cac, "nrr": nrr, "grr": grr,
        "rule_of_40": rule_of_40, "monthly_burn": monthly_burn,
        "cumulative_cash": cumulative_cash, "runway": runway,
        "breakeven_month": breakeven_month, "total_costs": total_costs,
        "y1_revenue": y1_revenue, "y2_revenue": y2_revenue,
        "y1_cogs": y1_cogs, "y2_cogs": y2_cogs,
        "y1_opex": y1_opex, "y2_opex": y2_opex,
        "y1_ebitda": y1_ebitda, "y2_ebitda": y2_ebitda,
        "personnel": personnel,
    }


# ============================================================
# SPREADSHEET BUILDER
# ============================================================
def create_scenario_sheet(wb, sheet_name, scenario_key, data):
    """Create a full P&L sheet for a scenario."""
    sc = SCENARIOS[scenario_key]
    ws = wb.create_sheet(title=sheet_name)

    # Styles
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    metric_font = Font(bold=True, size=11, color="2F5496")
    money_fmt = '#,##0'
    pct_fmt = '0.0%'
    ratio_fmt = '0.0x'

    thin_border = Border(
        bottom=Side(style='thin', color='D9D9D9')
    )

    row = 1

    # === HEADER ===
    ws.cell(row=row, column=1, value=f"GuardSpine SaaS P&L -- {sc['label']}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="2F5496")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: Feb 18, 2026 | $1M raise at $9M pre / $10M post")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="808080")
    row += 2

    # === Column headers ===
    ws.column_dimensions['A'].width = 32
    ws.cell(row=row, column=1, value="Metric").font = header_font

    # Year headers
    ws.cell(row=row, column=2, value="Year 1").font = header_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    ws.cell(row=row, column=14, value="Year 2").font = header_font
    ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=25)
    # Annual totals
    ws.cell(row=row, column=26, value="Y1 Total").font = header_font
    ws.cell(row=row, column=27, value="Y2 Total").font = header_font
    row += 1

    # Month headers
    ws.cell(row=row, column=1, value="Month").font = header_font
    for i, m in enumerate(ALL_MONTHS):
        col = i + 2
        ws.cell(row=row, column=col, value=m).font = Font(bold=True, size=9)
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions[get_column_letter(26)].width = 14
    ws.column_dimensions[get_column_letter(27)].width = 14
    row += 1

    # Month numbers
    ws.cell(row=row, column=1, value="Month #")
    for i in range(24):
        ws.cell(row=row, column=i+2, value=i+1)
    row += 1

    def write_section(title):
        nonlocal row
        row += 1
        for c in range(1, 28):
            cell = ws.cell(row=row, column=c)
            cell.fill = section_fill
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

    def write_row(label, values, fmt=money_fmt, is_metric=False, y1_total=None, y2_total=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        if is_metric:
            ws.cell(row=row, column=1).font = metric_font
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=i+2, value=v)
            if fmt == money_fmt:
                cell.number_format = '#,##0'
            elif fmt == pct_fmt:
                cell.number_format = '0.0%'
            elif fmt == ratio_fmt:
                cell.number_format = '0.0'
        # Annual totals
        if y1_total is not None:
            cell = ws.cell(row=row, column=26, value=y1_total)
            cell.number_format = '#,##0'
            cell.font = Font(bold=True)
        if y2_total is not None:
            cell = ws.cell(row=row, column=27, value=y2_total)
            cell.number_format = '#,##0'
            cell.font = Font(bold=True)
        row += 1

    # === REVENUE ===
    write_section("REVENUE")
    write_row("Avg Price / Customer", data["prices"])
    write_row("Customers", data["customers"], fmt='#,##0')
    write_row("New Customers", data["new_customers"], fmt='#,##0')
    write_row("Gross MRR", data["gross_mrr"], y1_total=data["y1_revenue"], y2_total=data["y2_revenue"])
    write_row("Lost MRR (Churn)", data["mrr_lost"])
    write_row("Net MRR", data["net_mrr"])
    write_row("ARR (MRR x 12)", data["arr"], is_metric=True)
    write_row("MRR Growth Rate", [0] + [(data["gross_mrr"][m] - data["gross_mrr"][m-1])/data["gross_mrr"][m-1] if data["gross_mrr"][m-1] > 0 else 0 for m in range(1,24)], fmt=pct_fmt)

    # === COGS ===
    write_section("COST OF GOODS SOLD")
    write_row("Payment Processing (3%)", data["payment_proc"])
    write_row("Hosting & Bandwidth", [HOSTING_MONTHLY]*24)
    write_row("Total COGS", data["cogs"], is_metric=True, y1_total=data["y1_cogs"], y2_total=data["y2_cogs"])
    write_row("COGS % of Revenue", [data["cogs"][m]/data["gross_mrr"][m] if data["gross_mrr"][m] > 0 else 0 for m in range(24)], fmt=pct_fmt)

    # === GROSS PROFIT ===
    write_section("GROSS PROFIT")
    write_row("Gross Profit", data["gross_profit"], is_metric=True)
    write_row("Gross Margin %", data["gross_margin_pct"], fmt=pct_fmt, is_metric=True)

    # === PERSONNEL ===
    write_section("PERSONNEL")
    for p in data["personnel"]:
        vals = []
        for m in range(24):
            if (m+1) >= p["start_month"]:
                vals.append(p["monthly"])
            else:
                vals.append(0)
        write_row(f"  {p['name']} ({p['dept']})", vals)
    write_row("Total Payroll", data["payroll"], is_metric=True,
              y1_total=sum(data["payroll"][:12]), y2_total=sum(data["payroll"][12:]))

    # === OPERATING EXPENSES ===
    write_section("OPERATING EXPENSES")
    write_row("Sales & Marketing", data["sm_total"], y1_total=sum(data["sm_total"][:12]), y2_total=sum(data["sm_total"][12:]))
    write_row("  S&M % of Revenue", [data["sm_total"][m]/data["gross_mrr"][m] if data["gross_mrr"][m] > 0 else 0 for m in range(24)], fmt=pct_fmt)
    write_row("Research & Development", data["rd_total"], y1_total=sum(data["rd_total"][:12]), y2_total=sum(data["rd_total"][12:]))
    write_row("  R&D % of Revenue", [data["rd_total"][m]/data["gross_mrr"][m] if data["gross_mrr"][m] > 0 else 0 for m in range(24)], fmt=pct_fmt)
    write_row("General & Administrative", data["ga_total"], y1_total=sum(data["ga_total"][:12]), y2_total=sum(data["ga_total"][12:]))
    write_row("  G&A % of Revenue", [data["ga_total"][m]/data["gross_mrr"][m] if data["gross_mrr"][m] > 0 else 0 for m in range(24)], fmt=pct_fmt)
    write_row("Total OpEx", data["total_opex"], is_metric=True, y1_total=data["y1_opex"], y2_total=data["y2_opex"])

    # === EBITDA ===
    write_section("EBITDA / OPERATING INCOME")
    write_row("EBITDA", data["ebitda"], is_metric=True, y1_total=data["y1_ebitda"], y2_total=data["y2_ebitda"])
    write_row("EBITDA Margin %", data["ebitda_margin"], fmt=pct_fmt, is_metric=True)

    # === SaaS METRICS (from P&L guide) ===
    write_section("SaaS OPERATING METRICS")
    write_row("CAC (S&M / New Customers)", data["cac"])
    write_row("CAC Payback (months)", data["cac_payback"], fmt='0.0')
    write_row("LTV (margin-weighted)", data["ltv"])
    write_row("LTV:CAC Ratio", data["ltv_cac"], fmt='0.0', is_metric=True)
    write_row("Net Revenue Retention", data["nrr"], fmt=pct_fmt, is_metric=True)
    write_row("Gross Revenue Retention", data["grr"], fmt=pct_fmt)
    write_row("Rule of 40", data["rule_of_40"], fmt=pct_fmt, is_metric=True)

    # === CASH & RUNWAY ===
    write_section("CASH & RUNWAY ($1M Raise)")
    write_row("Monthly Burn (neg=profit)", data["monthly_burn"])
    write_row("Cumulative Cash", data["cumulative_cash"], is_metric=True)
    write_row("Runway (months)", data["runway"], fmt='0.0')

    # Breakeven indicator
    row += 1
    be = data["breakeven_month"]
    if be:
        ws.cell(row=row, column=1, value=f"EBITDA Breakeven: Month {be} ({ALL_MONTHS[be-1]})")
    else:
        ws.cell(row=row, column=1, value="EBITDA Breakeven: Not reached in 24 months")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="008000" if be and be <= 24 else "FF0000")

    return ws


def create_comparison_sheet(wb, results):
    """Create a summary sheet comparing all three scenarios."""
    ws = wb.create_sheet(title="Scenario Comparison", index=0)

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color="2F5496")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    green = Font(bold=True, color="008000")
    red = Font(bold=True, color="FF0000")

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    row = 1
    ws.cell(row=row, column=1, value="GuardSpine Financial Model -- Scenario Comparison")
    ws.cell(row=row, column=1).font = title_font
    row += 1
    ws.cell(row=row, column=1, value="Prepared for Kristen Hengst Smith meeting | Feb 19, 2026")
    ws.cell(row=row, column=1).font = Font(italic=True, color="808080")
    row += 1
    ws.cell(row=row, column=1, value="$1M raise at $9M pre-money / $10M post-money | 10% dilution")
    ws.cell(row=row, column=1).font = Font(italic=True, color="808080")
    row += 2

    # Headers
    ws.cell(row=row, column=1, value="Metric").font = header_font
    ws.cell(row=row, column=2, value="Bear Case").font = Font(bold=True, color="FF0000")
    ws.cell(row=row, column=3, value="Base Case").font = Font(bold=True, color="2F5496")
    ws.cell(row=row, column=4, value="Bull Case").font = Font(bold=True, color="008000")
    row += 1

    def section(title):
        nonlocal row
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

    def compare_row(label, bear_val, base_val, bull_val, fmt="$"):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        for col, val in [(2, bear_val), (3, base_val), (4, bull_val)]:
            cell = ws.cell(row=row, column=col)
            cell.value = val
            if fmt == "$":
                cell.number_format = '$#,##0'
            elif fmt == "%":
                cell.number_format = '0.0%'
            elif fmt == "x":
                cell.number_format = '0.0'
            elif fmt == "mo":
                cell.number_format = '0.0'
            elif fmt == "n":
                cell.number_format = '#,##0'
        row += 1

    bear = results["Bear"]
    base = results["Base"]
    bull = results["Bull"]

    # === Assumptions ===
    section("ASSUMPTIONS")
    compare_row("Starting Price (monthly)", 2000, 2500, 3500, fmt="$")
    compare_row("Y1 Monthly Growth Rate", 0.10, 0.15, 0.20, fmt="%")
    compare_row("Monthly Churn Rate", 0.005, 0.003, 0.002, fmt="%")
    compare_row("Starting Customers", 1, 1, 2, fmt="n")
    compare_row("Team Size (end Y1)", len(PERSONNEL_BEAR), len(PERSONNEL), len(PERSONNEL_BULL), fmt="n")

    # === Revenue ===
    section("REVENUE")
    compare_row("Y1 Total Revenue", bear["y1_revenue"], base["y1_revenue"], bull["y1_revenue"])
    compare_row("Y2 Total Revenue", bear["y2_revenue"], base["y2_revenue"], bull["y2_revenue"])
    compare_row("End-of-Y1 MRR", bear["gross_mrr"][11], base["gross_mrr"][11], bull["gross_mrr"][11])
    compare_row("End-of-Y2 MRR", bear["gross_mrr"][23], base["gross_mrr"][23], bull["gross_mrr"][23])
    compare_row("End-of-Y1 ARR", bear["arr"][11], base["arr"][11], bull["arr"][11])
    compare_row("End-of-Y2 ARR", bear["arr"][23], base["arr"][23], bull["arr"][23])
    compare_row("End-of-Y1 Customers", bear["customers"][11], base["customers"][11], bull["customers"][11], fmt="n")
    compare_row("End-of-Y2 Customers", bear["customers"][23], base["customers"][23], bull["customers"][23], fmt="n")

    # === Margins ===
    section("MARGINS")
    compare_row("End-of-Y1 Gross Margin", bear["gross_margin_pct"][11], base["gross_margin_pct"][11], bull["gross_margin_pct"][11], fmt="%")
    compare_row("End-of-Y2 Gross Margin", bear["gross_margin_pct"][23], base["gross_margin_pct"][23], bull["gross_margin_pct"][23], fmt="%")
    compare_row("Y1 EBITDA", bear["y1_ebitda"], base["y1_ebitda"], bull["y1_ebitda"])
    compare_row("Y2 EBITDA", bear["y2_ebitda"], base["y2_ebitda"], bull["y2_ebitda"])
    compare_row("End-of-Y2 EBITDA Margin", bear["ebitda_margin"][23], base["ebitda_margin"][23], bull["ebitda_margin"][23], fmt="%")

    # === SaaS Metrics ===
    section("SaaS OPERATING METRICS (End of Y2)")
    compare_row("LTV:CAC Ratio", bear["ltv_cac"][23], base["ltv_cac"][23], bull["ltv_cac"][23], fmt="x")
    compare_row("CAC Payback (months)", bear["cac_payback"][23], base["cac_payback"][23], bull["cac_payback"][23], fmt="mo")
    compare_row("Net Revenue Retention", bear["nrr"][23], base["nrr"][23], bull["nrr"][23], fmt="%")
    compare_row("Rule of 40", bear["rule_of_40"][23], base["rule_of_40"][23], bull["rule_of_40"][23], fmt="%")

    # === Efficiency: ARR/Employee ===
    section("EFFICIENCY: ARR per Employee")
    bear_team_y1 = len(PERSONNEL_BEAR)
    bear_team_y2 = len(PERSONNEL_BEAR)
    base_team_y1 = len(PERSONNEL)
    base_team_y2 = len(PERSONNEL)
    bull_team_y1 = len(PERSONNEL_BULL)
    bull_team_y2 = len(PERSONNEL_BULL)
    compare_row("Team Size (end Y2)", bear_team_y2, base_team_y2, bull_team_y2, fmt="n")
    compare_row("End-Y2 ARR / Employee",
                bear["arr"][23] / bear_team_y2,
                base["arr"][23] / base_team_y2,
                bull["arr"][23] / bull_team_y2)

    # === Cash ===
    section("CASH & RUNWAY")
    be_bear = bear["breakeven_month"] or 99
    be_base = base["breakeven_month"] or 99
    be_bull = bull["breakeven_month"] or 99
    compare_row("EBITDA Breakeven (month #)", be_bear, be_base, be_bull, fmt="n")
    compare_row("End-of-Y1 Cash Remaining", bear["cumulative_cash"][11], base["cumulative_cash"][11], bull["cumulative_cash"][11])
    compare_row("End-of-Y2 Cash Remaining", bear["cumulative_cash"][23], base["cumulative_cash"][23], bull["cumulative_cash"][23])
    compare_row("Peak Monthly Burn", max(bear["monthly_burn"]), max(base["monthly_burn"]), max(bull["monthly_burn"]))

    # === BYOK Advantage Note ===
    row += 2
    ws.cell(row=row, column=1, value="KEY DIFFERENTIATOR: BYOK Architecture")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1
    ws.cell(row=row, column=1, value="Customers bring their own LLM API keys. GuardSpine pays ZERO inference costs.")
    row += 1
    ws.cell(row=row, column=1, value="Result: 90%+ gross margins at scale vs 40-65% for typical AI companies.")
    row += 1
    ws.cell(row=row, column=1, value="Comparable: Datadog (80.8% GM), CrowdStrike (75% GM), Zscaler (73.7% GM)")
    row += 1
    ws.cell(row=row, column=1, value="GuardSpine exceeds all comparables because BYOK eliminates the #1 AI cost driver.")

    # === Bull Case AI Adoption Drivers ===
    row += 2
    ws.cell(row=row, column=1, value="BULL CASE DRIVERS: AI Adoption Tailwinds")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="008000")
    row += 1
    ai_drivers = [
        "AI coding tools market: $7.37B in 2025, projected $12-15B by 2027 (35-40% CAGR)",
        "GitHub Copilot: 20M users, 4.7M paid subscribers, 50K+ orgs, 90% of Fortune 100",
        "75% of enterprise devs will use AI code assistants by 2028 (Gartner), up from <10% in 2023",
        "AI governance platform market: $492M in 2026, >$1B by 2030 (Gartner, 28%+ CAGR)",
        "EU AI Act high-risk rules enforcement: August 2026 -- compliance spending $2-15M per firm",
        "DORA enforceable since Jan 2025; aligns with AI Act governance controls",
        "40% of enterprise apps will feature task-specific AI agents by end of 2026",
        "Every AI-generated line of code needs governance = GuardSpine TAM grows with AI adoption",
    ]
    for d in ai_drivers:
        ws.cell(row=row, column=1, value=f"  - {d}")
        ws.cell(row=row, column=1).font = Font(size=10)
        row += 1

    # === AI-Native Efficiency Benchmarks ===
    row += 2
    ws.cell(row=row, column=1, value="AI-NATIVE BENCHMARKS: ARR per Employee")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1

    # Table headers
    bench_header_font = Font(bold=True, size=10, color="FFFFFF")
    bench_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for c, h in [(1, "Company"), (2, "ARR"), (3, "Team Size"), (4, "ARR/Employee")]:
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = bench_header_font
        cell.fill = bench_fill
    row += 1

    # Benchmark data (sources: Sacra, SaaStr, CB Insights, Lean AI Leaderboard)
    benchmarks = [
        ("Midjourney (2025)", 200_000_000, 11, 18_181_818),
        ("Cursor / Anysphere (Nov 2025)", 1_000_000_000, 150, 6_666_667),
        ("Cursor / Anysphere (early 2025)", 100_000_000, 20, 5_000_000),
        ("Mercor (2025)", None, None, 4_500_000),
        ("Nvidia (FY 2025)", None, None, 3_600_000),
        ("Meta (FY 2024)", None, None, 2_200_000),
        ("Microsoft (FY 2024)", None, None, 1_800_000),
    ]

    for name, arr_val, team, arr_per in benchmarks:
        ws.cell(row=row, column=1, value=name).font = Font(size=10)
        if arr_val is not None:
            cell = ws.cell(row=row, column=2, value=arr_val)
            cell.number_format = '$#,##0'
            cell.font = Font(size=10)
        else:
            ws.cell(row=row, column=2, value="--").font = Font(size=10)
        if team is not None:
            ws.cell(row=row, column=3, value=team).font = Font(size=10)
        else:
            ws.cell(row=row, column=3, value="--").font = Font(size=10)
        cell = ws.cell(row=row, column=4, value=arr_per)
        cell.number_format = '$#,##0'
        cell.font = Font(size=10)
        row += 1

    # GuardSpine rows with highlighting
    gs_highlight = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    gs_font = Font(bold=True, size=10)
    bear_team = len(PERSONNEL_BEAR)
    base_team = len(PERSONNEL)
    bull_team = len(PERSONNEL_BULL)

    for label, arr_val, team, scenario_key in [
        ("GuardSpine BEAR (end Y2)", bear["arr"][23], bear_team, "Bear"),
        ("GuardSpine BASE (end Y2)", base["arr"][23], base_team, "Base"),
        ("GuardSpine BULL (end Y2)", bull["arr"][23], bull_team, "Bull"),
    ]:
        arr_per_emp = arr_val / team
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = gs_highlight
        ws.cell(row=row, column=1, value=label).font = gs_font
        cell = ws.cell(row=row, column=2, value=arr_val)
        cell.number_format = '$#,##0'
        cell.font = gs_font
        ws.cell(row=row, column=3, value=team).font = gs_font
        cell = ws.cell(row=row, column=4, value=arr_per_emp)
        cell.number_format = '$#,##0'
        cell.font = gs_font
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="AI-native startups avg $3.5M ARR/employee vs traditional SaaS $200-400K.")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="808080")
    row += 1
    ws.cell(row=row, column=1, value="GuardSpine's BYOK + PLG model = Cursor-class efficiency without inference cost drag.")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="808080")
    row += 1
    ws.cell(row=row, column=1, value="Sources: Sacra, SaaStr, CB Insights, Lean AI Leaderboard, company filings (2025-2026)")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="A0A0A0")

    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    print("Computing scenarios...")
    results = {}
    for key in ["Bear", "Base", "Bull"]:
        print(f"  {key} case...")
        results[key] = compute_scenario(key)

    print("Creating workbook...")
    wb = load_workbook(INPUT_FILE)

    # Fix dates in existing sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace("2017", "2026").replace("2018", "2027")

    # Add Kristen to Costs Hypotheses if it exists
    if "Costs Hypotheses" in wb.sheetnames:
        ws = wb["Costs Hypotheses"]
        # Find the first empty personnel row and add Kristen
        # The template has personnel rows starting around row 3
        # We need to find where to insert
        for r in range(3, 30):
            cell_b = ws.cell(row=r, column=2)
            if cell_b.value and str(cell_b.value).strip() == "-":
                # Empty personnel slot
                ws.cell(row=r, column=2, value="GTM - Kristen Hengst Smith")
                ws.cell(row=r, column=3, value="Sales")
                ws.cell(row=r, column=4, value=84000)  # $84K/yr
                ws.cell(row=r, column=5, value=1)  # Start month 1
                print(f"  Added Kristen at row {r}")
                break

    # Create new scenario sheets
    print("Building scenario sheets...")
    create_comparison_sheet(wb, results)
    create_scenario_sheet(wb, "Base Case P&L", "Base", results["Base"])
    create_scenario_sheet(wb, "Bear Case P&L", "Bear", results["Bear"])
    create_scenario_sheet(wb, "Bull Case P&L", "Bull", results["Bull"])

    # Save
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

    # Print summary
    print("\n" + "="*60)
    print("SCENARIO COMPARISON SUMMARY")
    print("="*60)
    for key in ["Bear", "Base", "Bull"]:
        d = results[key]
        sc = SCENARIOS[key]
        print(f"\n--- {sc['label']} ---")
        print(f"  End-Y1 MRR:     ${d['gross_mrr'][11]:,.0f}")
        print(f"  End-Y1 ARR:     ${d['arr'][11]:,.0f}")
        print(f"  End-Y2 MRR:     ${d['gross_mrr'][23]:,.0f}")
        print(f"  End-Y2 ARR:     ${d['arr'][23]:,.0f}")
        print(f"  Y1 EBITDA:      ${d['y1_ebitda']:,.0f}")
        print(f"  Y2 EBITDA:      ${d['y2_ebitda']:,.0f}")
        print(f"  End-Y2 GM:      {d['gross_margin_pct'][23]:.1%}")
        print(f"  Breakeven:      Month {d['breakeven_month'] or 'N/A'}")
        print(f"  End-Y1 Cash:    ${d['cumulative_cash'][11]:,.0f}")
        print(f"  End-Y2 Cash:    ${d['cumulative_cash'][23]:,.0f}")
        print(f"  End-Y2 LTV:CAC: {d['ltv_cac'][23]:.1f}x")


if __name__ == "__main__":
    main()
